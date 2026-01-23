"""
Vistas del módulo de Contratistas - MEJORADO FASE 1
apps/contratistas/views.py
"""
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum, Count
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from decimal import Decimal
from datetime import datetime
import decimal  # Para manejar excepciones de Decimal

from .models import Contratista, ContratoProyecto, AvaluoContratista, PlanillaContratista, DetallePlanillaContratista
from apps.proyectos.models import Proyecto
from .forms import ContratistaForm, ContratoProyectoForm, PagoContratistaForm
from apps.core.utils import get_tipo_cambio_actual

from django.http import JsonResponse
from django.views import View
from django.shortcuts import get_object_or_404,redirect

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.utils import timezone
from django.views.generic import TemplateView


PagoContratista = AvaluoContratista

class ContratistaListView(LoginRequiredMixin, ListView):
    """Vista de lista de contratistas con filtros avanzados"""
    model = Contratista
    template_name = 'contratistas/lista.html'
    context_object_name = 'contratistas'
    paginate_by = 50  # ✅ CAMBIADO: De 20 a 50 items por página
    
    def get_queryset(self):
        queryset = Contratista.objects.filter(eliminado=False).select_related('creado_por')
        
        # Filtro por búsqueda (cédula o nombre)
        buscar = self.request.GET.get('buscar', '').strip()
        if buscar:
            queryset = queryset.filter(
                Q(numero_cedula__icontains=buscar) |
                Q(nombre__icontains=buscar) |
                Q(apellido__icontains=buscar)
            )
        
        # Filtro por proyecto
        proyecto_id = self.request.GET.get('proyecto', '').strip()
        if proyecto_id:
            # Filtrar contratistas que tengan contratos en ese proyecto
            queryset = queryset.filter(
                contratos__proyecto_id=proyecto_id,
                contratos__eliminado=False
            ).distinct()
        
        # Filtro por forma de pago (de sus pagos)
        forma_pago = self.request.GET.get('forma_pago', '').strip()
        if forma_pago:
            queryset = queryset.filter(
                contratos__pagos__forma_pago=forma_pago,
                contratos__pagos__eliminado=False
            ).distinct()
        
        # Filtro por estado de contrato
        estado = self.request.GET.get('estado', '').strip()
        if estado:
            queryset = queryset.filter(
                contratos__estado=estado,
                contratos__eliminado=False
            ).distinct()
        
        # ✅ NUEVO: Filtro por rango de fechas de contrato
        fecha_desde = self.request.GET.get('fecha_desde', '').strip()
        fecha_hasta = self.request.GET.get('fecha_hasta', '').strip()
        
        if fecha_desde:
            try:
                fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                queryset = queryset.filter(
                    contratos__fecha_inicio__gte=fecha_desde_obj,
                    contratos__eliminado=False
                ).distinct()
            except ValueError:
                pass  # Ignorar si la fecha no es válida
        
        if fecha_hasta:
            try:
                fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                queryset = queryset.filter(
                    contratos__fecha_inicio__lte=fecha_hasta_obj,
                    contratos__eliminado=False
                ).distinct()
            except ValueError:
                pass  # Ignorar si la fecha no es válida
        
        return queryset.order_by('apellido', 'nombre')
    
    def get_context_data(self, **kwargs):
        from django.db.models import Sum
        from decimal import Decimal
        
        context = super().get_context_data(**kwargs)
        
        # ==========================================
        # PARÁMETROS DE FILTRADO
        # ==========================================
        context['buscar'] = self.request.GET.get('buscar', '')
        context['proyecto_filtro'] = self.request.GET.get('proyecto', '')
        context['forma_pago_filtro'] = self.request.GET.get('forma_pago', '')
        context['moneda'] = self.request.GET.get('moneda', 'cordobas')
        
        # ==========================================
        # LISTAS PARA FILTROS
        # ==========================================
        context['proyectos'] = Proyecto.objects.filter(eliminado=False).order_by('nombre')
        context['formas_pago'] = AvaluoContratista.FORMA_PAGO_CHOICES
        
        # ==========================================
        # TIPO DE CAMBIO
        # ==========================================
        tipo_cambio = get_tipo_cambio_actual()
        context['tipo_cambio'] = tipo_cambio
        context['tipo_cambio_global'] = tipo_cambio
        
        moneda = context['moneda']
        
        # ==========================================
        # ✅ ESTADÍSTICAS GLOBALES (PARA LAS TARJETAS)
        # ==========================================
        
        # Total de contratistas activos
        contratistas_activos = Contratista.objects.filter(eliminado=False, activo=True)
        
        # Total de contratos activos
        total_contratos = ContratoProyecto.objects.filter(eliminado=False).count()
        
        # ✅ TOTAL PAGADO: Desde planillas pagadas
        total_pagado_stats = DetallePlanillaContratista.objects.filter(
            planilla__estado='pagada'
        ).aggregate(
            total_cordobas=Sum('monto_cordobas'),
            total_dolares=Sum('monto_dolares'),
            cantidad_pagos=Count('id')
        )
        
        total_pagado_cordobas = total_pagado_stats['total_cordobas'] or Decimal('0.00')
        total_pagado_dolares = total_pagado_stats['total_dolares'] or Decimal('0.00')
        cantidad_pagos_realizados = total_pagado_stats['cantidad_pagos'] or 0
        
        # Total valor de contratos
        valor_contratos = ContratoProyecto.objects.filter(
            eliminado=False
        ).aggregate(
            total=Sum('valor_contrato')
        )['total'] or Decimal('0.00')
        
        pendiente_cordobas = valor_contratos - total_pagado_cordobas
        pendiente_dolares = pendiente_cordobas / tipo_cambio if tipo_cambio > 0 else Decimal('0.00')
        
        # Porcentaje de avance general
        porcentaje_avance = (total_pagado_cordobas / valor_contratos * 100) if valor_contratos > 0 else 0
        
        # ✅ AGREGAR AL CONTEXTO
        context['stats'] = {
            'total_contratistas': contratistas_activos.count(),
            'total_contratos': total_contratos,
            'total_pagos': cantidad_pagos_realizados,
            'total_pagado_cordobas': total_pagado_cordobas,
            'total_pagado_dolares': total_pagado_dolares,
            'valor_total_contratos': valor_contratos,
            'pendiente_cordobas': pendiente_cordobas,
            'pendiente_dolares': pendiente_dolares,
            'porcentaje_avance': porcentaje_avance,
        }
        
        # ==========================================
        # CALCULAR DATOS FINANCIEROS PARA CADA CONTRATISTA
        # ==========================================
        for contratista in context['contratistas']:
            # Contratos del contratista
            contratos = ContratoProyecto.objects.filter(
                contratista=contratista,
                eliminado=False
            )
            
            # Aplicar filtro de proyecto si existe
            if context['proyecto_filtro']:
                contratos = contratos.filter(proyecto_id=context['proyecto_filtro'])
            
            # Valor total de contratos
            valor_contratos_contratista = sum([c.valor_contrato for c in contratos])
            
            # ✅ TOTAL PAGADO: Suma de avalúos en planillas PAGADAS
            detalles_pagados = DetallePlanillaContratista.objects.filter(
                avaluo__contrato__contratista=contratista,
                avaluo__contrato__eliminado=False,
                planilla__estado='pagada'
            )
            
            # Aplicar filtro de proyecto si existe
            if context['proyecto_filtro']:
                detalles_pagados = detalles_pagados.filter(
                    avaluo__contrato__proyecto_id=context['proyecto_filtro']
                )
            
            total_pagado = detalles_pagados.aggregate(
                total=Sum('monto_cordobas')
            )['total'] or Decimal('0.00')
            
            # Pendiente = Valor contratos - Total pagado
            pendiente = valor_contratos_contratista - total_pagado
            
            # Avance = (Total pagado / Valor contratos) * 100
            avance = (total_pagado / valor_contratos_contratista * 100) if valor_contratos_contratista > 0 else 0
            
            # ✅ CANTIDAD DE PAGOS REALIZADOS (avalúos en planillas pagadas)
            cantidad_pagos_query = DetallePlanillaContratista.objects.filter(
                avaluo__contrato__contratista=contratista,
                avaluo__contrato__eliminado=False,
                planilla__estado='pagada'
            )
            
            # Aplicar filtro de proyecto si existe
            if context['proyecto_filtro']:
                cantidad_pagos_query = cantidad_pagos_query.filter(
                    avaluo__contrato__proyecto_id=context['proyecto_filtro']
                )
            
            cantidad_pagos = cantidad_pagos_query.count()
            
            # ✅ AGREGAR DATOS AL OBJETO CONTRATISTA
            contratista.total_contratos_count = contratos.count()
            contratista.valor_contratos = valor_contratos_contratista
            contratista.total_pagado_valor = total_pagado
            contratista.pendiente = pendiente
            contratista.avance = avance
            contratista.cantidad_pagos = cantidad_pagos
            
            # Convertir a dólares si es necesario
            if moneda == 'dolares':
                contratista.valor_contratos_display = valor_contratos_contratista / tipo_cambio if tipo_cambio > 0 else Decimal('0.00')
                contratista.total_pagado_display = total_pagado / tipo_cambio if tipo_cambio > 0 else Decimal('0.00')
                contratista.pendiente_display = pendiente / tipo_cambio if tipo_cambio > 0 else Decimal('0.00')
            else:
                contratista.valor_contratos_display = valor_contratos_contratista
                contratista.total_pagado_display = total_pagado
                contratista.pendiente_display = pendiente
        
        return context

class ContratistaCreateView(LoginRequiredMixin, CreateView):
    """Vista para crear un nuevo contratista"""
    model = Contratista
    form_class = ContratistaForm
    template_name = 'contratistas/crear.html'
    success_url = reverse_lazy('contratistas_lista')
    
    def form_valid(self, form):
        # Asignar el usuario que crea el contratista
        form.instance.creado_por = self.request.user
        
        messages.success(
            self.request,
            f'✅ Contratista {form.instance.nombre_completo} creado exitosamente.'
        )
        
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(
            self.request,
            '❌ Error al crear el contratista. Por favor revisa los campos.'
        )
        return super().form_invalid(form)


class ContratistaUpdateView(LoginRequiredMixin, UpdateView):
    """Vista para editar un contratista existente"""
    model = Contratista
    form_class = ContratistaForm
    template_name = 'contratistas/editar.html'
    success_url = reverse_lazy('contratistas_lista')
    
    def get_queryset(self):
        # Solo contratistas no eliminados
        return Contratista.objects.filter(eliminado=False)
    
    def form_valid(self, form):
        messages.success(
            self.request,
            f'✅ Contratista {form.instance.nombre_completo} actualizado exitosamente.'
        )
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(
            self.request,
            '❌ Error al actualizar el contratista. Por favor revisa los campos.'
        )
        return super().form_invalid(form)
    

class ContratistaDetalleAPIView(LoginRequiredMixin, View):
    """
    API para obtener el detalle completo de un contratista
    Retorna JSON con toda la información
    """
    
    def get(self, request, pk):
        try:
            # Obtener contratista
            contratista = get_object_or_404(Contratista, pk=pk, eliminado=False)
            
            # Obtener tipo de cambio
            tipo_cambio = get_tipo_cambio_actual()
            
            # Datos básicos del contratista
            data = {
                'id': contratista.id,
                'nombre_completo': contratista.nombre_completo,
                'nombre': contratista.nombre,
                'apellido': contratista.apellido,
                'numero_cedula': contratista.numero_cedula,
                'telefono': contratista.telefono or '-',
                'direccion': contratista.direccion or '-',
                'departamento': contratista.departamento or '-',
                'municipio': contratista.municipio or '-',
                'foto_cedula_url': contratista.foto_cedula.url if contratista.foto_cedula else None,
                
                # Datos bancarios
                'banco': contratista.banco or '-',
                'numero_cuenta': contratista.numero_cuenta or '-',
                'tipo_cuenta': contratista.get_tipo_cuenta_display() if contratista.tipo_cuenta else '-',
                'moneda_cuenta': contratista.get_moneda_cuenta_display() if contratista.moneda_cuenta else '-',
                
                # Metadata
                'activo': contratista.activo,
                'fecha_creacion': contratista.creado_en.strftime('%d/%m/%Y %H:%M'),
                
                # Contratos
                'contratos': [],
                
                # Resumen financiero
                'total_contratos': 0,
                'valor_total_contratos': Decimal('0.00'),
                'total_pagado': Decimal('0.00'),
                'total_pendiente': Decimal('0.00'),
                'porcentaje_avance': 0,
                'cantidad_pagos': 0,
            }
            
            # Obtener contratos del contratista
            contratos = ContratoProyecto.objects.filter(
                contratista=contratista,
                eliminado=False
            ).select_related('proyecto').order_by('-fecha_inicio')
            
            data['total_contratos'] = contratos.count()
            
            contratos_list = []
            total_valor = Decimal('0.00')
            total_pagado = Decimal('0.00')
            
            for contrato in contratos:
                # Pagos del contrato
                pagos = PagoContratista.objects.filter(
                    contrato=contrato,
                    eliminado=False,
                    estado='aprobado'
                ).order_by('-fecha_pago')
                
                pagado_contrato = pagos.aggregate(
                    total=Sum('monto_cordobas')
                )['total'] or Decimal('0.00')
                
                pendiente_contrato = contrato.valor_contrato - pagado_contrato
                avance_contrato = (pagado_contrato / contrato.valor_contrato * 100) if contrato.valor_contrato > 0 else 0
                
                # Lista de pagos del contrato
                pagos_list = []
                for pago in pagos:
                    pagos_list.append({
                        'id': pago.id,
                        'codigo': pago.codigo,
                        'fecha': pago.fecha_pago.strftime('%d/%m/%Y'),
                        'concepto': pago.concepto,
                        'monto_cordobas': float(pago.monto_cordobas),
                        'monto_dolares': float(pago.monto_dolares),
                        'forma_pago': pago.get_forma_pago_display(),
                        'estado': pago.get_estado_display(),
                        'archivo_soporte': pago.archivo_soporte.url if pago.archivo_soporte else None,
                    })
                
                contratos_list.append({
                    'id': contrato.id,
                    'codigo': contrato.codigo,
                    'proyecto': {
                        'id': contrato.proyecto.id,
                        'nombre': contrato.proyecto.nombre,
                        'descripcion': contrato.proyecto.descripcion or '-',
                    },
                    'actividades': contrato.actividades,
                    'unidad_medida': contrato.unidad_medida or '-',
                    'valor_contrato': float(contrato.valor_contrato),
                    'fecha_inicio': contrato.fecha_inicio.strftime('%d/%m/%Y'),
                    'fecha_fin': contrato.fecha_fin.strftime('%d/%m/%Y') if contrato.fecha_fin else '-',
                    'estado': contrato.get_estado_display(),
                    'descripcion': contrato.descripcion or '-',
                    'total_pagado': float(pagado_contrato),
                    'pendiente': float(pendiente_contrato),
                    'porcentaje_avance': round(avance_contrato, 2),
                    'cantidad_pagos': pagos.count(),
                    'pagos': pagos_list,
                })
                
                total_valor += contrato.valor_contrato
                total_pagado += pagado_contrato
            
            data['contratos'] = contratos_list
            data['valor_total_contratos'] = float(total_valor)
            data['total_pagado'] = float(total_pagado)
            data['total_pendiente'] = float(total_valor - total_pagado)
            data['porcentaje_avance'] = round((total_pagado / total_valor * 100) if total_valor > 0 else 0, 2)
            
            # Cantidad total de pagos
            data['cantidad_pagos'] = sum(c['cantidad_pagos'] for c in contratos_list)
            
            # Tipo de cambio para conversiones
            data['tipo_cambio'] = float(tipo_cambio)
            
            return JsonResponse(data)
            
        except Exception as e:
            return JsonResponse({
                'error': str(e)
            }, status=500)


class ContratistaEstadoCuentaAPIView(LoginRequiredMixin, View):
    """
    API para obtener el estado de cuenta de un contratista
    Muestra movimientos cronológicos con saldo acumulado
    """
    
    def get(self, request, pk):
        try:
            contratista = get_object_or_404(Contratista, pk=pk, eliminado=False)
            tipo_cambio = get_tipo_cambio_actual()
            
            # Obtener todos los contratos y pagos
            contratos = ContratoProyecto.objects.filter(
                contratista=contratista,
                eliminado=False
            ).select_related('proyecto').order_by('fecha_inicio')
            
            movimientos = []
            saldo_acumulado = Decimal('0.00')
            
            for contrato in contratos:
                # Agregar contrato como movimiento inicial
                saldo_acumulado += contrato.valor_contrato
                
                movimientos.append({
                    'fecha': contrato.fecha_inicio.strftime('%d/%m/%Y'),
                    'tipo': 'contrato',
                    'descripcion': f'Contrato {contrato.codigo} - {contrato.proyecto.nombre}',
                    'proyecto': contrato.proyecto.nombre,
                    'debe': float(contrato.valor_contrato),
                    'haber': 0,
                    'saldo': float(saldo_acumulado),
                })
                
                # Agregar pagos del contrato
                pagos = PagoContratista.objects.filter(
                    contrato=contrato,
                    eliminado=False,
                    estado='aprobado'
                ).order_by('fecha_pago')
                
                for pago in pagos:
                    saldo_acumulado -= pago.monto_cordobas
                    
                    movimientos.append({
                        'fecha': pago.fecha_pago.strftime('%d/%m/%Y'),
                        'tipo': 'pago',
                        'descripcion': f'Pago {pago.codigo} - {pago.concepto}',
                        'proyecto': contrato.proyecto.nombre,
                        'debe': 0,
                        'haber': float(pago.monto_cordobas),
                        'saldo': float(saldo_acumulado),
                        'forma_pago': pago.get_forma_pago_display(),
                    })
            
            # Ordenar por fecha
            movimientos.sort(key=lambda x: x['fecha'])
            
            # Calcular totales
            total_debe = sum(m['debe'] for m in movimientos)
            total_haber = sum(m['haber'] for m in movimientos)
            
            data = {
                'contratista': {
                    'nombre': contratista.nombre_completo,
                    'cedula': contratista.numero_cedula,
                },
                'movimientos': movimientos,
                'totales': {
                    'debe': total_debe,
                    'haber': total_haber,
                    'saldo': total_debe - total_haber,
                },
                'tipo_cambio': float(tipo_cambio),
            }
            
            return JsonResponse(data)
            
        except Exception as e:
            return JsonResponse({
                'error': str(e)
            }, status=500)

class ContratistaEstadoCuentaView(LoginRequiredMixin, TemplateView):
    """
    Vista de Estado de Cuenta del Contratista (Página Completa)
    Muestra historial financiero detallado con gráficos y exportación
    """
    template_name = 'contratistas/estado_cuenta.html'
    
    def get_context_data(self, **kwargs):
        from django.db.models import Sum
        from decimal import Decimal
        
        context = super().get_context_data(**kwargs)
        
        # Obtener contratista
        contratista_id = self.kwargs.get('pk')
        contratista = get_object_or_404(Contratista, pk=contratista_id, eliminado=False)
        
        context['contratista'] = contratista
        context['tipo_cambio'] = get_tipo_cambio_actual()
        
        # ==========================================
        # RESUMEN FINANCIERO GLOBAL
        # ==========================================
        
        contratos = ContratoProyecto.objects.filter(
            contratista=contratista,
            eliminado=False
        )
        
        # Valor total de contratos
        valor_total_contratos = sum([c.valor_contrato for c in contratos])
        
        # Total pagado (desde planillas pagadas)
        detalles_pagados = DetallePlanillaContratista.objects.filter(
            avaluo__contrato__contratista=contratista,
            avaluo__contrato__eliminado=False,
            planilla__estado='pagada'
        )
        
        total_pagado = detalles_pagados.aggregate(
            total=Sum('monto_cordobas')
        )['total'] or Decimal('0.00')
        
        saldo_pendiente = valor_total_contratos - total_pagado
        porcentaje_avance = (total_pagado / valor_total_contratos * 100) if valor_total_contratos > 0 else 0
        cantidad_pagos = detalles_pagados.count()
        
        context['resumen'] = {
            'valor_total_contratos': valor_total_contratos,
            'total_pagado': total_pagado,
            'saldo_pendiente': saldo_pendiente,
            'porcentaje_avance': porcentaje_avance,
            'cantidad_pagos': cantidad_pagos,
            'cantidad_contratos': contratos.count(),
        }
        
        # ==========================================
        # CONTRATOS CON DETALLES
        # ==========================================
        
        contratos_detalle = []
        
        for contrato in contratos.select_related('proyecto'):
            # Pagos de este contrato
            detalles_contrato = DetallePlanillaContratista.objects.filter(
                avaluo__contrato=contrato,
                planilla__estado='pagada'
            ).select_related('planilla', 'avaluo')
            
            pagado_contrato = detalles_contrato.aggregate(
                total=Sum('monto_cordobas')
            )['total'] or Decimal('0.00')
            
            pendiente_contrato = contrato.valor_contrato - pagado_contrato
            avance_contrato = (pagado_contrato / contrato.valor_contrato * 100) if contrato.valor_contrato > 0 else 0
            
            contratos_detalle.append({
                'contrato': contrato,
                'pagado': pagado_contrato,
                'pendiente': pendiente_contrato,
                'avance': avance_contrato,
                'cantidad_pagos': detalles_contrato.count(),
                'pagos': detalles_contrato.order_by('planilla__fecha_generacion')
            })
        
        context['contratos_detalle'] = contratos_detalle
        
        # ==========================================
        # HISTORIAL DE PAGOS CRONOLÓGICO
        # ==========================================
        
        historial_pagos = DetallePlanillaContratista.objects.filter(
            avaluo__contrato__contratista=contratista,
            avaluo__contrato__eliminado=False,
            planilla__estado='pagada'
        ).select_related(
            'planilla',
            'avaluo__contrato__proyecto',
            'avaluo__contrato'
        ).order_by('planilla__fecha_generacion')
        
        # Calcular saldo acumulado
        saldo_acumulado = Decimal('0.00')
        historial_con_saldo = []
        
        for detalle in historial_pagos:
            saldo_acumulado += detalle.monto_cordobas
            historial_con_saldo.append({
                'detalle': detalle,
                'saldo_acumulado': saldo_acumulado
            })
        
        context['historial_pagos'] = historial_con_saldo
        
        # ==========================================
        # RESUMEN POR PROYECTO
        # ==========================================
        
        from django.db.models import Count
        from apps.proyectos.models import Proyecto
        
        # Usar diccionario para evitar duplicados
        proyectos_dict = {}
        
        for contrato in contratos:
            proyecto_id = contrato.proyecto.id
            
            # Si el proyecto ya está en el diccionario, skip
            if proyecto_id in proyectos_dict:
                continue
            
            # Obtener todos los contratos de este proyecto
            contratos_proyecto = contratos.filter(proyecto=contrato.proyecto)
            
            # Total pagado en este proyecto
            pagado_proyecto = DetallePlanillaContratista.objects.filter(
                avaluo__contrato__contratista=contratista,
                avaluo__contrato__proyecto=contrato.proyecto,
                planilla__estado='pagada'
            ).aggregate(total=Sum('monto_cordobas'))['total'] or Decimal('0.00')
            
            valor_contratos_proyecto = sum([c.valor_contrato for c in contratos_proyecto])
            pendiente_proyecto = valor_contratos_proyecto - pagado_proyecto
            
            proyectos_dict[proyecto_id] = {
                'proyecto': contrato.proyecto,
                'cantidad_contratos': contratos_proyecto.count(),
                'valor_contratos': valor_contratos_proyecto,
                'pagado': pagado_proyecto,
                'pendiente': pendiente_proyecto,
            }
        
        # Convertir diccionario a lista
        resumen_proyectos = list(proyectos_dict.values())
        
        context['resumen_proyectos'] = resumen_proyectos
        
        return context


class ContratoCreateView(LoginRequiredMixin, CreateView):
    model = ContratoProyecto
    form_class = ContratoProyectoForm
    template_name = 'contratistas/contrato_form.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.proyecto = get_object_or_404(Proyecto, pk=self.kwargs['proyecto_id'], eliminado=False)
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        #kwargs['proyecto'] = self.proyecto
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proyecto'] = self.proyecto
        context['titulo'] = 'Crear Contrato'
        context['boton'] = 'Crear Contrato'
        return context
    
    def form_valid(self, form):
        form.instance.proyecto = self.proyecto
        form.instance.creado_por = self.request.user
        messages.success(self.request, f'✅ Contrato creado exitosamente')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # ✅ AGREGAR ESTA FUNCIÓN
        print("=" * 50)
        print("ERRORES EN EL FORMULARIO:")
        print(form.errors)
        print("=" * 50)
        return super().form_invalid(form)
    
    def get_success_url(self):
        return reverse('proyecto_detalle', kwargs={'pk': self.proyecto.id})

class ContratoUpdateView(LoginRequiredMixin, UpdateView):
    model = ContratoProyecto
    form_class = ContratoProyectoForm
    template_name = 'contratistas/contrato_form.html'
    
    def get_queryset(self):
        return ContratoProyecto.objects.filter(eliminado=False)
    
    def dispatch(self, request, *args, **kwargs):
        self.proyecto = self.get_object().proyecto
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        #kwargs['proyecto'] = self.proyecto
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proyecto'] = self.proyecto
        context['titulo'] = 'Editar Contrato'
        context['boton'] = 'Guardar Cambios'
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        # Guardar el proyecto actualizado para la redirección
        self.proyecto_destino = self.object.proyecto
        return response
        
    def get_success_url(self):
        #return reverse('proyecto_detalle', kwargs={'pk': self.proyecto.id})
        return reverse('proyecto_detalle', kwargs={'pk': self.object.proyecto.id})


class ContratoDeleteView(LoginRequiredMixin, DeleteView):
    model = ContratoProyecto
    
    def post(self, request, *args, **kwargs):
        contrato = self.get_object()
        proyecto_id = contrato.proyecto.id
        contrato.eliminado = True
        contrato.save()
        messages.success(request, f'✅ Contrato eliminado')
        return redirect('proyecto_detalle', pk=proyecto_id)

class ContratistaDetalleView(DetailView):
    """Vista de detalle del contratista con toda su información"""
    model = Contratista
    template_name = 'contratistas/contratista_detalle.html'
    context_object_name = 'contratista'
    
    def get_queryset(self):
        return Contratista.objects.filter(eliminado=False)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contratista = self.object
        
        # Obtener todos los contratos del contratista
        contratos = ContratoProyecto.objects.filter(
            contratista=contratista,
            eliminado=False
        ).select_related('proyecto').prefetch_related('avaluos')
        
        # Obtener todos los avalúos ordenados por fecha
        todos_avaluos = AvaluoContratista.objects.filter(
            contrato__contratista=contratista,
            contrato__eliminado=False,
            eliminado=False
        ).select_related('contrato', 'contrato__proyecto').order_by('-fecha_pago')
        
        context['contratos'] = contratos
        context['todos_avaluos'] = todos_avaluos
        context['total_avaluos'] = todos_avaluos.count()
        
        return context
    
class PagoContratistaCreateView(LoginRequiredMixin, CreateView):
    """Vista para crear un nuevo pago a contratista"""
    model = AvaluoContratista
    form_class = PagoContratistaForm
    template_name = 'contratistas/pago_form.html'
    
    def dispatch(self, request, *args, **kwargs):
        # Obtener el contrato
        self.contrato = get_object_or_404(
            ContratoProyecto,
            pk=self.kwargs['contrato_id'],
            eliminado=False
        )
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['contrato'] = self.contrato
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contrato'] = self.contrato
        context['contratista'] = self.contrato.contratista
        context['proyecto'] = self.contrato.proyecto
        context['titulo'] = 'Registrar Pago'
        context['boton'] = 'Registrar Pago'
        
        contrato = context['contrato']
        if contrato:
            context['contrato'] = contrato
            
            # Calcular total avaluado hasta ahora
            from django.db.models import Sum
            total = contrato.avaluos.filter(eliminado=False).aggregate(
                total=Sum('monto_cordobas')
            )['total'] or 0
            
            context['total_avaluado'] = total
        # Información financiera del contrato
        context['info_financiera'] = {
            'valor_contrato': self.contrato.valor_contrato,
            'total_pagado': self.contrato.total_pagado,
            'pendiente': self.contrato.total_pendiente,
            'porcentaje_avance': self.contrato.porcentaje_avance,
            'cantidad_pagos': self.contrato.cantidad_avaluos,
        }
        
        return context
    
    def form_valid(self, form):
        # Asignar el contrato y el usuario que registra
        form.instance.contrato = self.contrato
        form.instance.ingresado_por = self.request.user
        form.instance.estado = 'pendiente'
        
        messages.success(
            self.request,
            f'✅ Pago registrado exitosamente. Pendiente de aprobación.'
        )
        
        return super().form_valid(form)
    
    def get_success_url(self):
        # Redirigir al detalle del contrato
        return reverse('proyecto_detalle', kwargs={'pk': self.contrato.proyecto.id})

class PagoContratistaUpdateView(LoginRequiredMixin, UpdateView):
    """Vista para editar un pago (solo si está pendiente)"""
    model = PagoContratista
    form_class = PagoContratistaForm
    template_name = 'contratistas/pago_form.html'
    
    def get_queryset(self):
        # Solo pagos pendientes pueden editarse
        return PagoContratista.objects.filter(
            eliminado=False,
            estado='pendiente'
        )
    
    def dispatch(self, request, *args, **kwargs):
        pago = self.get_object()
        self.contrato = pago.contrato
        
        # Validar que el pago está en estado pendiente
        if pago.estado != 'pendiente':
            messages.error(
                request,
                '❌ Solo se pueden editar pagos en estado pendiente.'
            )
            return redirect('proyecto_detalle', pk=self.contrato.proyecto.id)
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['contrato'] = self.contrato
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contrato'] = self.contrato
        context['contratista'] = self.contrato.contratista
        context['proyecto'] = self.contrato.proyecto
        context['titulo'] = 'Editar Pago'
        context['boton'] = 'Guardar Cambios'
        
        # Información financiera del contrato
        context['info_financiera'] = {
            'valor_contrato': self.contrato.valor_contrato,
            'total_pagado': self.contrato.total_pagado,
            'pendiente': self.contrato.total_pendiente,
            'porcentaje_avance': self.contrato.porcentaje_avance,
            'cantidad_pagos': self.contrato.cantidad_pagos,
        }
        
        return context
    
    def form_valid(self, form):
        messages.success(
            self.request,
            f'✅ Pago actualizado exitosamente.'
        )
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('proyecto_detalle', kwargs={'pk': self.contrato.proyecto.id})

class PagoContratistaDeleteView(LoginRequiredMixin, DeleteView):
    """Vista para eliminar un pago (soft delete)"""
    model = PagoContratista
    
    def get_queryset(self):
        # Solo pagos pendientes pueden eliminarse
        return PagoContratista.objects.filter(
            eliminado=False,
            estado='pendiente'
        )
    
    def post(self, request, *args, **kwargs):
        pago = self.get_object()
        
        if pago.estado != 'pendiente':
            messages.error(
                request,
                '❌ Solo se pueden eliminar pagos en estado pendiente.'
            )
        else:
            proyecto_id = pago.contrato.proyecto.id
            pago.eliminado = True
            pago.save()
            messages.success(request, f'✅ Pago eliminado exitosamente')
            return redirect('proyecto_detalle', pk=proyecto_id)
        
        return redirect('proyecto_detalle', pk=pago.contrato.proyecto.id)

class PagoContratistaDetalleView(LoginRequiredMixin, DetailView):
    """Vista de detalle del pago con timeline de aprobaciones"""
    model = PagoContratista
    template_name = 'contratistas/pago_detalle.html'
    context_object_name = 'pago'
    
    def get_queryset(self):
        return PagoContratista.objects.filter(eliminado=False).select_related(
            'contrato__contratista',
            'contrato__proyecto',
            'ingresado_por',
            'aprobado_gerente_por',
            'aprobado_contador_por'
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pago = self.object
        
        context['contrato'] = pago.contrato
        context['contratista'] = pago.contrato.contratista
        context['proyecto'] = pago.contrato.proyecto
        
        # Timeline de aprobaciones
        timeline = []
        
        # 1. Ingresado
        timeline.append({
            'estado': 'ingresado',
            'titulo': 'Pago Registrado',
            'usuario': pago.ingresado_por,
            'fecha': pago.fecha_ingreso,
            'icono': 'add_circle',
            'color': 'blue',
            'completado': True
        })
        
        # 2. Aprobado Gerente
        if pago.aprobado_gerente_por:
            timeline.append({
                'estado': 'aprobado_gerente',
                'titulo': 'Aprobado por Gerente',
                'usuario': pago.aprobado_gerente_por,
                'fecha': pago.fecha_aprobacion_gerente,
                'icono': 'check_circle',
                'color': 'purple',
                'completado': True
            })
        else:
            timeline.append({
                'estado': 'aprobado_gerente',
                'titulo': 'Pendiente Aprobación Gerente',
                'usuario': None,
                'fecha': None,
                'icono': 'schedule',
                'color': 'gray',
                'completado': False
            })
        
        # 3. Aprobado Contador
        if pago.aprobado_contador_por:
            timeline.append({
                'estado': 'aprobado_contador',
                'titulo': 'Aprobado por Contador',
                'usuario': pago.aprobado_contador_por,
                'fecha': pago.fecha_aprobacion_contador,
                'icono': 'verified',
                'color': 'green',
                'completado': True
            })
        else:
            timeline.append({
                'estado': 'aprobado_contador',
                'titulo': 'Pendiente Aprobación Contador',
                'usuario': None,
                'fecha': None,
                'icono': 'pending',
                'color': 'gray',
                'completado': False
            })
        
        context['timeline'] = timeline
        
        # Verificar permisos del usuario actual
        user = self.request.user
        context['puede_aprobar_gerente'] = (
            pago.estado == 'pendiente' and 
            (user.rol in ['administrador', 'gerente'])
        )
        context['puede_aprobar_contador'] = (
            pago.estado == 'aprobado_gerente' and 
            (user.rol in ['administrador', 'contador'])
        )
        context['puede_rechazar'] = (
            pago.estado in ['pendiente', 'aprobado_gerente'] and
            (user.rol in ['administrador', 'gerente', 'contador'])
        )
        
        return context


class PagoAprobarGerenteView(LoginRequiredMixin, View):
    """Vista para aprobar pago como gerente"""
    
    def post(self, request, *args, **kwargs):
        pago = get_object_or_404(
            PagoContratista,
            pk=kwargs['pk'],
            eliminado=False
        )
        
        # Verificar permisos
        if request.user.rol not in ['administrador', 'gerente']:
            messages.error(
                request,
                '❌ No tienes permisos para aprobar pagos como gerente.'
            )
            return redirect('pago_detalle', pk=pago.id)
        
        # Verificar estado
        if pago.estado != 'pendiente':
            messages.error(
                request,
                '❌ Solo se pueden aprobar pagos en estado PENDIENTE.'
            )
            return redirect('pago_detalle', pk=pago.id)
        
        try:
            # Aprobar
            pago.aprobar_gerente(request.user)
            
            messages.success(
                request,
                f'✅ Pago {pago.codigo} aprobado como GERENTE. '
                f'Ahora debe ser aprobado por el CONTADOR.'
            )
        except Exception as e:
            messages.error(
                request,
                f'❌ Error al aprobar el pago: {str(e)}'
            )
        
        return redirect('pago_detalle', pk=pago.id)


class PagoAprobarContadorView(LoginRequiredMixin, View):
    """Vista para aprobar pago como contador (aprobación final)"""
    
    def post(self, request, *args, **kwargs):
        pago = get_object_or_404(
            PagoContratista,
            pk=kwargs['pk'],
            eliminado=False
        )
        
        # Verificar permisos
        if request.user.rol not in ['administrador', 'contador']:
            messages.error(
                request,
                '❌ No tienes permisos para aprobar pagos como contador.'
            )
            return redirect('pago_detalle', pk=pago.id)
        
        # Verificar estado (puede ser pendiente O aprobado_gerente)
        if pago.estado not in ['pendiente', 'aprobado_gerente']:
            messages.error(
                request,
                '❌ Este pago no puede ser aprobado por el contador en su estado actual.'
            )
            return redirect('pago_detalle', pk=pago.id)
        
        try:
            # Aprobar
            pago.aprobar_contador(request.user)
            
            messages.success(
                request,
                f'✅ Pago {pago.codigo} APROBADO FINAL. '
                f'El pago ha sido registrado y suma al total del contrato.'
            )
        except Exception as e:
            messages.error(
                request,
                f'❌ Error al aprobar el pago: {str(e)}'
            )
        
        return redirect('pago_detalle', pk=pago.id)


class PagoRechazarView(LoginRequiredMixin, View):
    """Vista para rechazar un pago"""
    
    def post(self, request, *args, **kwargs):
        pago = get_object_or_404(
            PagoContratista,
            pk=kwargs['pk'],
            eliminado=False
        )
        
        # Verificar permisos
        if request.user.rol not in ['administrador', 'gerente', 'contador']:
            messages.error(
                request,
                '❌ No tienes permisos para rechazar pagos.'
            )
            return redirect('pago_detalle', pk=pago.id)
        
        # Verificar estado
        if pago.estado not in ['pendiente', 'aprobado_gerente']:
            messages.error(
                request,
                '❌ Solo se pueden rechazar pagos pendientes o aprobados por gerente.'
            )
            return redirect('pago_detalle', pk=pago.id)
        
        # Obtener motivo del rechazo
        motivo = request.POST.get('motivo', '').strip()
        if not motivo:
            messages.error(
                request,
                '❌ Debes proporcionar un motivo para rechazar el pago.'
            )
            return redirect('pago_detalle', pk=pago.id)
        
        try:
            # Rechazar
            pago.rechazar(request.user, motivo)
            
            messages.warning(
                request,
                f'⚠️ Pago {pago.codigo} RECHAZADO. Motivo: {motivo}'
            )
        except Exception as e:
            messages.error(
                request,
                f'❌ Error al rechazar el pago: {str(e)}'
            )
        
        return redirect('pago_detalle', pk=pago.id)


class PagosPendientesListView(LoginRequiredMixin, ListView):
    """Vista de lista de pagos pendientes de aprobación"""
    model = PagoContratista
    template_name = 'contratistas/pagos_pendientes.html'
    context_object_name = 'pagos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = PagoContratista.objects.filter(
            eliminado=False
        ).exclude(
            estado='aprobado'
        ).select_related(
            'contrato__contratista',
            'contrato__proyecto',
            'ingresado_por'
        ).order_by('-fecha_ingreso')
        
        # Filtro por estado
        estado = self.request.GET.get('estado', '').strip()
        if estado:
            queryset = queryset.filter(estado=estado)
        
        # Filtro por contratista
        contratista_id = self.request.GET.get('contratista', '').strip()
        if contratista_id:
            queryset = queryset.filter(contrato__contratista_id=contratista_id)
        
        # Filtro por proyecto
        proyecto_id = self.request.GET.get('proyecto', '').strip()
        if proyecto_id:
            queryset = queryset.filter(contrato__proyecto_id=proyecto_id)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estadísticas
        total_pendientes = PagoContratista.objects.filter(
            eliminado=False,
            estado='pendiente'
        ).count()
        
        total_aprobados_gerente = PagoContratista.objects.filter(
            eliminado=False,
            estado='aprobado_gerente'
        ).count()
        
        total_rechazados = PagoContratista.objects.filter(
            eliminado=False,
            estado='rechazado'
        ).count()
        
        context['stats'] = {
            'total_pendientes': total_pendientes,
            'total_aprobados_gerente': total_aprobados_gerente,
            'total_rechazados': total_rechazados,
        }
        
        # Listas para filtros
        context['contratistas'] = Contratista.objects.filter(eliminado=False).order_by('apellido', 'nombre')
        context['proyectos'] = Proyecto.objects.filter(eliminado=False).order_by('nombre')
        context['estados'] = PagoContratista.ESTADO_CHOICES
        
        # Parámetros actuales
        context['estado_seleccionado'] = self.request.GET.get('estado', '')
        context['contratista_seleccionado'] = self.request.GET.get('contratista', '')
        context['proyecto_seleccionado'] = self.request.GET.get('proyecto', '')
        
        return context

# ===========================================
# VISTAS DE PLANILLAS DE CONTRATISTAS
# ===========================================

class PlanillaListView(LoginRequiredMixin, ListView):
    """Lista de planillas de contratistas"""
    model = PlanillaContratista
    template_name = 'contratistas/planilla_lista.html'
    context_object_name = 'planillas'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = PlanillaContratista.objects.filter(
            eliminado=False
        ).select_related('proyecto', 'generada_por').order_by('-fecha_generacion')
        
        # Filtros
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        
        proyecto_id = self.request.GET.get('proyecto')
        if proyecto_id:
            queryset = queryset.filter(proyecto_id=proyecto_id)
        
        fecha_desde = self.request.GET.get('fecha_desde')
        if fecha_desde:
            queryset = queryset.filter(periodo_inicio__gte=fecha_desde)
        
        fecha_hasta = self.request.GET.get('fecha_hasta')
        if fecha_hasta:
            queryset = queryset.filter(periodo_fin__lte=fecha_hasta)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proyectos'] = Proyecto.objects.filter(eliminado=False)
        context['estados'] = PlanillaContratista.ESTADO_CHOICES
        
        # Estadísticas
        queryset = self.get_queryset()
        context['total_planillas'] = queryset.count()
        context['total_cordobas'] = queryset.aggregate(
            total=Sum('total_cordobas')
        )['total'] or Decimal('0')
        context['total_dolares'] = queryset.aggregate(
            total=Sum('total_dolares')
        )['total'] or Decimal('0')
        
        return context

class PlanillaCreateView(LoginRequiredMixin, CreateView):
    """Crear nueva planilla de contratistas"""
    model = PlanillaContratista
    template_name = 'contratistas/planilla_crear.html'
    fields = ['proyecto', 'periodo_inicio', 'periodo_fin']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Proyectos
        context['proyectos'] = Proyecto.objects.filter(eliminado=False).order_by('nombre')
        
        # Filtros
        proyecto_id = self.request.GET.get('proyecto')
        contratista_id = self.request.GET.get('contratista')
        
        # Solo mostrar avalúos si hay proyecto seleccionado
        if proyecto_id:
            avaluos = AvaluoContratista.objects.filter(
                eliminado=False,
                estado__in=['pendiente', 'aprobado_gerente', 'aprobado_contador'],
                contrato__proyecto_id=proyecto_id
            )
            
            # Excluir los que ya están en planilla
            avaluos_en_planilla = DetallePlanillaContratista.objects.values_list('avaluo_id', flat=True)
            avaluos = avaluos.exclude(id__in=avaluos_en_planilla)
            
            # Filtro opcional por contratista
            if contratista_id:
                avaluos = avaluos.filter(contrato__contratista_id=contratista_id)
            
            context['avaluos_disponibles'] = avaluos.select_related(
                'contrato__contratista',
                'contrato__proyecto'
            ).order_by('contrato__contratista__apellido', 'fecha_pago')
            
            context['filtros_activos'] = True
        else:
            context['avaluos_disponibles'] = AvaluoContratista.objects.none()
            context['filtros_activos'] = False
        
        context['tipo_cambio'] = get_tipo_cambio_actual()
        context['proyecto_seleccionado'] = proyecto_id
        context['contratista_seleccionado'] = contratista_id
        
        return context
    
    def form_valid(self, form):
        avaluos_ids = self.request.POST.getlist('avaluos')
        
        if not avaluos_ids:
            messages.warning(self.request, '⚠️ Selecciona al menos un avalúo')
            return self.form_invalid(form)
        
        # Crear planilla
        planilla = form.save(commit=False)
        planilla.generada_por = self.request.user
        planilla.tipo_cambio = get_tipo_cambio_actual()
        planilla.save()
        
        # Agregar avalúos
        for avaluo_id in avaluos_ids:
            avaluo = AvaluoContratista.objects.get(id=avaluo_id)
            DetallePlanillaContratista.objects.create(
                planilla=planilla,
                avaluo=avaluo
            )
        
        planilla.calcular_totales()
        
        messages.success(
            self.request,
            f'✅ Planilla {planilla.codigo} creada con {len(avaluos_ids)} avalúos'
        )
        
        return redirect('planillas_contratistas_detalle', pk=planilla.pk)

class ObtenerContratistasProyectoView(LoginRequiredMixin, View):
    """Retorna contratistas con avalúos disponibles en un proyecto"""
    
    def get(self, request):
        proyecto_id = request.GET.get('proyecto_id')
        
        if not proyecto_id:
            return JsonResponse({'contratistas': []})
        
        # Avalúos disponibles del proyecto
        avaluos_disponibles = AvaluoContratista.objects.filter(
            eliminado=False,
            estado__in=['pendiente', 'aprobado_gerente', 'aprobado_contador'],
            contrato__proyecto_id=proyecto_id
        )
        
        # Excluir avalúos ya en planilla
        avaluos_en_planilla = DetallePlanillaContratista.objects.values_list('avaluo_id', flat=True)
        avaluos_disponibles = avaluos_disponibles.exclude(id__in=avaluos_en_planilla)
        
        # Obtener IDs de contratistas únicos
        contratistas_ids = avaluos_disponibles.values_list(
            'contrato__contratista_id', 
            flat=True
        ).distinct()
        
        # Obtener contratistas
        contratistas = Contratista.objects.filter(
            id__in=contratistas_ids,
            eliminado=False
        ).order_by('apellido', 'nombre')
        
        # Formatear respuesta
        contratistas_data = [
            {
                'id': c.id,
                'nombre': c.nombre_completo
            }
            for c in contratistas
        ]
        
        return JsonResponse({'contratistas': contratistas_data})

class PlanillaDetalleView(LoginRequiredMixin, DetailView):
    """Vista de detalle de planilla de contratistas"""
    model = PlanillaContratista
    template_name = 'contratistas/planilla_detalle.html'
    context_object_name = 'planilla'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        planilla = self.object
        
        # Detalles de la planilla
        detalles = planilla.detalles.select_related(
            'avaluo__contrato__contratista',
            'avaluo__contrato__proyecto'
        ).order_by('contratista_nombre')
        
        # 🔍 DEBUG: Ver qué hay en detalles
        print("=" * 80)
        print(f"🔍 DEBUG: Planilla {planilla.codigo}")
        print(f"Total detalles: {detalles.count()}")
        
        for detalle in detalles:
            print(f"\n📋 Detalle ID: {detalle.id}")
            print(f"   Contratista: {detalle.contratista_nombre}")
            print(f"   Forma de pago: '{detalle.forma_pago}' (tipo: {type(detalle.forma_pago)})")
            print(f"   Monto C$: {detalle.monto_cordobas}")
            print(f"   Monto USD: {detalle.monto_dolares}")
        
        print("=" * 80)
        
        context['detalles'] = detalles
        
        # CALCULAR TOTALES POR FORMA DE PAGO
        totales_transferencia = detalles.filter(
            forma_pago__iexact='transferencia'  # ✅ iexact = case-insensitive
        ).aggregate(
            cordobas=Sum('monto_cordobas'),
            dolares=Sum('monto_dolares')
        )
        
        print(f"\n💰 Totales transferencia:")
        print(f"   Córdobas: {totales_transferencia['cordobas']}")
        print(f"   Dólares: {totales_transferencia['dolares']}")
        
        totales_efectivo = detalles.filter(
            forma_pago='efectivo'
        ).aggregate(
            cordobas=Sum('monto_cordobas'),
            dolares=Sum('monto_dolares')
        )
        
        totales_cheque = detalles.filter(
            forma_pago='cheque'
        ).aggregate(
            cordobas=Sum('monto_cordobas'),
            dolares=Sum('monto_dolares')
        )
        
        context['total_transferencia_cordobas'] = totales_transferencia['cordobas'] or 0
        context['total_transferencia_dolares'] = totales_transferencia['dolares'] or 0
        
        context['total_efectivo_cordobas'] = totales_efectivo['cordobas'] or 0
        context['total_efectivo_dolares'] = totales_efectivo['dolares'] or 0
        
        context['total_cheque_cordobas'] = totales_cheque['cordobas'] or 0
        context['total_cheque_dolares'] = totales_cheque['dolares'] or 0
        
        # Cantidad de contratistas únicos
        context['cantidad_contratistas'] = detalles.values(
            'contratista_cedula'
        ).distinct().count()
        
        return context

class PlanillaAprobarGerenteView(LoginRequiredMixin, View):
    """Aprobar planilla por gerente"""
    
    def post(self, request, pk):
        planilla = get_object_or_404(PlanillaContratista, pk=pk)
        
        if planilla.estado != 'borrador':
            messages.error(request, '❌ Solo se pueden aprobar planillas en borrador')
            return redirect('planillas_contratistas_detalle', pk=pk)
        
        # Cambiar estado de la planilla
        planilla.estado = 'aprobada_gerente'
        planilla.aprobada_gerente_por = request.user
        planilla.fecha_aprobacion_gerente = timezone.now()
        planilla.save()
        
        # ✅ NUEVO: Cambiar estado de avalúos
        for detalle in planilla.detalles.all():
            if detalle.avaluo.estado == 'pendiente':
                detalle.avaluo.estado = 'aprobado_gerente'
                detalle.avaluo.aprobado_gerente_por = request.user
                detalle.avaluo.fecha_aprobacion_gerente = timezone.now()
                detalle.avaluo.save()
        
        messages.success(request, f'✅ Planilla {planilla.codigo} aprobada por gerente')
        return redirect('planillas_contratistas_detalle', pk=pk)


class PlanillaAprobarContadorView(LoginRequiredMixin, View):
    """Aprobar planilla por contador"""
    
    def post(self, request, pk):
        planilla = get_object_or_404(PlanillaContratista, pk=pk)
        
        if planilla.estado != 'aprobada_gerente':
            messages.error(request, '❌ La planilla debe estar aprobada por gerente primero')
            return redirect('planillas_contratistas_detalle', pk=pk)
        
        # Cambiar estado de la planilla
        planilla.estado = 'aprobada_contador'
        planilla.aprobada_contador_por = request.user
        planilla.fecha_aprobacion_contador = timezone.now()
        planilla.save()
        
        # ✅ NUEVO: Cambiar estado de avalúos
        for detalle in planilla.detalles.all():
            if detalle.avaluo.estado in ['pendiente', 'aprobado_gerente']:
                detalle.avaluo.estado = 'aprobado_contador'
                detalle.avaluo.aprobado_contador_por = request.user
                detalle.avaluo.fecha_aprobacion_contador = timezone.now()
                detalle.avaluo.save()
        
        messages.success(request, f'✅ Planilla {planilla.codigo} aprobada por contador')
        return redirect('planillas_contratistas_detalle', pk=pk)


class PlanillaMarcarPagadaView(LoginRequiredMixin, View):
    """Marcar planilla como pagada"""
    
    def post(self, request, pk):
        planilla = get_object_or_404(PlanillaContratista, pk=pk)
        
        if planilla.estado != 'aprobada_contador':
            messages.error(request, '❌ La planilla debe estar aprobada por el contador primero')
            return redirect('planillas_contratistas_detalle', pk=pk)
        
        # Cambiar estado de la planilla
        planilla.estado = 'pagada'
        planilla.pagada_por = request.user
        planilla.fecha_pago = timezone.now()
        planilla.save()
        
        # ✅ NUEVO: Marcar avalúos como pagados
        for detalle in planilla.detalles.all():
            detalle.avaluo.estado = 'pagado'
            detalle.avaluo.save()
        
        messages.success(request, f'✅ Planilla {planilla.codigo} marcada como pagada')
        return redirect('planillas_contratistas_detalle', pk=pk)

class PlanillaExportarExcelView(LoginRequiredMixin, View):
    """Exportar planilla a Excel con formato"""
    
    def get(self, request, pk):
        from apps.contratistas.models import PlanillaContratista
        
        planilla = get_object_or_404(PlanillaContratista, pk=pk)
        detalles = planilla.detalles.select_related(
            'avaluo__contrato__contratista',
            'avaluo__contrato__proyecto'
        ).order_by('contratista_nombre')
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Planilla {planilla.codigo}"
        
        # ========================================
        # ENCABEZADO CON ESTILO
        # ========================================
        
        # Fila 1: PROYECTO (Fondo azul, texto blanco, negrita)
        ws['A1'] = f"PROYECTO: {planilla.proyecto.nombre.upper()}"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:L1')
        ws.row_dimensions[1].height = 30
        
        # Fila 2: PLANILLA DE PAGOS (Fondo azul claro)
        ws['A2'] = "PLANILLA DE PAGOS"
        ws['A2'].font = Font(size=14, bold=True, color='1F4E78')
        ws['A2'].fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A2:L2')
        ws.row_dimensions[2].height = 25
        
        # Fila 3: SUBCONTRATISTAS (Fondo azul claro)
        ws['A3'] = "SUBCONTRATISTAS"
        ws['A3'].font = Font(size=13, bold=True, color='1F4E78')
        ws['A3'].fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A3:L3')
        ws.row_dimensions[3].height = 22
        
        # Fila 4: FECHA
        meses = {
            1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
            5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
            9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
        }
        mes_texto = meses[planilla.fecha_generacion.month]
        ws['A4'] = f"FECHA: {planilla.fecha_generacion.day} DE {mes_texto} DE {planilla.fecha_generacion.year}"
        ws['A4'].font = Font(size=11, bold=True)
        ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Fila 5: PERIODO
        mes_inicio = meses[planilla.periodo_inicio.month]
        mes_fin = meses[planilla.periodo_fin.month]
        ws['A5'] = f"PERIODO: DEL {planilla.periodo_inicio.day} DE {mes_inicio} AL {planilla.periodo_fin.day} DE {mes_fin} DEL {planilla.periodo_fin.year}"
        ws['A5'].font = Font(size=11, bold=True)
        ws['A5'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Fila 6: TIPO DE CAMBIO (Con fondo amarillo suave)
        ws['A6'] = "TIPO DE CAMBIO BCN:"
        ws['B6'] = float(planilla.tipo_cambio)
        ws['A6'].font = Font(size=11, bold=True)
        ws['A6'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        ws['B6'].font = Font(size=11, bold=True, color='C00000')
        ws['B6'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        ws['B6'].number_format = '0.0000'
        ws['B6'].alignment = Alignment(horizontal='center')
        
        # Fila vacía
        current_row = 8
        
        # ========================================
        # ENCABEZADO DE TABLA CON ESTILO
        # ========================================
        
        headers = ['N°', 'NOMBRE Y APELLIDO', 'CÉDULA', 'Actividad Catálogo', 'DESCRIPCIÓN', 
                   'Pago C$', 'Pago en USD', 'Forma de pago', 'Banco', 'Moneda', 
                   'Numero de cuenta', 'FIRMA']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='medium', color='1F4E78'),
                right=Side(style='medium', color='1F4E78'),
                top=Side(style='medium', color='1F4E78'),
                bottom=Side(style='medium', color='1F4E78')
            )
        
        # Altura de fila de encabezado
        ws.row_dimensions[current_row].height = 35
        
        current_row += 1
        start_data_row = current_row
        
        # ========================================
        # DATOS DE AVALÚOS CON ESTILO
        # ========================================
        
        for idx, detalle in enumerate(detalles, 1):
            # Alternar colores de fila (zebra striping)
            if idx % 2 == 0:
                fill_color = 'F2F2F2'  # Gris claro
            else:
                fill_color = 'FFFFFF'  # Blanco
            
            ws.cell(row=current_row, column=1, value=idx)  # N°
            ws.cell(row=current_row, column=2, value=detalle.contratista_nombre)  # Nombre
            ws.cell(row=current_row, column=3, value=detalle.contratista_cedula)  # Cédula
            ws.cell(row=current_row, column=4, value='')  # Actividad Catálogo (vacío por ahora)
            ws.cell(row=current_row, column=5, value=detalle.actividad)  # Descripción (denormalizado en detalle)
            ws.cell(row=current_row, column=6, value=float(detalle.monto_cordobas))  # C$
            ws.cell(row=current_row, column=7, value=float(detalle.monto_dolares))  # USD
            
            # Forma de pago con capitalización
            forma_pago_texto = detalle.forma_pago.capitalize() if detalle.forma_pago else ''
            ws.cell(row=current_row, column=8, value=forma_pago_texto)  # Forma pago
            
            ws.cell(row=current_row, column=9, value=detalle.banco or '')  # Banco (denormalizado en detalle)
            ws.cell(row=current_row, column=10, value=detalle.moneda_cuenta or '')  # Moneda (denormalizado en detalle)
            ws.cell(row=current_row, column=11, value=detalle.numero_cuenta or '')  # Cuenta (denormalizado en detalle)
            ws.cell(row=current_row, column=12, value='')  # Firma
            
            # Aplicar estilos a todas las celdas de la fila
            for col in range(1, 13):
                cell = ws.cell(row=current_row, column=col)
                
                # Fondo alternado
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                # Alineación
                if col == 1:  # N°
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col in [6, 7]:  # Montos
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Formato moneda
                if col == 6:  # Córdobas
                    cell.number_format = 'C$#,##0.00'
                    cell.font = Font(bold=True, color='0070C0')
                elif col == 7:  # Dólares
                    cell.number_format = '$#,##0.00'
                    cell.font = Font(bold=True, color='00B050')
                
                # Bordes
                cell.border = Border(
                    left=Side(style='thin', color='BFBFBF'),
                    right=Side(style='thin', color='BFBFBF'),
                    top=Side(style='thin', color='BFBFBF'),
                    bottom=Side(style='thin', color='BFBFBF')
                )
            
            # Altura de fila
            ws.row_dimensions[current_row].height = 25
            current_row += 1
        
        end_data_row = current_row - 1
        
        # ========================================
        # TOTALES CON ESTILO
        # ========================================
        
        current_row += 1
        
        # TOTAL EN CÓRDOBAS (Fondo verde claro)
        ws.cell(row=current_row, column=1, value="TOTAL EN CÓRDOBAS")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color='FFFFFF')
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(f'A{current_row}:E{current_row}')
        
        ws.cell(row=current_row, column=6, value=f'=SUM(F{start_data_row}:F{end_data_row})')
        ws.cell(row=current_row, column=6).font = Font(bold=True, size=13, color='FFFFFF')
        ws.cell(row=current_row, column=6).fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        ws.cell(row=current_row, column=6).number_format = 'C$#,##0.00'
        ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=current_row, column=6).border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        ws.row_dimensions[current_row].height = 28
        current_row += 1
        
        # TOTAL EN DÓLARES (Fondo azul claro)
        ws.cell(row=current_row, column=1, value="TOTAL EN DÓLARES")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color='FFFFFF')
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(f'A{current_row}:F{current_row}')
        
        ws.cell(row=current_row, column=7, value=f'=SUM(G{start_data_row}:G{end_data_row})')
        ws.cell(row=current_row, column=7).font = Font(bold=True, size=13, color='FFFFFF')
        ws.cell(row=current_row, column=7).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws.cell(row=current_row, column=7).number_format = 'U$#,##0.00'
        ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=current_row, column=7).border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        ws.row_dimensions[current_row].height = 28
        
        # ========================================
        # AJUSTAR ANCHOS DE COLUMNA
        # ========================================
        
        column_widths = {
            'A': 6,   # N°
            'B': 30,  # Nombre
            'C': 18,  # Cédula
            'D': 15,  # Actividad
            'E': 45,  # Descripción
            'F': 15,  # Pago C$
            'G': 15,  # Pago USD
            'H': 18,  # Forma pago
            'I': 18,  # Banco
            'J': 12,  # Moneda
            'K': 18,  # Cuenta
            'L': 15   # Firma
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # ========================================
        # PREPARAR RESPUESTA HTTP
        # ========================================
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{planilla.codigo}_Planilla_Contratistas.xlsx"'
        
        wb.save(response)
        return response
