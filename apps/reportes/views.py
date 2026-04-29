"""
Vistas del módulo de Reportes
apps/reportes/views.py
"""
from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, View
from django.db.models import Sum, Count, Q
from decimal import Decimal
from datetime import datetime, timedelta

from .models import ConfiguracionEmpresa
from apps.proyectos.models import Proyecto
from apps.planillas.models import Planilla, DetallePlanilla, PlanillaReembolso
from apps.contratistas.models import (
    Contratista, 
    ContratoProyecto, 
    AvaluoContratista,
    PlanillaContratista,
    DetallePlanillaContratista
)
from apps.core.utils import get_tipo_cambio_actual
from apps.trabajadores.models import Trabajador
from apps.admin_panel.permissions import PermissionRequiredMixin

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from collections import defaultdict

from django.conf import settings
import os

class ReportesIndexView(LoginRequiredMixin, TemplateView):
    """
    Vista principal del módulo de reportes
    Muestra el índice con todos los tipos de reportes disponibles
    """
    template_name = 'reportes/index.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Configuración de la empresa
        context['configuracion'] = ConfiguracionEmpresa.get_configuracion()
        
        # Tipo de cambio actual
        context['tipo_cambio'] = get_tipo_cambio_actual()
        
        # Lista de proyectos activos
        context['proyectos'] = Proyecto.objects.filter(
            eliminado=False
        ).order_by('nombre')
        
        # Estadísticas generales
        context['total_proyectos'] = Proyecto.objects.filter(eliminado=False).count()
        context['total_planillas'] = Planilla.objects.filter(eliminado=False).count()
        context['total_planillas_contratistas'] = PlanillaContratista.objects.filter(eliminado=False).count()
        
        return context

class ReportePorProyectoView(LoginRequiredMixin, TemplateView):
    """
    Vista para generar reporte detallado por proyecto
    
    FUNCIONAMIENTO:
    1. Usuario selecciona proyecto → Muestra últimas 20 planillas automáticamente
    2. Opcionalmente puede filtrar por rango de fechas
    3. Usuario hace clic en una planilla → Muestra detalle completo
    """
    template_name = 'reportes/reporte_proyecto.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Configuración de la empresa
        context['configuracion'] = ConfiguracionEmpresa.get_configuracion()
        
        # Tipo de cambio
        tipo_cambio = get_tipo_cambio_actual()
        context['tipo_cambio'] = tipo_cambio
        
        # Lista de proyectos para el selector
        context['proyectos'] = Proyecto.objects.filter(
            eliminado=False
        ).order_by('nombre')
        
        # Obtener parámetros
        proyecto_id = self.request.GET.get('proyecto')
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        planilla_id = self.request.GET.get('planilla_id')
        tipo_planilla = self.request.GET.get('tipo_planilla')
        
        # ==========================================
        # MOSTRAR LISTA DE PLANILLAS DISPONIBLES
        # ==========================================
        if proyecto_id and fecha_inicio and fecha_fin and not self.request.GET.get('consolidar'):
            try:
                proyecto = Proyecto.objects.get(id=proyecto_id, eliminado=False)
                context['proyecto_seleccionado'] = proyecto
                
                # Preparar queryset base de planillas de trabajadores
                planillas_trabajadores_qs = Planilla.objects.filter(
                    proyecto=proyecto,
                    estado='pagada',
                    eliminado=False
                )
                
                # Preparar queryset base de planillas de contratistas
                planillas_contratistas_qs = PlanillaContratista.objects.filter(
                    proyecto=proyecto,
                    estado='pagada',
                    eliminado=False
                )
                
                # Si hay filtro de fechas, aplicarlo
                if fecha_inicio and fecha_fin:
                    fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                    fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                    
                    planillas_trabajadores_qs = planillas_trabajadores_qs.filter(
                        periodo_inicio__gte=fecha_inicio_obj,
                        periodo_fin__lte=fecha_fin_obj
                    )
                    
                    planillas_contratistas_qs = planillas_contratistas_qs.filter(
                        periodo_inicio__gte=fecha_inicio_obj,
                        periodo_fin__lte=fecha_fin_obj
                    )
                    
                    context['fecha_inicio'] = fecha_inicio
                    context['fecha_fin'] = fecha_fin
                    context['filtro_aplicado'] = True
                
                # Obtener planillas ordenadas por fecha (PRIMERO ordenar, LUEGO limitar)
                planillas_trabajadores_qs = planillas_trabajadores_qs.select_related('proyecto').order_by('-fecha_generacion')
                planillas_contratistas_qs = planillas_contratistas_qs.select_related('proyecto').order_by('-fecha_generacion')
                
                # Limitar a las últimas 20 planillas si no hay filtro
                if not (fecha_inicio and fecha_fin):
                    planillas_trabajadores = planillas_trabajadores_qs[:20]
                    planillas_contratistas = planillas_contratistas_qs[:20]
                else:
                    planillas_trabajadores = planillas_trabajadores_qs
                    planillas_contratistas = planillas_contratistas_qs
                
                # Preparar lista de reportes disponibles
                reportes_disponibles = []
                
                # Agregar planillas de trabajadores
                for planilla in planillas_trabajadores:
                    total_trabajadores = DetallePlanilla.objects.filter(
                        planilla=planilla
                    ).count()
                    
                    # Usar el campo correcto: salario_devengado (que es el total a pagar)
                    total_cordobas = DetallePlanilla.objects.filter(
                        planilla=planilla
                    ).aggregate(
                        total=Sum('ingreso_total')
                    )['total'] or Decimal('0.00')
                    
                    # Calcular dólares usando tipo de cambio
                    total_dolares = total_cordobas / tipo_cambio if total_cordobas else Decimal('0.00')
                    
                    reportes_disponibles.append({
                        'id': f'trabajadores_{planilla.id}',
                        'tipo': 'Trabajadores',
                        'fecha': planilla.fecha_generacion,
                        'periodo_inicio': planilla.periodo_inicio,
                        'periodo_fin': planilla.periodo_fin,
                        'cantidad': total_trabajadores,
                        'total_cordobas': total_cordobas,
                        'total_dolares': total_dolares,
                        'planilla_id': planilla.id,
                        'tipo_planilla': 'trabajadores'
                    })
                
                # Agregar planillas de contratistas
                for planilla in planillas_contratistas:
                    total_contratistas = DetallePlanillaContratista.objects.filter(
                        planilla=planilla
                    ).count()
                    
                    total_cordobas = DetallePlanillaContratista.objects.filter(
                        planilla=planilla
                    ).aggregate(
                        total=Sum('monto_cordobas')
                    )['total'] or Decimal('0.00')
                    
                    total_dolares = DetallePlanillaContratista.objects.filter(
                        planilla=planilla
                    ).aggregate(
                        total=Sum('monto_dolares')
                    )['total'] or Decimal('0.00')
                    
                    reportes_disponibles.append({
                        'id': f'contratistas_{planilla.id}',
                        'tipo': 'Contratistas',
                        'fecha': planilla.fecha_generacion,
                        'periodo_inicio': planilla.periodo_inicio,
                        'periodo_fin': planilla.periodo_fin,
                        'cantidad': total_contratistas,
                        'total_cordobas': total_cordobas,
                        'total_dolares': total_dolares,
                        'planilla_id': planilla.id,
                        'tipo_planilla': 'contratistas'
                    })
                
                # Ordenar por fecha descendente
                reportes_disponibles.sort(key=lambda x: x['fecha'], reverse=True)
                
                context['reportes_disponibles'] = reportes_disponibles
                context['mostrar_lista'] = True
                
            except Proyecto.DoesNotExist:
                context['error'] = 'El proyecto seleccionado no existe'
        
        # ==========================================
        # GENERAR REPORTE CONSOLIDADO
        # ==========================================
        elif proyecto_id and self.request.GET.get('consolidar') == '1':
            try:
                proyecto = Proyecto.objects.get(id=proyecto_id, eliminado=False)
                context['proyecto_seleccionado'] = proyecto
                
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                context['fecha_inicio'] = fecha_inicio
                context['fecha_fin'] = fecha_fin
                
                # Planillas de trabajadores en el rango
                planillas_trab = Planilla.objects.filter(
                    proyecto=proyecto,
                    estado='pagada',
                    eliminado=False,
                    periodo_inicio__gte=fecha_inicio_obj,
                    periodo_fin__lte=fecha_fin_obj
                )
                
                # Planillas de contratistas en el rango
                planillas_cont = PlanillaContratista.objects.filter(
                    proyecto=proyecto,
                    estado='pagada',
                    eliminado=False,
                    periodo_inicio__gte=fecha_inicio_obj,
                    periodo_fin__lte=fecha_fin_obj
                )
                
                # Detalles de trabajadores
                detalles_trab = DetallePlanilla.objects.filter(
                    planilla__in=planillas_trab
                ).select_related('trabajador')
                
                oficiales = []
                ayudantes = []
                
                for detalle in detalles_trab:
                    trabajador_data = {
                        'nombre': detalle.trabajador.nombre_completo,
                        'cedula': detalle.trabajador.numero_cedula or '',
                        'cargo': detalle.cargo or 'N/A',
                        'dias_trabajados': detalle.dias_laborados or 0,
                        'dias_feriados': detalle.dias_feriados or 0,
                        'horas_extras': detalle.horas_extras or Decimal('0.00'),
                        'salario_dia_base': detalle.salario_dia_base or Decimal('0.00'),
                        'valor_septimo': detalle.valor_septimo_dia or Decimal('0.00'),
                        'salario_devengado': detalle.salario_devengado or Decimal('0.00'),
                        'salario_horas_extras': detalle.salario_horas_extras or Decimal('0.00'),
                        'ingreso_feriado': detalle.ingreso_dia_feriado or Decimal('0.00'),
                        'combustible': detalle.combustible or Decimal('0.00'),
                        'otros': detalle.otros_gastos or Decimal('0.00'),
                        'deducciones': detalle.deducciones or Decimal('0.00'),
                        'total_cordobas': detalle.ingreso_total or Decimal('0.00'),
                        'total_dolares': (detalle.ingreso_total / tipo_cambio) if detalle.ingreso_total else Decimal('0.00'),
                    }
                    
                    cargo_lower = (detalle.cargo or '').lower()
                    area_lower = (detalle.area or '').lower()
                    if any(w in cargo_lower for w in ['oficial', 'maestro', 'ingeniero', 'fontanero', 'electricista', 'soldador', 'albañil']) or 'oficial' in area_lower:
                        oficiales.append(trabajador_data)
                    else:
                        ayudantes.append(trabajador_data)
                
                # Detalles de contratistas
                detalles_cont = DetallePlanillaContratista.objects.filter(
                    planilla__in=planillas_cont
                ).select_related('avaluo__contrato__contratista')
                
                contratistas_data = []
                for detalle in detalles_cont:
                    contratistas_data.append({
                        'nombre': detalle.avaluo.contrato.contratista.nombre_completo,
                        'cedula': detalle.avaluo.contrato.contratista.numero_cedula or '',
                        'contrato': f"Contrato #{detalle.avaluo.contrato.id}",
                        'descripcion': detalle.avaluo.concepto or 'N/A',
                        'monto_cordobas': detalle.monto_cordobas or Decimal('0.00'),
                        'monto_dolares': detalle.monto_dolares or Decimal('0.00'),
                    })
                
                # Totales
                context['oficiales'] = oficiales
                context['ayudantes'] = ayudantes
                context['contratistas'] = contratistas_data
                
                context['total_oficiales_cordobas'] = sum(o['total_cordobas'] for o in oficiales)
                context['total_oficiales_dolares'] = sum(o['total_dolares'] for o in oficiales)
                context['total_ayudantes_cordobas'] = sum(a['total_cordobas'] for a in ayudantes)
                context['total_ayudantes_dolares'] = sum(a['total_dolares'] for a in ayudantes)
                context['total_contratistas_cordobas'] = sum(c['monto_cordobas'] for c in contratistas_data)
                context['total_contratistas_dolares'] = sum(c['monto_dolares'] for c in contratistas_data)
                
                # Resumen por área
                resumen = [
                    {'area': 'Oficiales', 'total_cordobas': context['total_oficiales_cordobas'], 'total_dolares': context['total_oficiales_dolares']},
                    {'area': 'Ayudantes', 'total_cordobas': context['total_ayudantes_cordobas'], 'total_dolares': context['total_ayudantes_dolares']},
                    {'area': 'Sub-Contratistas', 'total_cordobas': context['total_contratistas_cordobas'], 'total_dolares': context['total_contratistas_dolares']},
                ]
                
                context['resumen'] = resumen
                context['gran_total_cordobas'] = sum(r['total_cordobas'] for r in resumen)
                context['gran_total_dolares'] = sum(r['total_dolares'] for r in resumen)
                context['planillas_incluidas_trab'] = planillas_trab.count()
                context['planillas_incluidas_cont'] = planillas_cont.count()
                context['datos_reporte'] = True
                
            except Proyecto.DoesNotExist:
                context['error'] = 'El proyecto seleccionado no existe'
            except Exception as e:
                context['error'] = f'Error al generar reporte: {str(e)}'
        
        return context


class ReportePlanillaAdministrativaView(LoginRequiredMixin, TemplateView):
    """
    Vista para generar reporte de planilla administrativa
    
    Personal administrativo NO asociado a proyectos de construcción
    (gerentes, contadores, secretarias, etc.)
    """
    template_name = 'reportes/reporte_administrativa.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Configuración de la empresa
        context['configuracion'] = ConfiguracionEmpresa.get_configuracion()
        
        # Tipo de cambio
        tipo_cambio = get_tipo_cambio_actual()
        context['tipo_cambio'] = tipo_cambio
        
        # Obtener parámetros de filtro
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        planilla_id = self.request.GET.get('planilla_id')
        
        # ==========================================
        # MOSTRAR LISTA DE PLANILLAS DISPONIBLES
        # ==========================================
        if fecha_inicio and fecha_fin and not planilla_id:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                
                # Buscar proyecto administrativo
                # Puede tener diferentes nombres: "Administración", "Administrativa", etc.
                proyecto_admin = Proyecto.objects.filter(
                    Q(nombre__icontains='administra') | 
                    Q(nombre__icontains='general') |
                    Q(nombre__iexact='Administración General'),
                    eliminado=False
                ).first()
                
                if not proyecto_admin:
                    context['error'] = 'No se encontró el proyecto administrativo. Por favor, créalo con el nombre "Administración General".'
                    context['fecha_inicio'] = fecha_inicio
                    context['fecha_fin'] = fecha_fin
                    return context
                
                # Obtener planillas administrativas en el rango
                planillas_administrativas = Planilla.objects.filter(
                    proyecto=proyecto_admin,
                    estado='pagada',
                    fecha_generacion__gte=fecha_inicio_obj,
                    fecha_generacion__lte=fecha_fin_obj,
                    eliminado=False
                ).select_related('proyecto').order_by('-fecha_generacion')[:20]
                
                # Preparar lista de reportes disponibles
                reportes_disponibles = []
                
                for planilla in planillas_administrativas:
                    total_personal = DetallePlanilla.objects.filter(
                        planilla=planilla
                    ).count()
                    
                    total_cordobas = DetallePlanilla.objects.filter(
                        planilla=planilla
                    ).aggregate(
                        total=Sum('ingreso_total')
                    )['total'] or Decimal('0.00')
                    
                    total_dolares = total_cordobas / tipo_cambio if total_cordobas > 0 else Decimal('0.00')
                    
                    reportes_disponibles.append({
                        'planilla_id': planilla.id,
                        'fecha': planilla.fecha_generacion,
                        'periodo_inicio': planilla.periodo_inicio,
                        'periodo_fin': planilla.periodo_fin,
                        'cantidad': total_personal,
                        'total_cordobas': total_cordobas,
                        'total_dolares': total_dolares,
                    })
                
                context['reportes_disponibles'] = reportes_disponibles
                context['proyecto_admin'] = proyecto_admin
                context['fecha_inicio'] = fecha_inicio
                context['fecha_fin'] = fecha_fin
                context['mostrar_lista'] = True
                
            except Exception as e:
                context['error'] = f'Error al buscar planillas: {str(e)}'
        
        # ==========================================
        # MOSTRAR DETALLE DEL REPORTE SELECCIONADO
        # ==========================================
        elif planilla_id:
            try:
                planilla = Planilla.objects.get(id=planilla_id, eliminado=False)
                proyecto = planilla.proyecto
                
                # Obtener detalles del personal administrativo
                detalles = DetallePlanilla.objects.filter(
                    planilla=planilla
                ).select_related('trabajador').order_by('cargo', 'trabajador__nombre_completo')
                
                # Agrupar por área/cargo
                personal_por_area = {}
                total_general_c = Decimal('0.00')
                total_general_d = Decimal('0.00')
                
                for detalle in detalles:
                    area = detalle.cargo or 'Sin área definida'
                    
                    if area not in personal_por_area:
                        personal_por_area[area] = {
                            'personal': [],
                            'total_cordobas': Decimal('0.00'),
                            'total_dolares': Decimal('0.00'),
                        }
                    
                    salario_c = detalle.ingreso_total or Decimal('0.00')
                    salario_d = salario_c / tipo_cambio if salario_c > 0 else Decimal('0.00')
                    
                    personal_por_area[area]['personal'].append({
                        'nombre': detalle.trabajador.nombre_completo,
                        'cargo': detalle.cargo or 'N/A',
                        'dias': detalle.dias_laborados or 0,
                        'salario_cordobas': salario_c,
                        'salario_dolares': salario_d,
                    })
                    
                    personal_por_area[area]['total_cordobas'] += salario_c
                    personal_por_area[area]['total_dolares'] += salario_d
                    
                    total_general_c += salario_c
                    total_general_d += salario_d
                
                # Convertir dict a lista para el template
                areas_data = []
                for area_nombre, area_info in personal_por_area.items():
                    porcentaje = (area_info['total_cordobas'] / total_general_c * 100) if total_general_c > 0 else 0
                    areas_data.append({
                        'nombre': area_nombre,
                        'personal': area_info['personal'],
                        'total_cordobas': area_info['total_cordobas'],
                        'total_dolares': area_info['total_dolares'],
                        'porcentaje': porcentaje,
                    })
                
                # Ordenar por monto descendente
                areas_data.sort(key=lambda x: x['total_cordobas'], reverse=True)
                
                context['areas_data'] = areas_data
                context['total_general_cordobas'] = total_general_c
                context['total_general_dolares'] = total_general_d
                context['proyecto_admin'] = proyecto
                context['fecha_inicio'] = planilla.periodo_inicio
                context['fecha_fin'] = planilla.periodo_fin
                context['fecha_generacion'] = planilla.fecha_generacion
                context['datos_reporte'] = True
                
            except Planilla.DoesNotExist:
                context['error'] = 'La planilla seleccionada no existe'
        
        return context

class ReporteGastosVariosView(LoginRequiredMixin, TemplateView):
    """
    Vista para generar reporte de gastos varios y reembolsos
    
    Gastos NO asociados a trabajadores ni proyectos específicos
    Agrupados por concepto/categoría para análisis
    """
    template_name = 'reportes/reporte_gastos_varios.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Configuración de la empresa
        context['configuracion'] = ConfiguracionEmpresa.get_configuracion()
        
        # Tipo de cambio
        tipo_cambio = get_tipo_cambio_actual()
        context['tipo_cambio'] = tipo_cambio
        
        # Obtener parámetros de filtro
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        comparar = self.request.GET.get('comparar', 'no')
        
        if fecha_inicio and fecha_fin:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                
                context['fecha_inicio'] = fecha_inicio
                context['fecha_fin'] = fecha_fin
                context['comparar'] = comparar
                
                # ==========================================
                # OBTENER GASTOS DEL PERIODO ACTUAL
                # ==========================================
                
                # Obtener planillas pagadas en el rango
                planillas = Planilla.objects.filter(
                    estado='pagada',
                    fecha_generacion__gte=fecha_inicio_obj,
                    fecha_generacion__lte=fecha_fin_obj,
                    eliminado=False
                )
                
                # Obtener todos los reembolsos de esas planillas
                reembolsos_actuales = PlanillaReembolso.objects.filter(
                    planilla__in=planillas
                ).select_related('planilla').order_by('concepto', 'numero_item')
                
                # Agrupar por concepto/categoría
                gastos_por_concepto = defaultdict(lambda: {
                    'items': [],
                    'total_cordobas': Decimal('0.00'),
                    'total_dolares': Decimal('0.00'),
                    'cantidad': 0
                })
                
                total_general_c = Decimal('0.00')
                total_general_d = Decimal('0.00')
                
                for reembolso in reembolsos_actuales:
                    concepto = reembolso.concepto.strip().upper()
                    monto_c = reembolso.monto_cordobas or Decimal('0.00')
                    monto_d = monto_c / tipo_cambio if monto_c > 0 else Decimal('0.00')
                    
                    gastos_por_concepto[concepto]['items'].append({
                        'numero': reembolso.numero_item,
                        'concepto': reembolso.concepto,
                        'monto_cordobas': monto_c,
                        'monto_dolares': monto_d,
                        'observaciones': reembolso.observaciones or '',
                        'fecha': reembolso.planilla.fecha_generacion,
                    })
                    
                    gastos_por_concepto[concepto]['total_cordobas'] += monto_c
                    gastos_por_concepto[concepto]['total_dolares'] += monto_d
                    gastos_por_concepto[concepto]['cantidad'] += 1
                    
                    total_general_c += monto_c
                    total_general_d += monto_d
                
                # ==========================================
                # COMPARACIÓN CON PERIODO ANTERIOR (opcional)
                # ==========================================
                if comparar == 'si' and total_general_c > 0:
                    # Calcular fechas del periodo anterior
                    dias_periodo = (fecha_fin_obj - fecha_inicio_obj).days + 1
                    fecha_inicio_anterior = fecha_inicio_obj - timedelta(days=dias_periodo)
                    fecha_fin_anterior = fecha_inicio_obj - timedelta(days=1)
                    
                    # Obtener planillas del periodo anterior
                    planillas_anterior = Planilla.objects.filter(
                        estado='pagada',
                        fecha_generacion__gte=fecha_inicio_anterior,
                        fecha_generacion__lte=fecha_fin_anterior,
                        eliminado=False
                    )
                    
                    # Obtener reembolsos del periodo anterior
                    reembolsos_anterior = PlanillaReembolso.objects.filter(
                        planilla__in=planillas_anterior
                    )
                    
                    # Agrupar por concepto
                    gastos_anterior_por_concepto = defaultdict(lambda: Decimal('0.00'))
                    total_anterior_c = Decimal('0.00')
                    
                    for reembolso in reembolsos_anterior:
                        concepto = reembolso.concepto.strip().upper()
                        monto_c = reembolso.monto_cordobas or Decimal('0.00')
                        gastos_anterior_por_concepto[concepto] += monto_c
                        total_anterior_c += monto_c
                    
                    # Calcular variaciones por concepto
                    for concepto, datos in gastos_por_concepto.items():
                        total_actual = datos['total_cordobas']
                        total_prev = gastos_anterior_por_concepto.get(concepto, Decimal('0.00'))
                        
                        variacion_c = total_actual - total_prev
                        variacion_d = variacion_c / tipo_cambio if variacion_c != 0 else Decimal('0.00')
                        
                        if total_prev > 0:
                            porcentaje_variacion = ((total_actual - total_prev) / total_prev) * 100
                        else:
                            porcentaje_variacion = 100 if total_actual > 0 else 0
                        
                        datos['total_anterior_cordobas'] = total_prev
                        datos['total_anterior_dolares'] = total_prev / tipo_cambio if total_prev > 0 else Decimal('0.00')
                        datos['variacion_cordobas'] = variacion_c
                        datos['variacion_dolares'] = variacion_d
                        datos['porcentaje_variacion'] = porcentaje_variacion
                    
                    # Variación total
                    variacion_total_c = total_general_c - total_anterior_c
                    variacion_total_d = variacion_total_c / tipo_cambio if variacion_total_c != 0 else Decimal('0.00')
                    porcentaje_variacion_total = ((total_general_c - total_anterior_c) / total_anterior_c) * 100 if total_anterior_c > 0 else 100
                    
                    context['total_anterior_cordobas'] = total_anterior_c
                    context['total_anterior_dolares'] = total_anterior_c / tipo_cambio if total_anterior_c > 0 else Decimal('0.00')
                    context['variacion_total_cordobas'] = variacion_total_c
                    context['variacion_total_dolares'] = variacion_total_d
                    context['porcentaje_variacion_total'] = porcentaje_variacion_total
                
                # Convertir defaultdict a lista para el template
                conceptos_data = []
                for concepto, datos in gastos_por_concepto.items():
                    porcentaje_participacion = (datos['total_cordobas'] / total_general_c * 100) if total_general_c > 0 else 0
                    
                    conceptos_data.append({
                        'concepto': concepto,
                        'items': datos['items'],
                        'cantidad': datos['cantidad'],
                        'total_cordobas': datos['total_cordobas'],
                        'total_dolares': datos['total_dolares'],
                        'porcentaje_participacion': porcentaje_participacion,
                        'total_anterior_cordobas': datos.get('total_anterior_cordobas'),
                        'total_anterior_dolares': datos.get('total_anterior_dolares'),
                        'variacion_cordobas': datos.get('variacion_cordobas'),
                        'variacion_dolares': datos.get('variacion_dolares'),
                        'porcentaje_variacion': datos.get('porcentaje_variacion'),
                    })
                
                # Ordenar por monto descendente
                conceptos_data.sort(key=lambda x: x['total_cordobas'], reverse=True)
                
                context['conceptos_data'] = conceptos_data
                context['total_general_cordobas'] = total_general_c
                context['total_general_dolares'] = total_general_d
                context['tiene_datos'] = len(conceptos_data) > 0
                context['datos_reporte'] = True
                
            except Exception as e:
                context['error'] = f'Error al generar el reporte: {str(e)}'
        
        return context

class ReporteConsolidadoProyectosView(LoginRequiredMixin, TemplateView):
    """
    Vista para generar reporte consolidado de todos los proyectos
    
    Muestra:
    - Todos los proyectos con sus montos
    - Comparación entre periodos
    - Porcentaje de participación
    - Variaciones
    """
    template_name = 'reportes/reporte_consolidado.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Configuración de la empresa
        context['configuracion'] = ConfiguracionEmpresa.get_configuracion()
        
        # Tipo de cambio
        tipo_cambio = get_tipo_cambio_actual()
        context['tipo_cambio'] = tipo_cambio
        
        # Obtener parámetros de filtro
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        comparar = self.request.GET.get('comparar', 'no')  # Comparar con periodo anterior
        
        if fecha_inicio and fecha_fin:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                
                context['fecha_inicio'] = fecha_inicio
                context['fecha_fin'] = fecha_fin
                context['comparar'] = comparar
                
                # Obtener todos los proyectos activos
                proyectos = Proyecto.objects.filter(eliminado=False).order_by('nombre')
                
                datos_proyectos = []
                total_general_cordobas = Decimal('0.00')
                total_general_dolares = Decimal('0.00')
                
                for proyecto in proyectos:
                    # ====================================
                    # PERIODO ACTUAL
                    # ====================================
                    
                    # Planillas de trabajadores
                    planillas_trabajadores = Planilla.objects.filter(
                        proyecto=proyecto,
                        estado='pagada',
                        fecha_generacion__gte=fecha_inicio_obj,
                        fecha_generacion__lte=fecha_fin_obj,
                        eliminado=False
                    )
                    
                    total_trabajadores_c = DetallePlanilla.objects.filter(
                        planilla__in=planillas_trabajadores
                    ).aggregate(
                        total=Sum('ingreso_total')
                    )['total'] or Decimal('0.00')
                    
                    # Planillas de contratistas
                    planillas_contratistas = PlanillaContratista.objects.filter(
                        proyecto=proyecto,
                        estado='pagada',
                        fecha_generacion__gte=fecha_inicio_obj,
                        fecha_generacion__lte=fecha_fin_obj,
                        eliminado=False
                    )
                    
                    total_contratistas_c = DetallePlanillaContratista.objects.filter(
                        planilla__in=planillas_contratistas
                    ).aggregate(
                        total=Sum('monto_cordobas')
                    )['total'] or Decimal('0.00')
                    
                    # Total del proyecto en periodo actual
                    total_proyecto_c = total_trabajadores_c + total_contratistas_c
                    total_proyecto_d = total_proyecto_c / tipo_cambio if total_proyecto_c > 0 else Decimal('0.00')
                    
                    # ====================================
                    # PERIODO ANTERIOR (si se solicita comparación)
                    # ====================================
                    variacion_c = None
                    variacion_d = None
                    porcentaje_variacion = None
                    
                    if comparar == 'si':
                        # Calcular fechas del periodo anterior
                        dias_periodo = (fecha_fin_obj - fecha_inicio_obj).days + 1
                        fecha_inicio_anterior = fecha_inicio_obj - timedelta(days=dias_periodo)
                        fecha_fin_anterior = fecha_inicio_obj - timedelta(days=1)
                        
                        # Planillas de trabajadores periodo anterior
                        planillas_trabajadores_ant = Planilla.objects.filter(
                            proyecto=proyecto,
                            estado='pagada',
                            fecha_generacion__gte=fecha_inicio_anterior,
                            fecha_generacion__lte=fecha_fin_anterior,
                            eliminado=False
                        )
                        
                        total_trabajadores_ant_c = DetallePlanilla.objects.filter(
                            planilla__in=planillas_trabajadores_ant
                        ).aggregate(
                            total=Sum('ingreso_total')
                        )['total'] or Decimal('0.00')
                        
                        # Planillas de contratistas periodo anterior
                        planillas_contratistas_ant = PlanillaContratista.objects.filter(
                            proyecto=proyecto,
                            estado='pagada',
                            fecha_generacion__gte=fecha_inicio_anterior,
                            fecha_generacion__lte=fecha_fin_anterior,
                            eliminado=False
                        )
                        
                        total_contratistas_ant_c = DetallePlanillaContratista.objects.filter(
                            planilla__in=planillas_contratistas_ant
                        ).aggregate(
                            total=Sum('monto_cordobas')
                        )['total'] or Decimal('0.00')
                        
                        total_anterior_c = total_trabajadores_ant_c + total_contratistas_ant_c
                        total_anterior_d = total_anterior_c / tipo_cambio if total_anterior_c > 0 else Decimal('0.00')
                        
                        # Calcular variación
                        variacion_c = total_proyecto_c - total_anterior_c
                        variacion_d = total_proyecto_d - total_anterior_d
                        
                        # Calcular porcentaje de variación
                        if total_anterior_c > 0:
                            porcentaje_variacion = ((total_proyecto_c - total_anterior_c) / total_anterior_c) * 100
                        else:
                            porcentaje_variacion = 100 if total_proyecto_c > 0 else 0
                    
                    # Solo incluir proyectos con actividad en el periodo
                    if total_proyecto_c > 0:
                        datos_proyectos.append({
                            'proyecto': proyecto,
                            'total_cordobas': total_proyecto_c,
                            'total_dolares': total_proyecto_d,
                            'variacion_cordobas': variacion_c,
                            'variacion_dolares': variacion_d,
                            'porcentaje_variacion': porcentaje_variacion,
                        })
                        
                        total_general_cordobas += total_proyecto_c
                        total_general_dolares += total_proyecto_d
                
                # Calcular porcentaje de participación
                for dato in datos_proyectos:
                    if total_general_cordobas > 0:
                        dato['porcentaje_participacion'] = (dato['total_cordobas'] / total_general_cordobas) * 100
                    else:
                        dato['porcentaje_participacion'] = 0
                
                context['datos_proyectos'] = datos_proyectos
                context['total_general_cordobas'] = total_general_cordobas
                context['total_general_dolares'] = total_general_dolares
                context['tiene_datos'] = len(datos_proyectos) > 0
                context['datos_reporte'] = True
                
            except Exception as e:
                context['error'] = f'Error al generar el reporte: {str(e)}'
        
        return context

class ReportePlanillaTotalView(LoginRequiredMixin, TemplateView):
    """
    Vista para generar REPORTE MAESTRO - Planilla Total Consolidada
    
    Integra TODOS los componentes del sistema:
    - Todos los proyectos de construcción
    - Planilla administrativa
    - Gastos varios y reembolsos
    
    Muestra variaciones y porcentaje de participación de cada componente
    """
    template_name = 'reportes/reporte_planilla_total.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Configuración
        context['configuracion'] = ConfiguracionEmpresa.get_configuracion()
        
        # Tipo de cambio
        tipo_cambio = get_tipo_cambio_actual()
        context['tipo_cambio'] = tipo_cambio
        
        # Obtener parámetros
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        comparar = self.request.GET.get('comparar', 'no')
        
        if fecha_inicio and fecha_fin:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                
                context['fecha_inicio'] = fecha_inicio
                context['fecha_fin'] = fecha_fin
                context['comparar'] = comparar
                
                # ==========================================
                # COMPONENTE 1: PROYECTOS DE CONSTRUCCIÓN
                # ==========================================
                
                # Buscar proyecto administrativo para excluirlo
                proyecto_admin = Proyecto.objects.filter(
                    Q(nombre__icontains='administra') | 
                    Q(nombre__icontains='general'),
                    eliminado=False
                ).first()
                
                # Proyectos de construcción (excluir administrativo)
                if proyecto_admin:
                    proyectos_construccion = Proyecto.objects.filter(
                        eliminado=False
                    ).exclude(id=proyecto_admin.id)
                else:
                    proyectos_construccion = Proyecto.objects.filter(eliminado=False)
                
                # Calcular total de proyectos
                total_proyectos_c = Decimal('0.00')
                total_proyectos_d = Decimal('0.00')
                
                for proyecto in proyectos_construccion:
                    # Trabajadores
                    planillas_trab = Planilla.objects.filter(
                        proyecto=proyecto,
                        estado='pagada',
                        fecha_generacion__gte=fecha_inicio_obj,
                        fecha_generacion__lte=fecha_fin_obj,
                        eliminado=False
                    )
                    
                    total_trab = DetallePlanilla.objects.filter(
                        planilla__in=planillas_trab
                    ).aggregate(total=Sum('ingreso_total'))['total'] or Decimal('0.00')
                    
                    # Contratistas
                    planillas_cont = PlanillaContratista.objects.filter(
                        proyecto=proyecto,
                        estado='pagada',
                        fecha_generacion__gte=fecha_inicio_obj,
                        fecha_generacion__lte=fecha_fin_obj,
                        eliminado=False
                    )
                    
                    total_cont = DetallePlanillaContratista.objects.filter(
                        planilla__in=planillas_cont
                    ).aggregate(total=Sum('monto_cordobas'))['total'] or Decimal('0.00')
                    
                    total_proyectos_c += (total_trab + total_cont)
                
                total_proyectos_d = total_proyectos_c / tipo_cambio if total_proyectos_c > 0 else Decimal('0.00')
                
                # ==========================================
                # COMPONENTE 2: PLANILLA ADMINISTRATIVA
                # ==========================================
                
                total_administrativa_c = Decimal('0.00')
                total_administrativa_d = Decimal('0.00')
                
                if proyecto_admin:
                    planillas_admin = Planilla.objects.filter(
                        proyecto=proyecto_admin,
                        estado='pagada',
                        fecha_generacion__gte=fecha_inicio_obj,
                        fecha_generacion__lte=fecha_fin_obj,
                        eliminado=False
                    )
                    
                    total_administrativa_c = DetallePlanilla.objects.filter(
                        planilla__in=planillas_admin
                    ).aggregate(total=Sum('ingreso_total'))['total'] or Decimal('0.00')
                    
                    total_administrativa_d = total_administrativa_c / tipo_cambio if total_administrativa_c > 0 else Decimal('0.00')
                
                # ==========================================
                # COMPONENTE 3: GASTOS VARIOS Y REEMBOLSOS
                # ==========================================
                
                planillas_todas = Planilla.objects.filter(
                    estado='pagada',
                    fecha_generacion__gte=fecha_inicio_obj,
                    fecha_generacion__lte=fecha_fin_obj,
                    eliminado=False
                )
                
                total_gastos_varios_c = PlanillaReembolso.objects.filter(
                    planilla__in=planillas_todas
                ).aggregate(total=Sum('monto_cordobas'))['total'] or Decimal('0.00')
                
                total_gastos_varios_d = total_gastos_varios_c / tipo_cambio if total_gastos_varios_c > 0 else Decimal('0.00')
                
                # ==========================================
                # TOTALES GENERALES
                # ==========================================
                
                gran_total_c = total_proyectos_c + total_administrativa_c + total_gastos_varios_c
                gran_total_d = gran_total_c / tipo_cambio if gran_total_c > 0 else Decimal('0.00')
                
                # Porcentajes de participación
                if gran_total_c > 0:
                    porc_proyectos = (total_proyectos_c / gran_total_c) * 100
                    porc_administrativa = (total_administrativa_c / gran_total_c) * 100
                    porc_gastos = (total_gastos_varios_c / gran_total_c) * 100
                else:
                    porc_proyectos = porc_administrativa = porc_gastos = 0
                
                # ==========================================
                # COMPARACIÓN CON PERIODO ANTERIOR (opcional)
                # ==========================================
                
                if comparar == 'si':
                    dias_periodo = (fecha_fin_obj - fecha_inicio_obj).days + 1
                    fecha_inicio_anterior = fecha_inicio_obj - timedelta(days=dias_periodo)
                    fecha_fin_anterior = fecha_inicio_obj - timedelta(days=1)
                    
                    # PROYECTOS ANTERIOR
                    total_proy_ant_c = Decimal('0.00')
                    for proyecto in proyectos_construccion:
                        planillas_trab_ant = Planilla.objects.filter(
                            proyecto=proyecto,
                            estado='pagada',
                            fecha_generacion__gte=fecha_inicio_anterior,
                            fecha_generacion__lte=fecha_fin_anterior,
                            eliminado=False
                        )
                        
                        total_trab_ant = DetallePlanilla.objects.filter(
                            planilla__in=planillas_trab_ant
                        ).aggregate(total=Sum('ingreso_total'))['total'] or Decimal('0.00')
                        
                        planillas_cont_ant = PlanillaContratista.objects.filter(
                            proyecto=proyecto,
                            estado='pagada',
                            fecha_generacion__gte=fecha_inicio_anterior,
                            fecha_generacion__lte=fecha_fin_anterior,
                            eliminado=False
                        )
                        
                        total_cont_ant = DetallePlanillaContratista.objects.filter(
                            planilla__in=planillas_cont_ant
                        ).aggregate(total=Sum('monto_cordobas'))['total'] or Decimal('0.00')
                        
                        total_proy_ant_c += (total_trab_ant + total_cont_ant)
                    
                    # ADMINISTRATIVA ANTERIOR
                    total_admin_ant_c = Decimal('0.00')
                    if proyecto_admin:
                        planillas_admin_ant = Planilla.objects.filter(
                            proyecto=proyecto_admin,
                            estado='pagada',
                            fecha_generacion__gte=fecha_inicio_anterior,
                            fecha_generacion__lte=fecha_fin_anterior,
                            eliminado=False
                        )
                        
                        total_admin_ant_c = DetallePlanilla.objects.filter(
                            planilla__in=planillas_admin_ant
                        ).aggregate(total=Sum('ingreso_total'))['total'] or Decimal('0.00')
                    
                    # GASTOS VARIOS ANTERIOR
                    planillas_todas_ant = Planilla.objects.filter(
                        estado='pagada',
                        fecha_generacion__gte=fecha_inicio_anterior,
                        fecha_generacion__lte=fecha_fin_anterior,
                        eliminado=False
                    )
                    
                    total_gastos_ant_c = PlanillaReembolso.objects.filter(
                        planilla__in=planillas_todas_ant
                    ).aggregate(total=Sum('monto_cordobas'))['total'] or Decimal('0.00')
                    
                    # Gran total anterior
                    gran_total_ant_c = total_proy_ant_c + total_admin_ant_c + total_gastos_ant_c
                    
                    # Calcular variaciones
                    var_proyectos_c = total_proyectos_c - total_proy_ant_c
                    var_administrativa_c = total_administrativa_c - total_admin_ant_c
                    var_gastos_c = total_gastos_varios_c - total_gastos_ant_c
                    var_total_c = gran_total_c - gran_total_ant_c
                    
                    # Porcentajes de variación
                    porc_var_proyectos = ((var_proyectos_c / total_proy_ant_c) * 100) if total_proy_ant_c > 0 else (100 if total_proyectos_c > 0 else 0)
                    porc_var_administrativa = ((var_administrativa_c / total_admin_ant_c) * 100) if total_admin_ant_c > 0 else (100 if total_administrativa_c > 0 else 0)
                    porc_var_gastos = ((var_gastos_c / total_gastos_ant_c) * 100) if total_gastos_ant_c > 0 else (100 if total_gastos_varios_c > 0 else 0)
                    porc_var_total = ((var_total_c / gran_total_ant_c) * 100) if gran_total_ant_c > 0 else (100 if gran_total_c > 0 else 0)
                    
                    context.update({
                        'total_proy_ant_c': total_proy_ant_c,
                        'total_admin_ant_c': total_admin_ant_c,
                        'total_gastos_ant_c': total_gastos_ant_c,
                        'gran_total_ant_c': gran_total_ant_c,
                        'var_proyectos_c': var_proyectos_c,
                        'var_administrativa_c': var_administrativa_c,
                        'var_gastos_c': var_gastos_c,
                        'var_total_c': var_total_c,
                        'porc_var_proyectos': porc_var_proyectos,
                        'porc_var_administrativa': porc_var_administrativa,
                        'porc_var_gastos': porc_var_gastos,
                        'porc_var_total': porc_var_total,
                    })
                
                # Componentes para el template
                componentes = [
                    {
                        'nombre': 'Proyectos de Construcción',
                        'total_cordobas': total_proyectos_c,
                        'total_dolares': total_proyectos_d,
                        'porcentaje': porc_proyectos,
                    },
                    {
                        'nombre': 'Planilla Administrativa',
                        'total_cordobas': total_administrativa_c,
                        'total_dolares': total_administrativa_d,
                        'porcentaje': porc_administrativa,
                    },
                    {
                        'nombre': 'Gastos Varios y Reembolsos',
                        'total_cordobas': total_gastos_varios_c,
                        'total_dolares': total_gastos_varios_d,
                        'porcentaje': porc_gastos,
                    },
                ]
                
                context.update({
                    'componentes': componentes,
                    'gran_total_cordobas': gran_total_c,
                    'gran_total_dolares': gran_total_d,
                    'tiene_datos': gran_total_c > 0,
                    'datos_reporte': True,
                })
                
            except Exception as e:
                context['error'] = f'Error al generar el reporte: {str(e)}'
        
        return context

class ExportarReporteProyectoExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Exportar reporte por proyecto a Excel con múltiples hojas:
    - RESUMEN: Totales por área
    - PLANILLA: Detalle de trabajadores (oficiales + ayudantes)
    - SUB-CONTRATISTAS: Detalle de contratistas/avalúos
    """
    permission_modulo = 'reportes'
    permission_accion = 'exportar'

    def get(self, request, *args, **kwargs):
        proyecto_id = request.GET.get('proyecto')
        planilla_id = request.GET.get('planilla_id')
        tipo_planilla = request.GET.get('tipo_planilla')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        if not proyecto_id and not planilla_id:
            return HttpResponse('Parámetros incompletos', status=400)
        
        try:
            config = ConfiguracionEmpresa.get_configuracion()
            tipo_cambio = get_tipo_cambio_actual()
            
            # Determinar proyecto
            if planilla_id and tipo_planilla:
                if tipo_planilla == 'trabajadores':
                    planilla_ref = Planilla.objects.get(id=planilla_id, eliminado=False)
                else:
                    planilla_ref = PlanillaContratista.objects.get(id=planilla_id, eliminado=False)
                proyecto = planilla_ref.proyecto
                periodo_inicio = planilla_ref.periodo_inicio
                periodo_fin = planilla_ref.periodo_fin
            else:
                proyecto = Proyecto.objects.get(id=proyecto_id, eliminado=False)
                periodo_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date() if fecha_inicio else None
                periodo_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date() if fecha_fin else None
            
            # =============================================
            # OBTENER DATOS DE TRABAJADORES
            # =============================================
            planillas_trab_qs = Planilla.objects.filter(
                proyecto=proyecto, estado='pagada', eliminado=False
            )
            if periodo_inicio and periodo_fin:
                planillas_trab_qs = planillas_trab_qs.filter(
                    periodo_inicio__gte=periodo_inicio, periodo_fin__lte=periodo_fin
                )
            
            detalles_trabajadores = DetallePlanilla.objects.filter(
                planilla__in=planillas_trab_qs
            ).select_related('trabajador')
            
            oficiales = []
            ayudantes = []
            
            for d in detalles_trabajadores:
                data = {
                    'nombre': d.trabajador.nombre_completo,
                    'cedula': d.trabajador.numero_cedula or '',
                    'cargo': d.cargo or 'N/A',
                    'area': d.area or '',
                    'dias_laborados': d.dias_laborados or 0,
                    'dias_feriados': d.dias_feriados or 0,
                    'horas_extras': float(d.horas_extras or 0),
                    'salario_dia_base': float(d.salario_dia_base or 0),
                    'valor_septimo': float(d.valor_septimo_dia or 0),
                    'salario_con_septimo': float(d.salario_diario_con_septimo or 0),
                    'valor_hora_base': float(d.valor_hora_base or 0),
                    'salario_devengado': float(d.salario_devengado or 0),
                    'salario_horas_extras': float(d.salario_horas_extras or 0),
                    'ingreso_feriado': float(d.ingreso_dia_feriado or 0),
                    'bonos': float(d.bonos or 0),
                    'combustible': float(d.combustible or 0),
                    'otros': float(d.otros_gastos or 0),
                    'deducciones': float(d.deducciones or 0),
                    'ingreso_total': float(d.ingreso_total or 0),
                    'total_dolares': float(d.ingreso_total / tipo_cambio) if d.ingreso_total else 0,
                }
                
                cargo_lower = (d.cargo or '').lower()
                area_lower = (d.area or '').lower()
                if any(w in cargo_lower for w in ['oficial', 'maestro', 'ingeniero', 'fontanero', 'electricista', 'soldador', 'albañil']):
                    oficiales.append(data)
                elif 'oficial' in area_lower:
                    oficiales.append(data)
                else:
                    ayudantes.append(data)
            
            # =============================================
            # OBTENER DATOS DE CONTRATISTAS
            # =============================================
            planillas_cont_qs = PlanillaContratista.objects.filter(
                proyecto=proyecto, estado='pagada', eliminado=False
            )
            if periodo_inicio and periodo_fin:
                planillas_cont_qs = planillas_cont_qs.filter(
                    periodo_inicio__gte=periodo_inicio, periodo_fin__lte=periodo_fin
                )
            
            detalles_contratistas = DetallePlanillaContratista.objects.filter(
                planilla__in=planillas_cont_qs
            ).select_related('avaluo__contrato__contratista')
            
            contratistas = []
            for d in detalles_contratistas:
                contratistas.append({
                    'nombre': d.avaluo.contrato.contratista.nombre_completo,
                    'cedula': d.avaluo.contrato.contratista.numero_cedula or '',
                    'descripcion': d.avaluo.concepto or 'N/A',
                    'monto_cordobas': float(d.monto_cordobas or 0),
                    'monto_dolares': float(d.monto_dolares or 0),
                })
            
            # =============================================
            # CALCULAR TOTALES
            # =============================================
            total_oficiales_c = sum(o['ingreso_total'] for o in oficiales)
            total_ayudantes_c = sum(a['ingreso_total'] for a in ayudantes)
            total_contratistas_c = sum(c['monto_cordobas'] for c in contratistas)
            
            total_oficiales_d = sum(o['total_dolares'] for o in oficiales)
            total_ayudantes_d = sum(a['total_dolares'] for a in ayudantes)
            total_contratistas_d = sum(c['monto_dolares'] for c in contratistas)
            
            gran_total_c = total_oficiales_c + total_ayudantes_c + total_contratistas_c
            gran_total_d = total_oficiales_d + total_ayudantes_d + total_contratistas_d
            
            # =============================================
            # CREAR EXCEL
            # =============================================
            wb = Workbook()
            wb.remove(wb.active)
            
            periodo_str = f'{periodo_inicio.strftime("%d/%m/%Y")} al {periodo_fin.strftime("%d/%m/%Y")}' if periodo_inicio and periodo_fin else 'Todas las planillas'
            
            # Hoja RESUMEN
            self._crear_hoja_resumen(wb, proyecto, periodo_str, tipo_cambio, {
                'Oficiales': (total_oficiales_c, total_oficiales_d),
                'Ayudantes': (total_ayudantes_c, total_ayudantes_d),
                'Sub-Contratistas': (total_contratistas_c, total_contratistas_d),
            }, gran_total_c, gran_total_d)
            
            # Hoja OFICIALES
            if oficiales:
                self._crear_hoja_planilla(wb, 'OFICIALES', 'PLANILLA OFICIALES', oficiales,
                                          total_oficiales_c, total_oficiales_d, proyecto, periodo_str, tipo_cambio)
            
            # Hoja AYUDANTES
            if ayudantes:
                self._crear_hoja_planilla(wb, 'AYUDANTES', 'PLANILLA DE AYUDANTES', ayudantes,
                                          total_ayudantes_c, total_ayudantes_d, proyecto, periodo_str, tipo_cambio)
            
            # Hoja SUB-CONTRATISTAS
            if contratistas:
                self._crear_hoja_contratistas(wb, contratistas, total_contratistas_c, total_contratistas_d,
                                              proyecto, periodo_str, tipo_cambio)
            
            # Respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'Planilla_Proyecto_{proyecto.nombre}.xlsx'.replace(' ', '_')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
            
        except (Planilla.DoesNotExist, PlanillaContratista.DoesNotExist, Proyecto.DoesNotExist):
            return HttpResponse('No encontrado', status=404)
        except Exception as e:
            return HttpResponse(f'Error: {str(e)}', status=500)
    
    def _estilos(self):
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
        border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        total_fill = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
        total_font = Font(bold=True, size=10, name='Arial', color='065F46')
        return header_fill, header_font, border, total_fill, total_font
    
    def _encabezado(self, ws, proyecto, periodo_str, tipo_cambio, subtitulo):
        self._agregar_logo(ws, 'A1')
        ws.merge_cells('A1:E1')
        ws['A1'] = f'PROYECTO: {proyecto.nombre.upper()}'
        ws['A1'].font = Font(bold=True, size=14, name='Arial', color='1F4788')
        
        ws.merge_cells('A2:E2')
        ws['A2'] = 'PLANILLA DE PAGOS'
        ws['A2'].font = Font(bold=True, size=12, name='Arial', color='1F4788')
        
        ws.merge_cells('A3:E3')
        ws['A3'] = subtitulo
        ws['A3'].font = Font(bold=True, size=11, name='Arial')
        
        ws['A4'] = 'FECHA:'
        ws['A4'].font = Font(bold=True, size=10, name='Arial')
        ws['B4'] = datetime.now().strftime('%d DE %B DE %Y').upper()
        ws['B4'].font = Font(size=10, name='Arial')
        
        ws['A5'] = 'PERIODO:'
        ws['A5'].font = Font(bold=True, size=10, name='Arial')
        ws['B5'] = periodo_str
        ws['B5'].font = Font(size=10, name='Arial')
        
        ws['A6'] = 'TIPO DE CAMBIO BCN:'
        ws['A6'].font = Font(bold=True, size=10, name='Arial')
        ws['D6'] = float(tipo_cambio)
        ws['D6'].font = Font(size=10, name='Arial')
    
    def _crear_hoja_resumen(self, wb, proyecto, periodo_str, tipo_cambio, totales, gran_total_c, gran_total_d):
        ws = wb.create_sheet('RESUMEN', 0)
        header_fill, header_font, border, total_fill, total_font = self._estilos()
        
        self._encabezado(ws, proyecto, periodo_str, tipo_cambio, '')
        
        # Tasa de cambio
        ws['A6'] = 'Tasa de Cambio'
        ws['C6'] = float(tipo_cambio)
        
        # Headers tabla
        row = 9
        headers = ['N°', 'Area', '', 'SALARIO TOTAL', 'TOTAL EN DOLAR PROYECTO']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        row = 10
        num = 1
        for area, (total_c, total_d) in totales.items():
            ws.cell(row=row, column=1, value=num).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=area).border = border
            ws.cell(row=row, column=3).border = border
            cell_c = ws.cell(row=row, column=4, value=total_c)
            cell_c.border = border
            cell_c.number_format = '#,##0.00'
            cell_d = ws.cell(row=row, column=5, value=total_d)
            cell_d.border = border
            cell_d.number_format = '#,##0.00'
            num += 1
            row += 1
        
        # Totales
        ws.cell(row=row, column=1, value='TOTAL CÓRDOBAS').font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=4, value=gran_total_c).font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = total_fill
        
        row += 1
        ws.cell(row=row, column=1, value='TOTAL DÓLARES').font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=4, value=gran_total_d).font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = total_fill
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
    
    def _crear_hoja_planilla(self, wb, nombre_hoja, subtitulo, datos, total_c, total_d, proyecto, periodo_str, tipo_cambio):
        ws = wb.create_sheet(nombre_hoja)
        header_fill, header_font, border, total_fill, total_font = self._estilos()
        
        self._encabezado(ws, proyecto, periodo_str, tipo_cambio, subtitulo)
        
        # Headers
        row = 8
        headers = ['N°', 'NOMBRE Y APELLIDO', 'CÉDULA', 'CARGO', 'DÍAS LAB.', 'DÍAS FER.',
                   'H. EXTRAS', 'DÍA BASE', 'VALOR 7MO', 'SAL. BASE+7MO', 'VALOR HORA',
                   'SAL. DEVENGADO', 'SAL. H.E.', 'INGR. FERIADO', 'BONO', 'COMBUSTIBLE',
                   'OTROS', 'DEDUCCIÓN', 'TOTAL C$', 'TOTAL USD', 'FIRMA']
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Datos
        row = 9
        for i, d in enumerate(datos, 1):
            vals = [
                i, d['nombre'], d['cedula'], d['cargo'], d['dias_laborados'], d['dias_feriados'],
                d['horas_extras'], d['salario_dia_base'], d['valor_septimo'], d['salario_con_septimo'],
                d['valor_hora_base'], d['salario_devengado'], d['salario_horas_extras'],
                d['ingreso_feriado'], d['bonos'], d['combustible'], d['otros'], d['deducciones'],
                d['ingreso_total'], d['total_dolares'], ''
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.font = Font(size=9, name='Arial')
                if col >= 8 and col <= 20:
                    cell.number_format = '#,##0.00'
                if col in [1, 5, 6]:
                    cell.alignment = Alignment(horizontal='center')
            row += 1
        
        # Total
        ws.cell(row=row, column=1, value='TOTAL:').font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=19, value=total_c).font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=19).number_format = '#,##0.00'
        ws.cell(row=row, column=20, value=total_d).font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=20).number_format = '#,##0.00'
        for col in range(1, 22):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = total_fill
        
        row += 1
        ws.cell(row=row, column=1, value='TOTAL DÓLARES:').font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=19, value=total_d).font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=19).number_format = '#,##0.00'
        
        # Anchos
        widths = [5, 30, 18, 18, 8, 8, 8, 12, 12, 14, 10, 14, 12, 12, 10, 12, 10, 12, 14, 14, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    
    def _crear_hoja_contratistas(self, wb, datos, total_c, total_d, proyecto, periodo_str, tipo_cambio):
        ws = wb.create_sheet('SUB-CONTRATISTA')
        header_fill, header_font, border, total_fill, total_font = self._estilos()
        
        self._encabezado(ws, proyecto, periodo_str, tipo_cambio, 'SUBCONTRATISTAS')
        
        # Headers
        row = 8
        headers = ['N°', 'NOMBRE Y APELLIDO', 'CÉDULA', 'DESCRIPCIÓN', 'PAGO C$', 'PAGO USD', 'FIRMA']
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        row = 9
        for i, d in enumerate(datos, 1):
            vals = [i, d['nombre'], d['cedula'], d['descripcion'], d['monto_cordobas'], d['monto_dolares'], '']
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.font = Font(size=9, name='Arial')
                if col in [5, 6]:
                    cell.number_format = '#,##0.00'
            row += 1
        
        # Total
        ws.cell(row=row, column=1, value='TOTAL EN CÓRDOBAS').font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=5, value=total_c).font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = total_fill
        
        row += 1
        ws.cell(row=row, column=1, value='TOTAL EN DÓLARES').font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=5, value=total_d).font = Font(bold=True, name='Arial')
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = total_fill
        
        # Anchos
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
    
    def _agregar_logo(self, ws, celda='A1'):
        try:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_quadycons.png')
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                img.width = 100
                img.height = 100
                ws.add_image(img, celda)
                ws.row_dimensions[1].height = 40
        except:
            pass

# ============================================================================
# 1. EXPORTAR CONSOLIDADO DE PROYECTOS
# ============================================================================

class ExportarConsolidadoExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Exportar consolidado de proyectos a Excel"""
    permission_modulo = 'reportes'
    permission_accion = 'exportar'

    def get(self, request, *args, **kwargs):
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return HttpResponse('Faltan parámetros', status=400)
        
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            
            wb = Workbook()
            wb.remove(wb.active)
            
            config = ConfiguracionEmpresa.get_configuracion()
            tipo_cambio = get_tipo_cambio_actual()
            
            # Obtener datos
            proyectos = Proyecto.objects.filter(eliminado=False).order_by('nombre')
            datos_proyectos = []
            total_general_c = Decimal('0.00')
            
            for proyecto in proyectos:
                planillas_trab = Planilla.objects.filter(
                    proyecto=proyecto, estado='pagada',
                    fecha_generacion__gte=fecha_inicio_obj,
                    fecha_generacion__lte=fecha_fin_obj, eliminado=False
                )
                
                total_trab = DetallePlanilla.objects.filter(
                    planilla__in=planillas_trab
                ).aggregate(total=Sum('ingreso_total'))['total'] or Decimal('0.00')
                
                planillas_cont = PlanillaContratista.objects.filter(
                    proyecto=proyecto, estado='pagada',
                    fecha_generacion__gte=fecha_inicio_obj,
                    fecha_generacion__lte=fecha_fin_obj, eliminado=False
                )
                
                total_cont = DetallePlanillaContratista.objects.filter(
                    planilla__in=planillas_cont
                ).aggregate(total=Sum('monto_cordobas'))['total'] or Decimal('0.00')
                
                total_proy = total_trab + total_cont
                if total_proy > 0:
                    datos_proyectos.append({
                        'proyecto': proyecto.nombre,
                        'total_c': total_proy,
                        'total_d': total_proy / tipo_cambio,
                    })
                    total_general_c += total_proy
            
            # Crear hoja
            ws = wb.create_sheet('Consolidado', 0)
            self._crear_hoja(ws, datos_proyectos, total_general_c, config, 
                           fecha_inicio, fecha_fin, tipo_cambio, 'Consolidado de Proyectos')
            
            # Respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'Consolidado_Proyectos_{fecha_inicio}_{fecha_fin}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
            
        except Exception as e:
            return HttpResponse(f'Error: {str(e)}', status=500)
    
    def _crear_hoja(self, ws, datos, total_c, config, fecha_inicio, fecha_fin, tipo_cambio, titulo):
        header_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Logo y título
        self._agregar_logo(ws, 'A1')
        ws.merge_cells('B1:D2')
        ws['B1'] = titulo
        ws['B1'].font = Font(bold=True, size=14)
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Info
        ws['A3'] = 'Periodo:'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = f'{fecha_inicio} al {fecha_fin}'
        ws['A4'] = 'Tipo de Cambio:'
        ws['A4'].font = Font(bold=True)
        ws['B4'] = f'C$ {tipo_cambio}'
        
        # Headers
        row = 6
        for col, header in enumerate(['Proyecto', 'Córdobas (C$)', 'Dólares (USD)', '% Part'], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        row = 7
        total_d = Decimal('0.00')
        for dato in datos:
            ws.cell(row=row, column=1, value=dato['proyecto']).border = border
            ws.cell(row=row, column=2, value=float(dato['total_c'])).border = border
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=float(dato['total_d'])).border = border
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            porc = (dato['total_c'] / total_c * 100) if total_c > 0 else 0
            ws.cell(row=row, column=4, value=float(porc)).border = border
            ws.cell(row=row, column=4).number_format = '0.00"%"'
            total_d += dato['total_d']
            row += 1
        
        # Total
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=2, value=float(total_c)).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=float(total_d)).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=100.0).number_format = '0.00"%"'
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def _agregar_logo(self, ws, celda='A1'):
        try:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_quadycons.png')
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                img.width = 100
                img.height = 100
                ws.add_image(img, celda)
                ws.row_dimensions[1].height = 75
                ws.row_dimensions[2].height = 10
                ws.column_dimensions['A'].width = 15
        except: pass


# ============================================================================
# 2. EXPORTAR PLANILLA ADMINISTRATIVA
# ============================================================================

class ExportarAdministrativaExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Exportar planilla administrativa a Excel"""
    permission_modulo = 'reportes'
    permission_accion = 'exportar'

    def get(self, request, *args, **kwargs):
        planilla_id = request.GET.get('planilla_id')
        
        if not planilla_id:
            return HttpResponse('Falta planilla_id', status=400)
        
        try:
            planilla = Planilla.objects.get(id=planilla_id, eliminado=False)
            tipo_cambio = get_tipo_cambio_actual()
            config = ConfiguracionEmpresa.get_configuracion()
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # Obtener detalles agrupados por cargo
            detalles = DetallePlanilla.objects.filter(
                planilla=planilla
            ).select_related('trabajador').order_by('cargo')
            
            areas = defaultdict(list)
            for detalle in detalles:
                area = detalle.cargo or 'Sin área'
                areas[area].append({
                    'nombre': detalle.trabajador.nombre_completo,
                    'cargo': detalle.cargo or 'N/A',
                    'dias': detalle.dias_laborados or 0,
                    'cordobas': detalle.ingreso_total or Decimal('0.00'),
                    'dolares': (detalle.ingreso_total / tipo_cambio) if detalle.ingreso_total else Decimal('0.00'),
                })
            
            # Crear hoja por área
            for area, personal in areas.items():
                ws = wb.create_sheet(area[:30])  # Max 30 chars
                self._crear_hoja_area(ws, area, personal, config, planilla, tipo_cambio)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'Administrativa_{planilla.periodo_inicio}_{planilla.periodo_fin}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
            
        except Planilla.DoesNotExist:
            return HttpResponse('Planilla no encontrada', status=404)
    
    def _crear_hoja_area(self, ws, area, personal, config, planilla, tipo_cambio):
        header_fill = PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Logo y título
        self._agregar_logo(ws, 'A1')
        ws.merge_cells('B1:E2')
        ws['B1'] = f'Planilla Administrativa - {area}'
        ws['B1'].font = Font(bold=True, size=14)
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Info
        ws['A3'] = 'Periodo:'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = f'{planilla.periodo_inicio} al {planilla.periodo_fin}'
        ws['A4'] = 'Tipo de Cambio:'
        ws['A4'].font = Font(bold=True)
        ws['B4'] = f'C$ {tipo_cambio}'
        
        # Headers
        row = 6
        for col, header in enumerate(['Nombre', 'Cargo', 'Días', 'Córdobas', 'Dólares'], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        row = 7
        total_c = Decimal('0.00')
        total_d = Decimal('0.00')
        for p in personal:
            ws.cell(row=row, column=1, value=p['nombre']).border = border
            ws.cell(row=row, column=2, value=p['cargo']).border = border
            ws.cell(row=row, column=3, value=p['dias']).border = border
            ws.cell(row=row, column=4, value=float(p['cordobas'])).border = border
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=float(p['dolares'])).border = border
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            total_c += p['cordobas']
            total_d += p['dolares']
            row += 1
        
        # Total
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=4, value=float(total_c)).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=float(total_d)).number_format = '#,##0.00'
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
    
    def _agregar_logo(self, ws, celda='A1'):
        try:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_quadycons.png')
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                img.width = 100
                img.height = 100
                ws.add_image(img, celda)
                ws.row_dimensions[1].height = 75
                ws.row_dimensions[2].height = 10
                ws.column_dimensions['A'].width = 15
        except: pass


# ============================================================================
# 3. EXPORTAR GASTOS VARIOS
# ============================================================================

class ExportarGastosVariosExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Exportar gastos varios a Excel"""
    permission_modulo = 'reportes'
    permission_accion = 'exportar'

    def get(self, request, *args, **kwargs):
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return HttpResponse('Faltan parámetros', status=400)
        
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            
            wb = Workbook()
            wb.remove(wb.active)
            
            config = ConfiguracionEmpresa.get_configuracion()
            tipo_cambio = get_tipo_cambio_actual()
            
            # Obtener gastos
            planillas = Planilla.objects.filter(
                estado='pagada',
                fecha_generacion__gte=fecha_inicio_obj,
                fecha_generacion__lte=fecha_fin_obj,
                eliminado=False
            )
            
            reembolsos = PlanillaReembolso.objects.filter(
                planilla__in=planillas
            ).order_by('concepto')
            
            # Agrupar por concepto
            conceptos = defaultdict(list)
            for r in reembolsos:
                concepto = r.concepto.strip().upper()
                conceptos[concepto].append({
                    'numero': r.numero_item,
                    'fecha': r.planilla.fecha_generacion,
                    'concepto': r.concepto,
                    'observaciones': r.observaciones or '',
                    'cordobas': r.monto_cordobas or Decimal('0.00'),
                    'dolares': (r.monto_cordobas / tipo_cambio) if r.monto_cordobas else Decimal('0.00'),
                })
            
            # Hoja resumen
            ws_resumen = wb.create_sheet('Resumen', 0)
            self._crear_resumen(ws_resumen, conceptos, config, fecha_inicio, fecha_fin, tipo_cambio)
            
            # Hoja por concepto
            for concepto, items in conceptos.items():
                ws = wb.create_sheet(concepto[:30])
                self._crear_hoja_concepto(ws, concepto, items, config, tipo_cambio)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'Gastos_Varios_{fecha_inicio}_{fecha_fin}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
            
        except Exception as e:
            return HttpResponse(f'Error: {str(e)}', status=500)
    
    def _crear_resumen(self, ws, conceptos, config, fecha_inicio, fecha_fin, tipo_cambio):
        header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        self._agregar_logo(ws, 'A1')
        ws.merge_cells('B1:D2')
        ws['B1'] = 'Gastos Varios - Resumen'
        ws['B1'].font = Font(bold=True, size=14)
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A3'] = 'Periodo:'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = f'{fecha_inicio} al {fecha_fin}'
        
        # Headers
        row = 6
        for col, header in enumerate(['Concepto', 'Cantidad', 'Córdobas', 'Dólares'], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        row = 7
        total_c = Decimal('0.00')
        total_d = Decimal('0.00')
        for concepto, items in conceptos.items():
            ws.cell(row=row, column=1, value=concepto).border = border
            ws.cell(row=row, column=2, value=len(items)).border = border
            subtotal_c = sum(i['cordobas'] for i in items)
            subtotal_d = sum(i['dolares'] for i in items)
            ws.cell(row=row, column=3, value=float(subtotal_c)).border = border
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=float(subtotal_d)).border = border
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            total_c += subtotal_c
            total_d += subtotal_d
            row += 1
        
        # Total
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=3, value=float(total_c)).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=float(total_d)).number_format = '#,##0.00'
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
    
    def _crear_hoja_concepto(self, ws, concepto, items, config, tipo_cambio):
        header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        self._agregar_logo(ws, 'A1')
        ws.merge_cells('B1:E2')
        ws['B1'] = concepto
        ws['B1'].font = Font(bold=True, size=14)
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        row = 6
        for col, header in enumerate(['N°', 'Fecha', 'Observaciones', 'Córdobas', 'Dólares'], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        row = 7
        total_c = Decimal('0.00')
        total_d = Decimal('0.00')
        for item in items:
            ws.cell(row=row, column=1, value=item['numero']).border = border
            ws.cell(row=row, column=2, value=item['fecha'].strftime('%d/%m/%Y')).border = border
            ws.cell(row=row, column=3, value=item['observaciones']).border = border
            ws.cell(row=row, column=4, value=float(item['cordobas'])).border = border
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=float(item['dolares'])).border = border
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            total_c += item['cordobas']
            total_d += item['dolares']
            row += 1
        
        # Total
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=4, value=float(total_c)).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=float(total_d)).number_format = '#,##0.00'
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
    
    def _agregar_logo(self, ws, celda='A1'):
        try:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_quadycons.png')
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                img.width = 100
                img.height = 100
                ws.add_image(img, celda)
                ws.row_dimensions[1].height = 75
                ws.row_dimensions[2].height = 10
                ws.column_dimensions['A'].width = 15
        except: pass


# ============================================================================
# 4. EXPORTAR PLANILLA TOTAL
# ============================================================================

class ExportarPlanillaTotalExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Exportar planilla total consolidada a Excel"""
    permission_modulo = 'reportes'
    permission_accion = 'exportar'
        
    def get(self, request, *args, **kwargs):
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return HttpResponse('Faltan parámetros', status=400)
        
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            
            wb = Workbook()
            wb.remove(wb.active)
            
            config = ConfiguracionEmpresa.get_configuracion()
            tipo_cambio = get_tipo_cambio_actual()
            
            # Calcular componentes (igual que la vista)
            proyecto_admin = Proyecto.objects.filter(
                Q(nombre__icontains='administra') | Q(nombre__icontains='general'),
                eliminado=False
            ).first()
            
            proyectos_construccion = Proyecto.objects.filter(eliminado=False)
            if proyecto_admin:
                proyectos_construccion = proyectos_construccion.exclude(id=proyecto_admin.id)
            
            # Proyectos
            total_proyectos_c = Decimal('0.00')
            for proyecto in proyectos_construccion:
                planillas = Planilla.objects.filter(
                    proyecto=proyecto, estado='pagada',
                    fecha_generacion__gte=fecha_inicio_obj,
                    fecha_generacion__lte=fecha_fin_obj, eliminado=False
                )
                total_proyectos_c += DetallePlanilla.objects.filter(
                    planilla__in=planillas
                ).aggregate(t=Sum('ingreso_total'))['t'] or Decimal('0.00')
                
                planillas_cont = PlanillaContratista.objects.filter(
                    proyecto=proyecto, estado='pagada',
                    fecha_generacion__gte=fecha_inicio_obj,
                    fecha_generacion__lte=fecha_fin_obj, eliminado=False
                )
                total_proyectos_c += DetallePlanillaContratista.objects.filter(
                    planilla__in=planillas_cont
                ).aggregate(t=Sum('monto_cordobas'))['t'] or Decimal('0.00')
            
            # Administrativa
            total_admin_c = Decimal('0.00')
            if proyecto_admin:
                planillas_admin = Planilla.objects.filter(
                    proyecto=proyecto_admin, estado='pagada',
                    fecha_generacion__gte=fecha_inicio_obj,
                    fecha_generacion__lte=fecha_fin_obj, eliminado=False
                )
                total_admin_c = DetallePlanilla.objects.filter(
                    planilla__in=planillas_admin
                ).aggregate(t=Sum('ingreso_total'))['t'] or Decimal('0.00')
            
            # Gastos
            planillas_todas = Planilla.objects.filter(
                estado='pagada',
                fecha_generacion__gte=fecha_inicio_obj,
                fecha_generacion__lte=fecha_fin_obj, eliminado=False
            )
            total_gastos_c = PlanillaReembolso.objects.filter(
                planilla__in=planillas_todas
            ).aggregate(t=Sum('monto_cordobas'))['t'] or Decimal('0.00')
            
            # Crear hoja
            ws = wb.create_sheet('Planilla Total', 0)
            self._crear_hoja(ws, total_proyectos_c, total_admin_c, total_gastos_c,
                           config, fecha_inicio, fecha_fin, tipo_cambio)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'Planilla_Total_{fecha_inicio}_{fecha_fin}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
            
        except Exception as e:
            return HttpResponse(f'Error: {str(e)}', status=500)
    
    def _crear_hoja(self, ws, total_proy, total_admin, total_gastos, config, fecha_inicio, fecha_fin, tipo_cambio):
        header_fill = PatternFill(start_color='8b5cf6', end_color='8b5cf6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        self._agregar_logo(ws, 'A1')
        ws.merge_cells('B1:D2')
        ws['B1'] = 'Planilla Total Consolidada'
        ws['B1'].font = Font(bold=True, size=14)
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A3'] = 'Periodo:'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = f'{fecha_inicio} al {fecha_fin}'
        ws['A4'] = 'Tipo de Cambio:'
        ws['A4'].font = Font(bold=True)
        ws['B4'] = f'C$ {tipo_cambio}'
        
        # Headers
        row = 6
        for col, header in enumerate(['Componente', 'Córdobas (C$)', 'Dólares (USD)', '% Part'], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        gran_total = total_proy + total_admin + total_gastos
        componentes = [
            ('Proyectos de Construcción', total_proy),
            ('Planilla Administrativa', total_admin),
            ('Gastos Varios y Reembolsos', total_gastos),
        ]
        
        row = 7
        for nombre, monto_c in componentes:
            monto_d = monto_c / tipo_cambio if monto_c > 0 else Decimal('0.00')
            porc = (monto_c / gran_total * 100) if gran_total > 0 else 0
            
            ws.cell(row=row, column=1, value=nombre).border = border
            ws.cell(row=row, column=2, value=float(monto_c)).border = border
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=float(monto_d)).border = border
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=float(porc)).border = border
            ws.cell(row=row, column=4).number_format = '0.00"%"'
            row += 1
        
        # Total
        gran_total_d = gran_total / tipo_cambio if gran_total > 0 else Decimal('0.00')
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=1, value='GRAN TOTAL')
        ws.cell(row=row, column=2, value=float(gran_total)).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=float(gran_total_d)).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=100.0).number_format = '0.00"%"'
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 18
    
    def _agregar_logo(self, ws, celda='A1'):
        try:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_quadycons.png')
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                img.width = 100
                img.height = 100
                ws.add_image(img, celda)
                ws.row_dimensions[1].height = 75
                ws.row_dimensions[2].height = 10
                ws.column_dimensions['A'].width = 15
        except: pass
