"""
Vistas para el módulo de Gastos y Reembolsos
apps/reportes/views_gastos.py
"""
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q
from decimal import Decimal

from apps.admin_panel.permissions import PermissionRequiredMixin
from .models import ReporteGastos, DetalleReporteGastos
from apps.planillas.models import TipoCambio


class ReporteGastosListView(LoginRequiredMixin, View):
    """Lista de reportes de gastos y reembolsos"""
    template_name = 'reportes/gastos_reembolsos_lista.html'
    
    def get(self, request):
        reportes = ReporteGastos.objects.filter(eliminado=False)
        
        # Filtros
        estado = request.GET.get('estado')
        if estado and estado != 'todos':
            reportes = reportes.filter(estado=estado)
        
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        if fecha_desde:
            reportes = reportes.filter(periodo_inicio__gte=fecha_desde)
        if fecha_hasta:
            reportes = reportes.filter(periodo_fin__lte=fecha_hasta)
        
        # Totales para cards
        from django.db.models import Sum
        total_cordobas = reportes.aggregate(total=Sum('total_cordobas'))['total'] or Decimal('0.00')
        tipo_cambio = TipoCambio.get_actual()
        total_dolares = (total_cordobas / tipo_cambio.valor).quantize(Decimal('0.01')) if tipo_cambio.valor > 0 else Decimal('0.00')
        
        context = {
            'reportes': reportes,
            'total_reportes': reportes.count(),
            'total_cordobas': total_cordobas,
            'total_dolares': total_dolares,
            'filtros': {
                'estado': estado or 'todos',
                'fecha_desde': fecha_desde or '',
                'fecha_hasta': fecha_hasta or '',
            },
        }
        return render(request, self.template_name, context)


class ReporteGastosCreateView(LoginRequiredMixin, View):
    """Crear nuevo reporte de gastos"""
    template_name = 'reportes/gastos_reembolsos_crear.html'
    
    def get(self, request):
        tipo_cambio = TipoCambio.get_actual()
        context = {
            'tipo_cambio': tipo_cambio,
        }
        return render(request, self.template_name, context)
    
    def post(self, request):
        try:
            periodo_inicio = request.POST.get('periodo_inicio')
            periodo_fin = request.POST.get('periodo_fin')
            
            if not periodo_inicio or not periodo_fin:
                messages.error(request, '❌ Debe indicar el período de inicio y fin.')
                return redirect('gastos_reembolsos_crear')
            
            tipo_cambio = TipoCambio.get_actual()
            
            reporte = ReporteGastos.objects.create(
                periodo_inicio=periodo_inicio,
                periodo_fin=periodo_fin,
                tipo_cambio=tipo_cambio.valor,
                generado_por=request.user,
                estado='borrador',
            )
            
            messages.success(request, f'✅ Reporte {reporte.codigo} creado. Agregue las líneas de gastos.')
            return redirect('gastos_reembolsos_detalle', pk=reporte.pk)
            
        except Exception as e:
            messages.error(request, f'❌ Error al crear reporte: {str(e)}')
            return redirect('gastos_reembolsos_crear')


class ReporteGastosDetalleView(LoginRequiredMixin, View):
    """Detalle del reporte con sus líneas de gastos"""
    template_name = 'reportes/gastos_reembolsos_detalle.html'
    
    def get(self, request, pk):
        reporte = get_object_or_404(ReporteGastos, pk=pk, eliminado=False)
        detalles = reporte.detalles.all()
        
        # Calcular porcentajes
        total = reporte.total_cordobas if reporte.total_cordobas > 0 else Decimal('1')
        detalles_con_porcentaje = []
        for d in detalles:
            porcentaje = ((d.monto_cordobas / total) * 100).quantize(Decimal('0.01')) if total > 0 and d.monto_cordobas > 0 else Decimal('0.00')
            detalles_con_porcentaje.append({
                'detalle': d,
                'porcentaje': porcentaje,
            })
        
        context = {
            'reporte': reporte,
            'detalles': detalles_con_porcentaje,
            'puede_editar': reporte.puede_editar,
        }
        return render(request, self.template_name, context)


class ReporteGastosAgregarLineaView(LoginRequiredMixin, View):
    """Agregar línea de gasto al reporte"""
    
    def post(self, request, pk):
        reporte = get_object_or_404(ReporteGastos, pk=pk, eliminado=False)
        
        if not reporte.puede_editar:
            messages.error(request, '❌ No se pueden agregar gastos a un reporte en este estado.')
            return redirect('gastos_reembolsos_detalle', pk=pk)
        
        concepto = request.POST.get('concepto', '').strip()
        monto_str = request.POST.get('monto_cordobas', '0')
        
        if not concepto:
            messages.error(request, '❌ El concepto es obligatorio.')
            return redirect('gastos_reembolsos_detalle', pk=pk)
        
        try:
            monto = Decimal(monto_str.replace(',', ''))
        except:
            messages.error(request, '❌ Monto inválido.')
            return redirect('gastos_reembolsos_detalle', pk=pk)
        
        detalle = DetalleReporteGastos(
            reporte=reporte,
            concepto=concepto,
            monto_cordobas=monto,
            observaciones=request.POST.get('observaciones', ''),
            creado_por=request.user,
        )
        
        # Archivo soporte
        if 'archivo_soporte' in request.FILES:
            archivo = request.FILES['archivo_soporte']
            if archivo.size > 10 * 1024 * 1024:
                messages.error(request, '❌ El archivo no puede superar los 10MB.')
                return redirect('gastos_reembolsos_detalle', pk=pk)
            if len(archivo.name) > 80:
                messages.error(request, f'❌ El nombre del archivo es demasiado largo ({len(archivo.name)} caracteres). Máximo 80.')
                return redirect('gastos_reembolsos_detalle', pk=pk)
            detalle.archivo_soporte = archivo
        
        detalle.save()
        messages.success(request, f'✅ Gasto "{concepto}" agregado.')
        return redirect('gastos_reembolsos_detalle', pk=pk)


class ReporteGastosEditarLineaView(LoginRequiredMixin, View):
    """Editar línea de gasto"""
    
    def post(self, request, pk, linea_pk):
        reporte = get_object_or_404(ReporteGastos, pk=pk, eliminado=False)
        detalle = get_object_or_404(DetalleReporteGastos, pk=linea_pk, reporte=reporte)
        
        if not reporte.puede_editar:
            messages.error(request, '❌ No se puede editar en este estado.')
            return redirect('gastos_reembolsos_detalle', pk=pk)
        
        concepto = request.POST.get('concepto', '').strip()
        monto_str = request.POST.get('monto_cordobas', '0')
        
        if concepto:
            detalle.concepto = concepto
        
        try:
            detalle.monto_cordobas = Decimal(monto_str.replace(',', ''))
        except:
            pass
        
        detalle.observaciones = request.POST.get('observaciones', detalle.observaciones)
        
        if 'archivo_soporte' in request.FILES:
            archivo = request.FILES['archivo_soporte']
            if archivo.size <= 10 * 1024 * 1024 and len(archivo.name) <= 80:
                detalle.archivo_soporte = archivo
        
        detalle.save()
        messages.success(request, f'✅ Gasto actualizado.')
        return redirect('gastos_reembolsos_detalle', pk=pk)


class ReporteGastosEliminarLineaView(LoginRequiredMixin, View):
    """Eliminar línea de gasto"""
    
    def post(self, request, pk, linea_pk):
        reporte = get_object_or_404(ReporteGastos, pk=pk, eliminado=False)
        detalle = get_object_or_404(DetalleReporteGastos, pk=linea_pk, reporte=reporte)
        
        if not reporte.puede_editar:
            messages.error(request, '❌ No se puede eliminar en este estado.')
            return redirect('gastos_reembolsos_detalle', pk=pk)
        
        concepto = detalle.concepto
        detalle.delete()
        
        # Renumerar items
        for i, d in enumerate(reporte.detalles.all().order_by('numero_item'), 1):
            if d.numero_item != i:
                DetalleReporteGastos.objects.filter(pk=d.pk).update(numero_item=i)
        
        reporte.calcular_totales()
        messages.success(request, f'✅ Gasto "{concepto}" eliminado.')
        return redirect('gastos_reembolsos_detalle', pk=pk)


class ReporteGastosAprobarGerenteView(LoginRequiredMixin, View):
    """Aprobar reporte como gerente"""
    
    def post(self, request, pk):
        reporte = get_object_or_404(ReporteGastos, pk=pk, eliminado=False)
        
        if reporte.estado != 'borrador':
            messages.error(request, '❌ Solo se pueden aprobar reportes en estado borrador.')
            return redirect('gastos_reembolsos_detalle', pk=pk)
        
        reporte.estado = 'aprobada_gerente'
        reporte.aprobada_gerente_por = request.user
        reporte.fecha_aprobacion_gerente = timezone.now()
        reporte.save()
        
        messages.success(request, f'✅ Reporte {reporte.codigo} aprobado por gerente.')
        return redirect('gastos_reembolsos_detalle', pk=pk)


class ReporteGastosAprobarContadorView(LoginRequiredMixin, View):
    """Aprobar reporte como contador"""
    
    def post(self, request, pk):
        reporte = get_object_or_404(ReporteGastos, pk=pk, eliminado=False)
        
        if reporte.estado not in ['borrador', 'aprobada_gerente']:
            messages.error(request, '❌ No se puede aprobar en este estado.')
            return redirect('gastos_reembolsos_detalle', pk=pk)
        
        reporte.estado = 'aprobada_contador'
        reporte.aprobada_contador_por = request.user
        reporte.fecha_aprobacion_contador = timezone.now()
        reporte.save()
        
        messages.success(request, f'✅ Reporte {reporte.codigo} aprobado por contador.')
        return redirect('gastos_reembolsos_detalle', pk=pk)


class ReporteGastosMarcarPagadaView(LoginRequiredMixin, View):
    """Marcar reporte como pagado"""
    
    def post(self, request, pk):
        reporte = get_object_or_404(ReporteGastos, pk=pk, eliminado=False)
        
        if reporte.estado != 'aprobada_contador':
            messages.error(request, '❌ Solo se pueden pagar reportes aprobados por contador.')
            return redirect('gastos_reembolsos_detalle', pk=pk)
        
        reporte.estado = 'pagada'
        reporte.pagada_por = request.user
        reporte.fecha_pago = timezone.now()
        reporte.save()
        
        messages.success(request, f'✅ Reporte {reporte.codigo} marcado como pagado.')
        return redirect('gastos_reembolsos_detalle', pk=pk)


class ReporteGastosAnularView(LoginRequiredMixin, View):
    """Anular reporte"""
    
    def post(self, request, pk):
        reporte = get_object_or_404(ReporteGastos, pk=pk, eliminado=False)
        
        if reporte.estado == 'pagada':
            messages.error(request, '❌ No se puede anular un reporte ya pagado.')
            return redirect('gastos_reembolsos_detalle', pk=pk)
        
        reporte.estado = 'anulada'
        reporte.save()
        
        messages.success(request, f'✅ Reporte {reporte.codigo} anulado.')
        return redirect('gastos_reembolsos_detalle', pk=pk)

class ReporteGastosExportarExcelView(LoginRequiredMixin, View):
    """Exportar reporte de gastos a Excel"""
    
    def get(self, request, pk):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        reporte = get_object_or_404(ReporteGastos, pk=pk, eliminado=False)
        detalles = reporte.detalles.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reembolso y Otros Gastos'
        
        # Estilos
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
        title_font = Font(bold=True, size=14, name='Arial', color='1F4788')
        subtitle_font = Font(bold=True, size=11, name='Arial', color='374151')
        normal_font = Font(size=10, name='Arial')
        bold_font = Font(bold=True, size=10, name='Arial')
        money_font = Font(size=10, name='Arial', color='059669')
        total_fill = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
        total_font = Font(bold=True, size=11, name='Arial', color='065F46')
        link_font = Font(size=9, name='Arial', color='3B82F6', underline='single')
        border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        center = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Anchos de columna
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 25
        
        # Fila 1-2: Título
        ws.merge_cells('A1:G1')
        ws['A1'] = 'QUADYCONS'
        ws['A1'].font = Font(bold=True, size=16, name='Arial', color='1F4788')
        ws['A1'].alignment = center
        
        ws.merge_cells('A2:G2')
        ws['A2'] = 'PLANILLA DE REEMBOLSO Y OTROS GASTOS'
        ws['A2'].font = title_font
        ws['A2'].alignment = center
        
        # Fila 3: Información
        ws.merge_cells('A3:D3')
        ws['A3'] = f'Código: {reporte.codigo}'
        ws['A3'].font = subtitle_font
        
        ws.merge_cells('E3:G3')
        ws['E3'] = f'Período: {reporte.periodo_inicio.strftime("%d/%m/%Y")} - {reporte.periodo_fin.strftime("%d/%m/%Y")}'
        ws['E3'].font = subtitle_font
        ws['E3'].alignment = Alignment(horizontal='right')
        
        # Fila 4: Tipo de cambio y generado por
        ws.merge_cells('A4:D4')
        ws['A4'] = f'Generado por: {reporte.generado_por.nombre_completo if reporte.generado_por else "—"}'
        ws['A4'].font = normal_font
        
        ws.merge_cells('E4:G4')
        ws['E4'] = f'Tipo de Cambio: C$ {reporte.tipo_cambio}'
        ws['E4'].font = normal_font
        ws['E4'].alignment = Alignment(horizontal='right')
        
        # Fila 6: Headers
        row = 6
        headers = ['No', 'CONCEPTO', 'MONTO CÓRDOBAS', 'MONTO DÓLARES', 'PORCENTAJE', 'OBSERVACIONES', 'SOPORTE']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        
        # Datos
        total_c = reporte.total_cordobas if reporte.total_cordobas > 0 else Decimal('1')
        base_url = request.build_absolute_uri('/')[:-1]  # ej: http://localhost:8009
        
        for i, detalle in enumerate(detalles, 1):
            row += 1
            porcentaje = ((detalle.monto_cordobas / total_c) * 100).quantize(Decimal('0.01')) if total_c > 0 and detalle.monto_cordobas > 0 else Decimal('0.00')
            
            # No
            cell = ws.cell(row=row, column=1, value=detalle.numero_item)
            cell.font = bold_font
            cell.alignment = center
            cell.border = border
            
            # Concepto
            cell = ws.cell(row=row, column=2, value=detalle.concepto)
            cell.font = normal_font
            cell.alignment = left_align
            cell.border = border
            
            # Monto Córdobas
            cell = ws.cell(row=row, column=3, value=float(detalle.monto_cordobas))
            cell.font = money_font
            cell.number_format = '#,##0.00'
            cell.alignment = right_align
            cell.border = border
            
            # Monto Dólares
            cell = ws.cell(row=row, column=4, value=float(detalle.monto_dolares))
            cell.font = Font(size=10, name='Arial', color='2563EB')
            cell.number_format = '#,##0.00'
            cell.alignment = right_align
            cell.border = border
            
            # Porcentaje
            cell = ws.cell(row=row, column=5, value=float(porcentaje) / 100)
            cell.font = normal_font
            cell.number_format = '0.00%'
            cell.alignment = center
            cell.border = border
            
            # Observaciones
            cell = ws.cell(row=row, column=6, value=detalle.observaciones or '')
            cell.font = normal_font
            cell.alignment = left_align
            cell.border = border
            
            # Soporte (URL)
            if detalle.archivo_soporte:
                soporte_url = f'{base_url}{detalle.archivo_soporte.url}'
                cell = ws.cell(row=row, column=7, value='Ver documento')
                cell.font = link_font
                cell.hyperlink = soporte_url
                cell.alignment = center
            else:
                cell = ws.cell(row=row, column=7, value='—')
                cell.font = Font(size=10, name='Arial', color='9CA3AF')
                cell.alignment = center
            cell.border = border
        
        # Fila Total
        row += 1
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = total_fill
            ws.cell(row=row, column=col).border = border
        
        ws.cell(row=row, column=2, value='TOTAL').font = total_font
        
        cell = ws.cell(row=row, column=3, value=float(reporte.total_cordobas))
        cell.font = total_font
        cell.number_format = '#,##0.00'
        cell.alignment = right_align
        
        cell = ws.cell(row=row, column=4, value=float(reporte.total_dolares))
        cell.font = Font(bold=True, size=11, name='Arial', color='2563EB')
        cell.number_format = '#,##0.00'
        cell.alignment = right_align
        
        cell = ws.cell(row=row, column=5, value=1.0)
        cell.font = total_font
        cell.number_format = '0.00%'
        cell.alignment = center
        
        # ============================================================
        # FIRMAS DIGITALES
        # ============================================================
        row += 3
        
        firmas = [
            ('Generado por:', reporte.generado_por.nombre_completo if reporte.generado_por else '', reporte.creado_en.strftime('%d/%m/%Y %H:%M') if reporte.creado_en else ''),
            ('Aprobado Gerente:', reporte.aprobada_gerente_por.nombre_completo if reporte.aprobada_gerente_por else 'Pendiente', reporte.fecha_aprobacion_gerente.strftime('%d/%m/%Y %H:%M') if reporte.fecha_aprobacion_gerente else ''),
            ('Aprobado Contador:', reporte.aprobada_contador_por.nombre_completo if reporte.aprobada_contador_por else 'Pendiente', reporte.fecha_aprobacion_contador.strftime('%d/%m/%Y %H:%M') if reporte.fecha_aprobacion_contador else ''),
        ]
        
        col_firmas = [1, 3, 5]  # Columnas A, C, E
        
        for i, (label, nombre, fecha) in enumerate(firmas):
            col = col_firmas[i]
            
            # Línea de firma
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            cell = ws.cell(row=row, column=col, value='_' * 35)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = center
            
            # Nombre
            ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
            cell = ws.cell(row=row + 1, column=col, value=nombre)
            cell.font = Font(name='Arial', size=9, bold=True)
            cell.alignment = center
            
            # Label
            ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1)
            cell = ws.cell(row=row + 2, column=col, value=label)
            cell.font = Font(name='Arial', size=8, color='666666')
            cell.alignment = center
            
            # Fecha
            if fecha:
                ws.merge_cells(start_row=row + 3, start_column=col, end_row=row + 3, end_column=col + 1)
                cell = ws.cell(row=row + 3, column=col, value=fecha)
                cell.font = Font(name='Arial', size=8, italic=True, color='9CA3AF')
                cell.alignment = center
        
        # Si está pagada, agregar firma de pago
        if reporte.fecha_pago and reporte.pagada_por:
            row += 5
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            cell = ws.cell(row=row, column=1, value='_' * 35)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = center
            
            ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=2)
            cell = ws.cell(row=row + 1, column=1, value=reporte.pagada_por.nombre_completo)
            cell.font = Font(name='Arial', size=9, bold=True)
            cell.alignment = center
            
            ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=2)
            cell = ws.cell(row=row + 2, column=1, value='Pagado por:')
            cell.font = Font(name='Arial', size=8, color='666666')
            cell.alignment = center
            
            ws.merge_cells(start_row=row + 3, start_column=1, end_row=row + 3, end_column=2)
            cell = ws.cell(row=row + 3, column=1, value=reporte.fecha_pago.strftime('%d/%m/%Y %H:%M'))
            cell.font = Font(name='Arial', size=8, italic=True, color='9CA3AF')
            cell.alignment = center
        
        # Generar respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Gastos_Reembolsos_{reporte.codigo}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

class ReporteGastosExportarPDFView(LoginRequiredMixin, View):
    """Exportar reporte de gastos a PDF"""
    
    def get(self, request, pk):
        import io
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        
        reporte = get_object_or_404(ReporteGastos, pk=pk, eliminado=False)
        detalles = reporte.detalles.all()
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=20,
            bottomMargin=20,
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'],
            fontSize=14, textColor=colors.HexColor('#1F4788'),
            spaceAfter=4, alignment=TA_CENTER, fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle', parent=styles['Heading2'],
            fontSize=10, textColor=colors.HexColor('#1F4788'),
            spaceAfter=4, alignment=TA_CENTER, fontName='Helvetica-Bold'
        )
        info_style = ParagraphStyle(
            'Info', parent=styles['Normal'],
            fontSize=8, textColor=colors.HexColor('#374151'),
            spaceAfter=2, fontName='Helvetica'
        )
        
        # Título
        elements.append(Paragraph('QUADYCONS', title_style))
        elements.append(Paragraph('PLANILLA DE REEMBOLSO Y OTROS GASTOS', subtitle_style))
        elements.append(Spacer(1, 4))
        
        # Info
        elements.append(Paragraph(
            f'Código: {reporte.codigo} &nbsp;&nbsp;&nbsp; '
            f'Período: {reporte.periodo_inicio.strftime("%d/%m/%Y")} - {reporte.periodo_fin.strftime("%d/%m/%Y")} &nbsp;&nbsp;&nbsp; '
            f'T/C: C$ {reporte.tipo_cambio}',
            info_style
        ))
        elements.append(Spacer(1, 10))
        
        # Headers tabla
        headers = ['No', 'CONCEPTO', 'MONTO C$', 'MONTO USD', '%', 'OBSERVACIONES', 'SOPORTE']
        
        total_c = reporte.total_cordobas if reporte.total_cordobas > 0 else Decimal('1')
        base_url = request.build_absolute_uri('/')[:-1]
        
        # Datos
        table_data = [headers]
        for detalle in detalles:
            porcentaje = ((detalle.monto_cordobas / total_c) * 100).quantize(Decimal('0.01')) if total_c > 0 and detalle.monto_cordobas > 0 else Decimal('0.00')
            if detalle.archivo_soporte:
                soporte_url = f'{base_url}{detalle.archivo_soporte.url}'
                soporte = Paragraph(f'<link href="{soporte_url}"><u>Ver documento</u></link>', ParagraphStyle('link', fontSize=7, textColor=colors.HexColor('#3B82F6'), fontName='Helvetica'))
            else:
                soporte = '—'
            
            table_data.append([
                str(detalle.numero_item),
                detalle.concepto,
                f'C$ {detalle.monto_cordobas:,.2f}',
                f'$ {detalle.monto_dolares:,.2f}',
                f'{porcentaje}%',
                detalle.observaciones or '—',
                soporte,
            ])
        
        # Fila total
        table_data.append([
            '', 'TOTAL',
            f'C$ {reporte.total_cordobas:,.2f}',
            f'$ {reporte.total_dolares:,.2f}',
            '100%', '', ''
        ])
        
        # Crear tabla
        col_widths = [0.4*inch, 2.8*inch, 1.2*inch, 1.2*inch, 0.7*inch, 2.2*inch, 0.7*inch]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            # Datos
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),
            ('ALIGN', (6, 1), (6, -1), 'CENTER'),
            # Total (última fila)
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F0FDF4')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (1, -1), (1, -1), colors.HexColor('#065F46')),
            ('TEXTCOLOR', (2, -1), (2, -1), colors.HexColor('#065F46')),
            ('TEXTCOLOR', (3, -1), (3, -1), colors.HexColor('#2563EB')),
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            # Colores alternados
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        elements.append(table)
        
        # Firmas
        elements.append(Spacer(1, 30))
        
        firmas_data = [
            ['____________________', '____________________', '____________________'],
            [
                reporte.generado_por.nombre_completo if reporte.generado_por else '',
                reporte.aprobada_gerente_por.nombre_completo if reporte.aprobada_gerente_por else 'Pendiente',
                reporte.aprobada_contador_por.nombre_completo if reporte.aprobada_contador_por else 'Pendiente',
            ],
            ['Generado por', 'Aprobado Gerente', 'Aprobado Contador'],
        ]
        
        # Agregar fechas si existen
        fechas = [
            reporte.creado_en.strftime('%d/%m/%Y %H:%M') if reporte.creado_en else '',
            reporte.fecha_aprobacion_gerente.strftime('%d/%m/%Y %H:%M') if reporte.fecha_aprobacion_gerente else '',
            reporte.fecha_aprobacion_contador.strftime('%d/%m/%Y %H:%M') if reporte.fecha_aprobacion_contador else '',
        ]
        firmas_data.append(fechas)
        
        firmas_table = Table(firmas_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
        firmas_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 1), 8),
            ('FONTSIZE', (0, 2), (-1, 2), 7),
            ('FONTSIZE', (0, 3), (-1, 3), 6),
            ('TEXTCOLOR', (0, 2), (-1, 2), colors.HexColor('#666666')),
            ('TEXTCOLOR', (0, 3), (-1, 3), colors.HexColor('#9CA3AF')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(firmas_table)
        
        # Estado
        elements.append(Spacer(1, 10))
        estado_style = ParagraphStyle(
            'Estado', parent=styles['Normal'],
            fontSize=7, textColor=colors.HexColor('#9CA3AF'),
            alignment=TA_CENTER, fontName='Helvetica-Oblique'
        )
        elements.append(Paragraph(
            f'Estado: {reporte.get_estado_display()} — Generado el {reporte.creado_en.strftime("%d/%m/%Y %H:%M")}',
            estado_style
        ))
        
        # Construir
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        filename = f'Gastos_Reembolsos_{reporte.codigo}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

