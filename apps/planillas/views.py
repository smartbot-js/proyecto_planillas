"""
Vistas del módulo de planillas
apps/planillas/views.py
"""
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.views.generic import ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, legal
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

import traceback
import io

from .models import Planilla, DetallePlanilla, TipoCambio, DiaFeriado
from .utils import (
    generar_planilla_administrativa,
    generar_planilla_desde_asistencias,
    validar_periodo_planilla,
    obtener_resumen_asistencias
)
from apps.proyectos.models import Proyecto
from apps.usuarios.models import Usuario
from apps.asistencias.models import Asistencia
from apps.core.utils import get_tipo_cambio_actual
from apps.admin_panel.permissions import PermissionRequiredMixin
from apps.trabajadores.models import Trabajador

class PlanillaListView(LoginRequiredMixin, ListView):
    """Vista para listar todas las planillas"""
    model = Planilla
    template_name = 'planillas/lista.html'
    context_object_name = 'planillas'
    paginate_by = 12
    
    def get_queryset(self):
        """Retorna el queryset filtrado de planillas (excluye eliminadas)"""
        queryset = Planilla.objects.filter(
            eliminado=False
        ).select_related(
            'proyecto',
            'generada_por',
            'aprobada_gerente_por',
            'aprobada_contador_por'
        ).order_by('-fecha_generacion')
        
        # Filtro de búsqueda por código
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(codigo__icontains=search) |
                Q(proyecto__nombre__icontains=search) |
                Q(observaciones__icontains=search)
            )
        
        # Filtro por proyecto
        proyecto_id = self.request.GET.get('proyecto')
        if proyecto_id and proyecto_id != 'todos':
            queryset = queryset.filter(proyecto_id=proyecto_id)
        
        # Filtro por estado
        estado = self.request.GET.get('estado')
        if estado and estado != 'todos':
            queryset = queryset.filter(estado=estado)
        
        # Filtro por rango de fechas
        fecha_desde = self.request.GET.get('fecha_desde')
        if fecha_desde:
            try:
                fecha_parsed = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                if fecha_parsed.year > 9999 or fecha_parsed.year < 1900:
                    raise ValueError('Año fuera de rango')
                queryset = queryset.filter(periodo_inicio__gte=fecha_parsed)
            except (ValueError, TypeError):
                if not getattr(self, '_fecha_desde_warned', False):
                    messages.warning(self.request, f'⚠️ Fecha desde inválida: "{fecha_desde}". Se ignoró este filtro.')
                    self._fecha_desde_warned = True
        
        fecha_hasta = self.request.GET.get('fecha_hasta')
        if fecha_hasta:
            try:
                fecha_parsed = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                if fecha_parsed.year > 9999 or fecha_parsed.year < 1900:
                    raise ValueError('Año fuera de rango')
                queryset = queryset.filter(periodo_fin__lte=fecha_parsed)
            except (ValueError, TypeError):
                if not getattr(self, '_fecha_hasta_warned', False):
                    messages.warning(self.request, f'⚠️ Fecha hasta inválida: "{fecha_hasta}". Se ignoró este filtro.')
                    self._fecha_hasta_warned = True
        
        # Filtrar por proyectos permitidos según rol
        if not self.request.user.es_administrador():
            proyectos_permitidos = self.request.user.get_proyectos_permitidos()
            queryset = queryset.filter(proyecto__in=proyectos_permitidos)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Agrega datos adicionales al contexto"""
        context = super().get_context_data(**kwargs)
        
        # Queryset base para estadísticas (filtrado por permisos)
        if self.request.user.es_administrador():
            base_queryset = Planilla.objects.filter(eliminado=False)
        else:
            proyectos_permitidos = self.request.user.get_proyectos_permitidos()
            base_queryset = Planilla.objects.filter(eliminado=False, proyecto__in=proyectos_permitidos)
        
        # Estadísticas generales
        context['total_planillas'] = base_queryset.count()
        context['planillas_borrador'] = base_queryset.filter(estado='borrador').count()
        context['planillas_aprobadas_gerente'] = base_queryset.filter(estado='aprobada_gerente').count()
        context['planillas_aprobadas_final'] = base_queryset.filter(estado='aprobada_final').count()
        context['planillas_pagadas'] = base_queryset.filter(estado='pagada').count()
        
        # Totales en córdobas
        totales = base_queryset.filter(
            estado__in=['aprobada_final', 'pagada']
        ).aggregate(
            total_cordobas=Sum('total_cordobas'),
            total_dolares=Sum('total_dolares')
        )
        context['total_monto_cordobas'] = totales['total_cordobas'] or 0
        context['total_monto_dolares'] = totales['total_dolares'] or 0
        
        # Listas para filtros (según permisos del usuario)
        context['proyectos'] = self.request.user.get_proyectos_permitidos().order_by('nombre')
        context['estados'] = Planilla.ESTADO_CHOICES
        
        # Valores actuales de filtros
        context['proyecto_actual'] = self.request.GET.get('proyecto', 'todos')
        context['estado_actual'] = self.request.GET.get('estado', 'todos')
        context['search_query'] = self.request.GET.get('search', '')
        context['fecha_desde'] = self.request.GET.get('fecha_desde', '')
        context['fecha_hasta'] = self.request.GET.get('fecha_hasta', '')
        
        # Tipo de cambio actual
        context['tipo_cambio_actual'] = TipoCambio.get_actual()
        
        return context

class PlanillaCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Vista para crear/generar una nueva planilla"""
    permission_modulo = 'planillas'
    permission_accion = 'crear'   
    template_name = 'planillas/crear.html'
    
    def get(self, request):
        """Muestra el formulario de generación"""
        
        # Obtener proyectos activos
        # Obtener proyectos según permisos del usuario
        proyectos = request.user.get_proyectos_permitidos().filter(
            estado__in=['planificacion', 'ejecucion']
        ).order_by('nombre')
        
        # Sugerir período por defecto (última catorcena)
        hoy = timezone.now().date()
        
        # Encontrar el jueves más reciente
        dias_desde_jueves = (hoy.weekday() - 3) % 7
        if dias_desde_jueves < 0:
            dias_desde_jueves += 7
        ultimo_jueves = hoy - timedelta(days=dias_desde_jueves)
        
        # Período sugerido: jueves a martes (13 días)
        periodo_inicio_sugerido = ultimo_jueves - timedelta(days=14)
        periodo_fin_sugerido = periodo_inicio_sugerido + timedelta(days=12)
        
        # Tipo de cambio actual
        tipo_cambio = TipoCambio.get_actual()
        
        context = {
            'proyectos': proyectos,
            'periodo_inicio_sugerido': periodo_inicio_sugerido,
            'periodo_fin_sugerido': periodo_fin_sugerido,
            'tipo_cambio': tipo_cambio,
        }
        
        return render(request, self.template_name, context)
    
    def post(self, request):
        """Procesa la generación de la planilla"""

        Planilla.tipo_cambio = get_tipo_cambio_actual()
        action = request.POST.get('action')
        
        print(f"DEBUG: Action recibida: {action}")
        print(f"DEBUG: POST data: {request.POST}")
        
        # ============================================================
        # ACCIÓN: PREVIEW
        # ============================================================
        if action == 'preview':
            return self.preview_planilla(request)
        
        # ============================================================
        # ACCIÓN: GENERAR
        # ============================================================
        elif action == 'generar':
            return self.generar_planilla(request)
        
        # Acción no válida
        messages.error(request, f'Acción no válida: {action}')
        return redirect('planilla_crear')
    
    def preview_planilla(self, request):
        """Muestra un preview de la planilla antes de generarla"""
        
        try:
            print("DEBUG: Iniciando preview_planilla")
            
            # Obtener datos del formulario
            proyecto_id = request.POST.get('proyecto')
            periodo_inicio_str = request.POST.get('periodo_inicio')
            periodo_fin_str = request.POST.get('periodo_fin')
            
            print(f"DEBUG: proyecto_id={proyecto_id}, inicio={periodo_inicio_str}, fin={periodo_fin_str}")
            
            # Validar datos
            if not all([proyecto_id, periodo_inicio_str, periodo_fin_str]):
                messages.error(request, 'Todos los campos son obligatorios')
                return redirect('planilla_crear')
            
            # Convertir fechas
            try:
                periodo_inicio = datetime.strptime(periodo_inicio_str, '%Y-%m-%d').date()
                periodo_fin = datetime.strptime(periodo_fin_str, '%Y-%m-%d').date()
            except ValueError as e:
                messages.error(request, f'Error en formato de fecha: {str(e)}')
                return redirect('planilla_crear')
            
            # Obtener proyecto
            try:
                proyecto = Proyecto.objects.get(id=proyecto_id, eliminado=False)
                print(f"DEBUG: Proyecto encontrado: {proyecto.nombre}")
            except Proyecto.DoesNotExist:
                messages.error(request, 'Proyecto no encontrado')
                return redirect('planilla_crear')
            
            # Bifurcación administrativa (NO requiere asistencias)
            if proyecto.is_administrativo:
                return self.preview_administrativa(request, proyecto, periodo_inicio, periodo_fin)
            
            # Validar período
            validacion = validar_periodo_planilla(periodo_inicio, periodo_fin)
            print(f"DEBUG: Validación período: {validacion}")
            
            if not validacion['valido']:
                for advertencia in validacion['advertencias']:
                    messages.error(request, advertencia)
                return redirect('planilla_crear')
            
            # Obtener resumen de asistencias
            resumen = obtener_resumen_asistencias(proyecto, periodo_inicio, periodo_fin)
            print(f"DEBUG: Resumen asistencias: {resumen}")
            
            # Verificar que haya asistencias validadas
            if not resumen['puede_generar']:
                messages.error(
                    request,
                    f'No hay asistencias validadas en el período seleccionado para el proyecto {proyecto.nombre}'
                )
                return redirect('planilla_crear')
            
            # Obtener asistencias
            asistencias = Asistencia.objects.filter(
                proyecto=proyecto,
                fecha__gte=periodo_inicio,
                fecha__lte=periodo_fin,
                validado=True
            ).select_related('trabajador')
            
            print(f"DEBUG: Asistencias encontradas: {asistencias.count()}")
            
            if not asistencias.exists():
                messages.error(request, 'No hay asistencias validadas en el período')
                return redirect('planilla_crear')
            
            # Agrupar por trabajador
            trabajadores_dict = {}
            for asistencia in asistencias:
                trabajador_id = asistencia.trabajador_id
                if trabajador_id not in trabajadores_dict:
                    trabajadores_dict[trabajador_id] = {
                        'trabajador': asistencia.trabajador,
                        'dias': 0,
                        'horas_normales': Decimal('0.00'),
                        'horas_extras': Decimal('0.00'),
                        'horas_dominicales': Decimal('0.00'),
                        'salario_hora': asistencia.trabajador.salario_normal or Decimal('0.00'),
                    }
                
                if asistencia.estado in ('cerrado', 'validado', 'editado'):
                    trabajadores_dict[trabajador_id]['dias'] += 1
                
                trabajadores_dict[trabajador_id]['horas_normales'] += asistencia.horas_normales or Decimal('0.00')
                trabajadores_dict[trabajador_id]['horas_extras'] += asistencia.horas_extras or Decimal('0.00')
                
                # Si es domingo, sumar a horas dominicales
                if asistencia.fecha.weekday() == 6:
                    trabajadores_dict[trabajador_id]['horas_dominicales'] += asistencia.horas_normales or Decimal('0.00')
            
            print(f"DEBUG: Trabajadores agrupados: {len(trabajadores_dict)}")
            
            # Calcular totales por trabajador
            preview_detalles = []
            total_general_cordobas = Decimal('0.00')
            tipo_cambio = TipoCambio.get_actual()
            
            for data in trabajadores_dict.values():
                trabajador = data['trabajador']
                # Día Base = salario_hora × 8 (fórmula Excel)
                salario_hora = trabajador.salario_normal or Decimal('0.00')
                dia_base = (salario_hora * Decimal('8')).quantize(Decimal('0.01'))
                dias = Decimal(str(data['dias']))
                
                # Fórmulas del Excel
                septimo = ((dia_base / Decimal('6')) * dias).quantize(Decimal('0.01')) if dia_base > 0 else Decimal('0.00')
                salario_base = (dias * dia_base).quantize(Decimal('0.01'))
                factor = Decimal('2.5') / Decimal('30')
                vac = (factor * salario_base).quantize(Decimal('0.01'))
                prestacionado = (salario_base + vac + vac + vac).quantize(Decimal('0.01'))
                
                tarifa_he = (dia_base / Decimal('8')) * Decimal('2') if dia_base > 0 else Decimal('0.00')
                sal_he = (tarifa_he * data['horas_extras']).quantize(Decimal('0.01'))
                sal_dom = (tarifa_he * data['horas_dominicales']).quantize(Decimal('0.01'))
                
                bonos_trabajador = trabajador.bonos or Decimal('0.00')
                ingreso_total = prestacionado + sal_he + sal_dom + bonos_trabajador
                total_general_cordobas += ingreso_total

                preview_detalles.append({
                    'trabajador': trabajador,
                    'dias': data['dias'],
                    'dia_base': dia_base,
                    'septimo_dia': septimo,
                    'salario_base': salario_base,
                    'vacaciones': vac,
                    'aguinaldo': vac,
                    'antiguedad': vac,
                    'prestacionado': prestacionado,
                    'horas_extras': data['horas_extras'],
                    'tarifa_he': tarifa_he.quantize(Decimal('0.01')),
                    'salario_horas_extras': sal_he,
                    'dias_feriados': 0,
                    'horas_feriado': Decimal('0.00'),
                    'ingreso_feriado': Decimal('0.00'),
                    'bonos': bonos_trabajador,
                    'horas_dominicales': data['horas_dominicales'],
                    'ingreso_total': ingreso_total,
                    'ingreso_total_dolares': (ingreso_total / tipo_cambio.valor).quantize(Decimal('0.01')),
                })

            total_general_dolares = (total_general_cordobas / tipo_cambio.valor).quantize(Decimal('0.01'))
            
            print(f"DEBUG: Detalles calculados: {len(preview_detalles)}")
            print(f"DEBUG: Total córdobas: {total_general_cordobas}")
            
            # Mostrar advertencias si las hay
            if validacion['advertencias']:
                for advertencia in validacion['advertencias']:
                    messages.warning(request, f"⚠️ {advertencia}")
            
            # Obtener proyectos y tipo de cambio para el contexto
            proyectos = Proyecto.objects.filter(
                eliminado=False,
                estado__in=['planificacion', 'ejecucion']
            ).order_by('nombre')
            
            context = {
                'proyectos': proyectos,
                'proyecto': proyecto,
                'periodo_inicio': periodo_inicio,
                'periodo_fin': periodo_fin,
                'validacion': validacion,
                'resumen': resumen,
                'preview_detalles': preview_detalles,
                'total_general_cordobas': total_general_cordobas,
                'total_general_dolares': total_general_dolares,
                'tipo_cambio': tipo_cambio,
                'mostrar_preview': True,
            }
            
            print("DEBUG: Renderizando template con preview")
            return render(request, self.template_name, context)
            
        except Exception as e:
            print(f"ERROR en preview_planilla: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            messages.error(request, f'Error al generar preview: {str(e)}')
            return redirect('planilla_crear')

    def preview_administrativa(self, request, proyecto, periodo_inicio, periodo_fin):
        """Preview planilla administrativa — no depende de asistencias"""
        try:
            dias_periodo = int(request.POST.get('dias_periodo', 12))
            tipo_cambio = TipoCambio.get_actual()

            trabajadores = Trabajador.objects.filter(
                proyecto_asignado=proyecto, eliminado=False, estado='activo'
            ).order_by('apellido', 'nombre')

            if not trabajadores.exists():
                messages.error(request, 'No hay trabajadores asignados a "Administración General"')
                return redirect('planilla_crear')

            preview_detalles = []
            total_general_cordobas = Decimal('0.00')

            for t in trabajadores:
                salario_hora = t.salario_normal or Decimal('0.00')
                dia_base = (salario_hora * Decimal('8')).quantize(Decimal('0.01'))
                dias = Decimal(str(dias_periodo))

                salario_base = (dias * dia_base).quantize(Decimal('0.01'))
                septimo = ((dia_base / Decimal('6')) * dias).quantize(Decimal('0.01'))
                factor = Decimal('2.5') / Decimal('30')
                vac = (factor * salario_base).quantize(Decimal('0.01'))
                prestacionado = (salario_base + vac + vac + vac).quantize(Decimal('0.01'))
                bonos_trabajador = t.bonos or Decimal('0.00')
                ingreso_total = prestacionado + bonos_trabajador
                total_general_cordobas += ingreso_total

                preview_detalles.append({
                    'trabajador': t,
                    'dias': dias_periodo,
                    'dia_base': dia_base,
                    'septimo_dia': septimo,
                    'salario_base': salario_base,
                    'vacaciones': vac,
                    'aguinaldo': vac,
                    'antiguedad': vac,
                    'prestacionado': prestacionado,
                    'horas_extras': Decimal('0.00'),
                    'tarifa_he': ((dia_base / Decimal('8')) * Decimal('2')).quantize(Decimal('0.01')) if dia_base > 0 else Decimal('0.00'),
                    'salario_horas_extras': Decimal('0.00'),
                    'dias_feriados': 0,
                    'horas_feriado': Decimal('0.00'),
                    'ingreso_feriado': Decimal('0.00'),
                    'bonos': bonos_trabajador,
                    'horas_dominicales': Decimal('0.00'),
                    'ingreso_total': ingreso_total,
                    'ingreso_total_dolares': (ingreso_total / tipo_cambio.valor).quantize(Decimal('0.01')),
                })

            total_general_dolares = (total_general_cordobas / tipo_cambio.valor).quantize(Decimal('0.01'))

            proyectos = Proyecto.objects.filter(
                eliminado=False, estado__in=['planificacion', 'ejecucion']
            ).order_by('nombre')

            context = {
                'proyectos': proyectos,
                'proyecto': proyecto,
                'periodo_inicio': periodo_inicio,
                'periodo_fin': periodo_fin,
                'dias_periodo': dias_periodo,
                'preview_detalles': preview_detalles,
                'total_general_cordobas': total_general_cordobas,
                'total_general_dolares': total_general_dolares,
                'tipo_cambio': tipo_cambio,
                'mostrar_preview': True,
                'es_administrativa': True,
            }
            return render(request, self.template_name, context)

        except Exception as e:
            messages.error(request, f'Error al generar preview administrativo: {str(e)}')
            return redirect('planilla_crear')

    def generar_planilla(self, request):
        """Genera la planilla final y la guarda en BD"""
        
        try:
            print("DEBUG: Iniciando generar_planilla")
            
            # Obtener datos del formulario
            proyecto_id = request.POST.get('proyecto')
            periodo_inicio = datetime.strptime(request.POST.get('periodo_inicio'), '%Y-%m-%d').date()
            periodo_fin = datetime.strptime(request.POST.get('periodo_fin'), '%Y-%m-%d').date()
            
            # Obtener proyecto
            proyecto = get_object_or_404(Proyecto, id=proyecto_id, eliminado=False)
            
            print(f"DEBUG: Generando planilla para {proyecto.nombre}")
            
            # Bifurcación administrativa (NO requiere asistencias)
            if proyecto.is_administrativo:
                dias_periodo = int(request.POST.get('dias_periodo', 12))
                planilla, detalles, errores = generar_planilla_administrativa(
                    proyecto=proyecto, periodo_inicio=periodo_inicio,
                    periodo_fin=periodo_fin, dias_periodo=dias_periodo, usuario=request.user
                )
                if errores:
                    for error in errores:
                        messages.error(request, f'❌ {error}')
                    if not planilla:
                        return redirect('planilla_crear')
                messages.success(request, f'✅ Planilla administrativa {planilla.codigo} generada con {len(detalles)} trabajadores.')
                return redirect('planilla_detalle', pk=planilla.pk)
            
            # ============================================================
            # ✅ VALIDACIÓN 1: Verificar si ya existe planilla
            # ============================================================
            planilla_existente = Planilla.objects.filter(
                proyecto=proyecto,
                periodo_inicio=periodo_inicio,
                periodo_fin=periodo_fin,
                eliminado=False
            ).first()
            
            if planilla_existente:
                messages.warning(
                    request,
                    f'⚠️ Ya existe una planilla para este proyecto y período: {planilla_existente.codigo}. '
                    f'Estado: {planilla_existente.get_estado_display()}. '
                    f'Si deseas crear una nueva, primero elimina la planilla existente.'
                )
                return redirect('planilla_detalle', pk=planilla_existente.pk)
            
            # ============================================================
            # ✅ VALIDACIÓN 2: Verificar que haya asistencias
            # ============================================================
            from apps.asistencias.models import Asistencia
            
            total_asistencias = Asistencia.objects.filter(
                proyecto=proyecto,
                fecha__gte=periodo_inicio,
                fecha__lte=periodo_fin
            ).count()
            
            asistencias_validadas = Asistencia.objects.filter(
                proyecto=proyecto,
                fecha__gte=periodo_inicio,
                fecha__lte=periodo_fin,
                validado=True
            ).count()
            
            if total_asistencias == 0:
                messages.error(
                    request,
                    f'❌ No hay ninguna asistencia registrada para el proyecto "{proyecto.nombre}" '
                    f'en el período del {periodo_inicio.strftime("%d/%m/%Y")} al {periodo_fin.strftime("%d/%m/%Y")}. '
                    f'Por favor, verifica las fechas o registra asistencias primero.'
                )
                return redirect('planilla_crear')
            
            if asistencias_validadas == 0:
                messages.error(
                    request,
                    f'❌ Hay {total_asistencias} asistencia(s) registrada(s) pero ninguna está validada. '
                    f'Por favor, valida las asistencias antes de generar la planilla.'
                )
                # Opcional: Redirigir a asistencias
                # return redirect('asistencias_lista')
                return redirect('planilla_crear')
            
            # ============================================================
            # ✅ VALIDACIÓN 3: Verificar que las asistencias tengan datos
            # ============================================================
            asistencias_sin_horas = Asistencia.objects.filter(
                proyecto=proyecto,
                fecha__gte=periodo_inicio,
                fecha__lte=periodo_fin,
                validado=True,
                horas_normales=0
            ).count()
            
            if asistencias_sin_horas > 0:
                messages.warning(
                    request,
                    f'⚠️ Advertencia: Hay {asistencias_sin_horas} asistencia(s) validada(s) sin horas registradas. '
                    f'Esto puede generar montos en cero para algunos trabajadores.'
                )
            
            # ============================================================
            # GENERAR PLANILLA
            # ============================================================
            # Detectar tipo de planilla
            if proyecto.is_administrativo:
                dias_periodo = int(request.POST.get('dias_periodo', 12))
                planilla, detalles, errores = generar_planilla_administrativa(
                    proyecto=proyecto,
                    periodo_inicio=periodo_inicio,
                    periodo_fin=periodo_fin,
                    dias_periodo=dias_periodo,
                    usuario=request.user
                )
            else:
                planilla, detalles, errores = generar_planilla_desde_asistencias(
                    proyecto=proyecto,
                    periodo_inicio=periodo_inicio,
                    periodo_fin=periodo_fin,
                    usuario=request.user
                )
            
            # Mostrar errores si los hay
            if errores:
                for error in errores:
                    messages.error(request, error)
                return redirect('planilla_crear')
            
            # Verificar si se generó correctamente
            if planilla:
                if len(detalles) == 0:
                    messages.warning(
                        request,
                        f'⚠️ La planilla {planilla.codigo} se creó pero no tiene trabajadores. '
                        f'Verifica que las asistencias tengan datos válidos.'
                    )
                else:
                    messages.success(
                        request,
                        f'✅ Planilla {planilla.codigo} generada exitosamente con {len(detalles)} trabajador(es). '
                        f'Total: C$ {planilla.total_cordobas:,.2f}'
                    )
                return redirect('planilla_detalle', pk=planilla.pk)
            else:
                messages.error(request, '❌ No se pudo generar la planilla. Revisa los errores anteriores.')
                return redirect('planilla_crear')
                
        except Exception as e:
            print(f"ERROR en generar_planilla: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            
            # Mensaje de error más descriptivo
            error_msg = str(e)
            
            if 'UNIQUE constraint failed' in error_msg:
                messages.error(
                    request,
                    '❌ Ya existe una planilla para este proyecto y período. '
                    'Por favor, elimina la planilla existente o cambia las fechas.'
                )
            elif 'No hay asistencias' in error_msg:
                messages.error(
                    request,
                    f'❌ {error_msg}'
                )
            else:
                messages.error(
                    request,
                    f'❌ Error al generar planilla: {error_msg}'
                )
            
            return redirect('planilla_crear')


class PlanillaDetalleView(LoginRequiredMixin, View):
    """Vista para ver el detalle completo de una planilla"""
    
    template_name = 'planillas/detalle.html'
    
    def get(self, request, pk):
        """Muestra el detalle de la planilla"""
        
        # Obtener la planilla
        planilla = get_object_or_404(
            Planilla.objects.select_related(
                'proyecto',
                'generada_por',
                'aprobada_gerente_por',
                'aprobada_contador_por'
            ),
            pk=pk,
            eliminado=False
        )
        
        # Obtener detalles de la planilla
        detalles = DetallePlanilla.objects.filter(
            planilla=planilla
        ).select_related('trabajador').order_by('area', 'trabajador__nombre', 'trabajador__apellido')
        
        # Agrupar por área
        detalles_por_area = {}
        for detalle in detalles:
            area = detalle.get_area_display()
            if area not in detalles_por_area:
                detalles_por_area[area] = []
            detalles_por_area[area].append(detalle)
        
        # Calcular totales por área
        totales_por_area = {}
        for area, detalles_area in detalles_por_area.items():
            total_cordobas = sum(d.ingreso_total for d in detalles_area)
            total_dolares = sum(d.ingreso_total_dolares for d in detalles_area)
            totales_por_area[area] = {
                'cantidad': len(detalles_area),
                'total_cordobas': total_cordobas,
                'total_dolares': total_dolares
            }
        
        # Permisos del usuario (usa rol_codigo property)
        rc = request.user.rol_codigo
        
        puede_editar = planilla.estado == 'borrador' and (
            request.user.is_superuser or 
            request.user == planilla.generada_por
        )
        
        puede_aprobar_gerente = (
            planilla.estado == 'borrador' and
            (request.user.is_superuser or rc in ['admin', 'gerente_general'])
        )
        
        puede_aprobar_contador = (
            planilla.estado == 'aprobada_gerente' and
            (request.user.is_superuser or rc in ['admin', 'contador', 'gerente_general'])
        )
        
        puede_marcar_pagada = (
            planilla.estado == 'aprobada_final' and
            (request.user.is_superuser or rc in ['admin', 'contador', 'gerente_general'])
        )
        
        context = {
            'planilla': planilla,
            'detalles': detalles,
            'detalles_por_area': detalles_por_area,
            'totales_por_area': totales_por_area,
            'puede_editar': puede_editar,
            'puede_aprobar_gerente': puede_aprobar_gerente,
            'puede_aprobar_contador': puede_aprobar_contador,
            'puede_marcar_pagada': puede_marcar_pagada,
        }
        
        return render(request, self.template_name, context)


class PlanillaEditarDetalleView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Vista para editar bonos y deducciones de un detalle de planilla"""
    permission_modulo = 'planillas'
    permission_accion = 'editar'

    def post(self, request, pk):
        """Actualiza los bonos, combustible, otros gastos y feriados de un detalle"""
        
        # Obtener objetos principales primero
        detalle = get_object_or_404(DetallePlanilla, pk=pk)
        planilla = detalle.planilla

        try:
            # 1. Verificar que la planilla esté en borrador
            if planilla.estado != 'borrador':
                messages.error(request, 'Solo se pueden editar planillas en estado borrador')
                return redirect('planilla_detalle', pk=planilla.pk)
            
            # 2. Obtener datos del formulario
            combustible_str = request.POST.get('combustible', '0')
            otros_str = request.POST.get('otros_gastos', '0')
            deducciones_str = request.POST.get('deducciones', '0')
            observaciones = request.POST.get('observaciones', '')

            # 3. Limpieza de datos (para aceptar '.' o ',')
            combustible_clean = combustible_str.strip().replace(',', '.') if combustible_str else '0'
            otros_clean = otros_str.strip().replace(',', '.') if otros_str else '0'
            deducciones_clean = deducciones_str.strip().replace(',', '.') if deducciones_str else '0'
            
            # 4. Asignar valores al detalle (bonos y feriados se calculan automáticamente)
            detalle.combustible = Decimal(combustible_clean)
            detalle.otros_gastos = Decimal(otros_clean)
            detalle.deducciones = Decimal(deducciones_clean)
            detalle.observaciones = observaciones
            
            # 5. Guardar y recalcular
            # El método .save() de DetallePlanilla
            # se encargará de llamar a calcular_valores()
            # y planilla.calcular_totales()
            detalle.save()
            
            # 6. Mensaje de éxito
            messages.success(
                request,
                f'✅ Detalle actualizado para {detalle.trabajador.nombre} {detalle.trabajador.apellido}'
            )
            
        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')
        
        # 7. Redirigir de vuelta a la planilla
        return redirect('planilla_detalle', pk=planilla.pk)


class PlanillaAprobarGerenteView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Vista para aprobar planilla como gerente"""
    permission_modulo = 'planillas'
    permission_accion = 'aprobar_gerente'

    def post(self, request, pk):
        """Aprueba la planilla como gerente"""
        
        try:
            planilla = get_object_or_404(Planilla, pk=pk, eliminado=False)
            
            # Verificar estado
            if planilla.estado != 'borrador':
                messages.error(request, 'Solo se pueden aprobar planillas en estado borrador')
                return redirect('planilla_detalle', pk=pk)
            
            # Verificar permisos
            if not (request.user.is_superuser or request.user.rol_codigo in ['admin', 'gerente_general']):
                messages.error(request, '⛔ No tienes permisos para aprobar como gerente')
                return redirect('planilla_detalle', pk=pk)
            
            # Aprobar
            planilla.estado = 'aprobada_gerente'
            planilla.aprobada_gerente_por = request.user
            planilla.aprobada_gerente_fecha = timezone.now()
            planilla.save()
            
            messages.success(
                request,
                f'✅ Planilla {planilla.codigo} aprobada como gerente. Ahora debe aprobarla el contador.'
            )
            
        except Exception as e:
            messages.error(request, f'Error al aprobar: {str(e)}')
        
        return redirect('planilla_detalle', pk=pk)


class PlanillaAprobarContadorView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Vista para aprobar planilla como contador"""
    permission_modulo = 'planillas'
    permission_accion = 'aprobar_contador'   
    
    def post(self, request, pk):
        """Aprueba la planilla como contador (aprobación final)"""
        
        try:
            planilla = get_object_or_404(Planilla, pk=pk, eliminado=False)
            
            # Verificar estado
            if planilla.estado != 'aprobada_gerente':
                messages.error(request, 'La planilla debe estar aprobada por el gerente primero')
                return redirect('planilla_detalle', pk=pk)
            
            # Verificar permisos
            if not (request.user.is_superuser or request.user.rol_codigo in ['admin', 'contador', 'gerente_general']):
                messages.error(request, '⛔ No tienes permisos para aprobar como contador')
                return redirect('planilla_detalle', pk=pk)
            
            # Aprobar
            planilla.estado = 'aprobada_final'
            planilla.aprobada_contador_por = request.user
            planilla.aprobada_contador_fecha = timezone.now()
            planilla.save()
            
            messages.success(
                request,
                f'✅ Planilla {planilla.codigo} aprobada finalmente. Lista para pago.'
            )
            
        except Exception as e:
            messages.error(request, f'Error al aprobar: {str(e)}')
        
        return redirect('planilla_detalle', pk=pk)


class PlanillaMarcarPagadaView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Vista para marcar planilla como pagada"""
    permission_modulo = 'planillas'
    permission_accion = 'aprobar_contador'

    def post(self, request, pk):
        """Marca la planilla como pagada"""
        
        try:
            planilla = get_object_or_404(Planilla, pk=pk, eliminado=False)
            
            # Verificar estado
            if planilla.estado != 'aprobada_final':
                messages.error(request, 'La planilla debe estar aprobada finalmente')
                return redirect('planilla_detalle', pk=pk)
            
            # Verificar permisos
            if not (request.user.is_superuser or request.user.rol_codigo in ['admin', 'contador', 'gerente_general']):
                messages.error(request, '⛔ No tienes permisos para marcar como pagada')
                return redirect('planilla_detalle', pk=pk)
            
            # Marcar como pagada
            planilla.estado = 'pagada'
            planilla.fecha_pago = timezone.now().date()
            planilla.save()
            
            messages.success(
                request,
                f'✅ Planilla {planilla.codigo} marcada como pagada.'
            )
            
        except Exception as e:
            messages.error(request, f'Error al marcar como pagada: {str(e)}')
        
        return redirect('planilla_detalle', pk=pk)


class PlanillaEliminarView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Vista para eliminar (soft delete) una planilla"""
    permission_modulo = 'planillas'
    permission_accion = 'eliminar'

    def post(self, request, pk):
        """Elimina la planilla (soft delete)"""
        
        try:
            planilla = get_object_or_404(Planilla, pk=pk, eliminado=False)
            
            # Solo se pueden eliminar borradores
            if planilla.estado != 'borrador':
                messages.error(request, 'Solo se pueden eliminar planillas en estado borrador')
                return redirect('planilla_detalle', pk=pk)
            
            # Verificar permisos
            if not (request.user.is_superuser or request.user == planilla.generada_por):
                messages.error(request, 'No tienes permisos para eliminar esta planilla')
                return redirect('planilla_detalle', pk=pk)
            
            # # Soft delete
            # planilla.eliminado = True
            # planilla.save()
            # Eliminar detalles y planilla definitivamente
            planilla.detalles.all().delete()
            planilla.reembolsos.all().delete()
            planilla.delete()
            
            messages.success(
                request,
                f'✅ Planilla {planilla.codigo} eliminada correctamente.'
            )
            
            return redirect('planillas_lista')
            
        except Exception as e:
            messages.error(request, f'Error al eliminar: {str(e)}')
            return redirect('planilla_detalle', pk=pk)

class PlanillaExportarExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Vista para exportar planilla a Excel"""
    permission_modulo = 'reportes'
    permission_accion = 'exportar'

    def get(self, request, pk):
            """Genera y descarga el archivo Excel con formato del Excel Quadycons"""
            
            planilla = get_object_or_404(
                Planilla.objects.select_related('proyecto', 'generada_por'),
                pk=pk,
                eliminado=False
            )
            
            detalles = DetallePlanilla.objects.filter(
                planilla=planilla
            ).select_related('trabajador').order_by('area', 'trabajador__nombre', 'trabajador__apellido')
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"Planilla {planilla.codigo}"
            
            # Estilos
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font_style = Font(name='Arial', size=9, bold=True, color='FFFFFF')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            area_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            money_format = '#,##0.00'
            
            # ============================================================
            # 1. ENCABEZADO
            # ============================================================
            
            ws.merge_cells('A1:X1')
            cell_titulo = ws['A1']
            cell_titulo.value = "PLANILLA DE PAGO"
            cell_titulo.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
            cell_titulo.fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
            cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            
            ws.merge_cells('A3:X3')
            cell_codigo = ws['A3']
            cell_codigo.value = f"{planilla.codigo} - {planilla.proyecto.nombre}"
            cell_codigo.font = Font(name='Arial', size=14, bold=True)
            cell_codigo.alignment = Alignment(horizontal='center')
            
            row = 5
            info_data = [
                ('Proyecto:', planilla.proyecto.nombre),
                ('Período:', f"{planilla.periodo_inicio.strftime('%d/%m/%Y')} - {planilla.periodo_fin.strftime('%d/%m/%Y')}"),
                ('Estado:', planilla.get_estado_display()),
                ('Tipo de Cambio:', f"C$ {planilla.tipo_cambio}"),
                ('Fecha:', planilla.fecha_generacion.strftime('%d/%m/%Y %H:%M')),
            ]
            
            for label, value in info_data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
                ws[f'B{row}'] = value
                ws[f'B{row}'].font = Font(name='Arial', size=10)
                ws.merge_cells(f'B{row}:D{row}')
                row += 1
            
            # ============================================================
            # 2. ENCABEZADOS DE TABLA (24 columnas del Excel Quadycons)
            # ============================================================
            
            row += 2
            header_row = row
            
            headers = [
                'N°', 'Nombre y Apellido', 'Cédula', 'Cargo', 'Días Lab.',
                'Día Base', '7mo Día', 'Sal. Base',
                'Vacaciones', 'Aguinaldo', 'Antigüedad', 'Prestacionado',
                'H. Extras', 'Tarifa H.E.', 'Sal. H.E.',
                'Días Fer.', 'Hrs. Fer.', 'Ingr. Feriado',
                'Bono', 'Combustible', 'Otros', 'Deducción',
                'Total C$', 'Total U$'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = header_font_style
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            ws.row_dimensions[header_row].height = 35
            
            # ============================================================
            # 3. DATOS POR ÁREA
            # ============================================================
            
            detalles_por_area = {}
            for detalle in detalles:
                area = detalle.get_area_display()
                if area not in detalles_por_area:
                    detalles_por_area[area] = []
                detalles_por_area[area].append(detalle)
            
            current_row = header_row + 1
            numero = 1
            gran_total_c = Decimal('0.00')
            gran_total_d = Decimal('0.00')
            
            for area, detalles_area in detalles_por_area.items():
                # Fila título del área
                ws.merge_cells(f'A{current_row}:X{current_row}')
                cell_area = ws.cell(row=current_row, column=1, value=f'{area.upper()} ({len(detalles_area)} trabajadores)')
                cell_area.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                cell_area.fill = area_fill
                cell_area.alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[current_row].height = 22
                current_row += 1
                
                total_area_c = Decimal('0.00')
                
                for detalle in detalles_area:
                    t = detalle.trabajador
                    db = detalle.salario_dia_base or Decimal('0')
                    tarifa_he = ((db / Decimal('8')) * Decimal('2')) if db > 0 else Decimal('0')
                    total_recibir = detalle.ingreso_total or Decimal('0')
                    total_usd = (total_recibir / planilla.tipo_cambio) if planilla.tipo_cambio > 0 else Decimal('0')
                    
                    total_area_c += total_recibir
                    gran_total_c += total_recibir
                    gran_total_d += total_usd
                    
                    data = [
                        numero,
                        t.nombre_completo,
                        t.numero_cedula or '',
                        detalle.cargo or '',
                        detalle.dias_laborados,
                        float(db),
                        float(detalle.valor_septimo_dia or 0),
                        float(detalle.salario_devengado or 0),
                        float(detalle.vacaciones or 0),
                        float(detalle.aguinaldo or 0),
                        float(detalle.antiguedad or 0),
                        float(detalle.salario_prestacionado or 0),
                        float(detalle.horas_extras or 0),
                        float(tarifa_he),
                        float(detalle.salario_horas_extras or 0),
                        detalle.dias_feriados or 0,
                        float(detalle.horas_feriado or 0),
                        float(detalle.ingreso_dia_feriado or 0),
                        float(detalle.bonos or 0),
                        float(detalle.combustible or 0),
                        float(detalle.otros_ingresos or 0),
                        float(detalle.deducciones or 0),
                        float(total_recibir),
                        float(total_usd),
                    ]
                    
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.border = border
                        cell.font = Font(name='Arial', size=9)
                        if col >= 6 and col <= 24:
                            cell.number_format = money_format
                            cell.alignment = Alignment(horizontal='right')
                    
                    numero += 1
                    current_row += 1
                
                # Subtotal del área
                ws.cell(row=current_row, column=1, value='').border = border
                ws.merge_cells(f'B{current_row}:D{current_row}')
                sub_cell = ws.cell(row=current_row, column=2, value=f'SUBTOTAL {area.upper()}')
                sub_cell.font = Font(name='Arial', size=9, bold=True)
                sub_cell.border = border
                ws.cell(row=current_row, column=23, value=float(total_area_c)).number_format = money_format
                ws.cell(row=current_row, column=23).font = Font(bold=True, size=9)
                ws.cell(row=current_row, column=23).border = border
                current_row += 1
            
            # ============================================================
            # 4. GRAN TOTAL
            # ============================================================
            
            current_row += 1
            ws.merge_cells(f'A{current_row}:V{current_row}')
            ws.cell(row=current_row, column=1, value='GRAN TOTAL').font = Font(name='Arial', size=11, bold=True)
            ws.cell(row=current_row, column=23, value=float(gran_total_c)).number_format = money_format
            ws.cell(row=current_row, column=23).font = Font(name='Arial', size=11, bold=True)
            ws.cell(row=current_row, column=24, value=float(gran_total_d)).number_format = money_format
            ws.cell(row=current_row, column=24).font = Font(name='Arial', size=11, bold=True)
            
            # ============================================================
            # 5. ANCHOS DE COLUMNA
            # ============================================================
            
            column_widths = {
                1: 5, 2: 30, 3: 16, 4: 18, 5: 8,
                6: 12, 7: 12, 8: 13,
                9: 11, 10: 11, 11: 11, 12: 14,
                13: 8, 14: 11, 15: 12,
                16: 8, 17: 8, 18: 13,
                19: 10, 20: 11, 21: 10, 22: 11,
                23: 14, 24: 13
            }
            
            for col_num, width in column_widths.items():
                ws.column_dimensions[get_column_letter(col_num)].width = width
            
            # ============================================================
            # FIRMAS: Generado por / Aprobado por
            # ============================================================
            
            current_row += 3
            
            firmas = [
                ('Generado por:', planilla.generada_por.nombre_completo if planilla.generada_por else ''),
                ('Aprobado Gerente:', planilla.aprobada_gerente_por.nombre_completo if planilla.aprobada_gerente_por else 'Pendiente'),
                ('Aprobado Contador:', planilla.aprobada_contador_por.nombre_completo if planilla.aprobada_contador_por else 'Pendiente'),
            ]
            
            col_firmas = [2, 10, 18]  # Columnas B, J, R
            
            for i, (label, nombre) in enumerate(firmas):
                col = col_firmas[i]
                # Línea de firma
                ws.cell(row=current_row, column=col, value='_' * 30).font = Font(name='Arial', size=9)
                # Nombre
                ws.cell(row=current_row + 1, column=col, value=nombre).font = Font(name='Arial', size=9, bold=True)
                ws.cell(row=current_row + 1, column=col).alignment = Alignment(horizontal='center')
                # Label
                ws.cell(row=current_row + 2, column=col, value=label).font = Font(name='Arial', size=8, color='666666')
                ws.cell(row=current_row + 2, column=col).alignment = Alignment(horizontal='center')

            # ============================================================
            # 6. GENERAR Y ENVIAR
            # ============================================================
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            filename = f"Planilla_{planilla.codigo}_{planilla.proyecto.nombre.replace(' ', '_')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response

class PlanillaExportarPDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Vista para exportar planilla a PDF"""
    permission_modulo = 'reportes'
    permission_accion = 'exportar'
        
    def get(self, request, pk):
        """Genera PDF con las 24 columnas del Excel Quadycons"""
        
        planilla = get_object_or_404(
            Planilla.objects.select_related('proyecto', 'generada_por',
                'aprobada_gerente_por', 'aprobada_contador_por'),
            pk=pk,
            eliminado=False
        )
        
        detalles = DetallePlanilla.objects.filter(
            planilla=planilla
        ).select_related('trabajador').order_by('area', 'trabajador__nombre', 'trabajador__apellido')
        
        buffer = io.BytesIO()
        from reportlab.lib.pagesizes import landscape, legal
        
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(legal),
            rightMargin=10,
            leftMargin=10,
            topMargin=12,
            bottomMargin=12,
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'],
            fontSize=13, textColor=colors.HexColor('#1F4788'),
            spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle', parent=styles['Heading2'],
            fontSize=9, textColor=colors.HexColor('#1F4788'),
            spaceAfter=4, alignment=TA_CENTER, fontName='Helvetica-Bold'
        )
        
        # ============================================================
        # 1. ENCABEZADO
        # ============================================================
        
        elements.append(Paragraph("PLANILLA DE PAGO", title_style))
        elements.append(Paragraph(
            f"{planilla.codigo} - {planilla.proyecto.nombre}", subtitle_style
        ))
        
        info_data = [
            ['Proyecto:', planilla.proyecto.nombre,
             'Período:', f"{planilla.periodo_inicio.strftime('%d/%m/%Y')} - {planilla.periodo_fin.strftime('%d/%m/%Y')}",
             'T/C:', f"C$ {planilla.tipo_cambio}"],
        ]
        info_table = Table(info_data, colWidths=[0.7*inch, 2.5*inch, 0.7*inch, 2.5*inch, 0.4*inch, 1*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTNAME', (4, 0), (4, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.1 * inch))
        
        # ============================================================
        # 2. DEFINIR COLUMNAS (24 del Excel)
        # ============================================================
        
        headers = [
            'N°', 'Nombre y Apellido', 'Cédula', 'Cargo', 'Días',
            'Día Base', '7mo Día', 'Sal.Base',
            'Vac.', 'Aguin.', 'Antig.', 'Prestac.',
            'H.E.', 'Tar.HE', 'Sal.HE',
            'D.Fer', 'H.Fer', 'Ing.Fer',
            'Bono', 'Comb.', 'Otros', 'Deduc.',
            'Total C$', 'Total U$'
        ]
        
        col_widths = [
            0.22*inch,  # N°
            1.30*inch,  # Nombre
            0.65*inch,  # Cédula
            0.55*inch,  # Cargo
            0.28*inch,  # Días
            0.50*inch,  # Día Base
            0.50*inch,  # 7mo
            0.55*inch,  # Sal.Base
            0.45*inch,  # Vac
            0.45*inch,  # Aguin
            0.45*inch,  # Antig
            0.55*inch,  # Prestac
            0.28*inch,  # H.E.
            0.42*inch,  # Tar.HE
            0.45*inch,  # Sal.HE
            0.28*inch,  # D.Fer
            0.28*inch,  # H.Fer
            0.45*inch,  # Ing.Fer
            0.38*inch,  # Bono
            0.38*inch,  # Comb
            0.38*inch,  # Otros
            0.38*inch,  # Deduc
            0.58*inch,  # Total C$
            0.55*inch,  # Total U$
        ]
        
        # ============================================================
        # 3. DATOS POR ÁREA
        # ============================================================
        
        detalles_por_area = {}
        for detalle in detalles:
            area = detalle.get_area_display()
            if area not in detalles_por_area:
                detalles_por_area[area] = []
            detalles_por_area[area].append(detalle)
        
        numero_trabajador = 1
        total_ancho = sum(col_widths)
        
        for area, detalles_area in detalles_por_area.items():
            # Título del área (mismo ancho que la tabla)
            area_row = [[f'{area.upper()} ({len(detalles_area)} trabajadores)']]
            area_table = Table(area_row, colWidths=[total_ancho])
            area_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#2F5496')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            elements.append(Spacer(1, 0.04 * inch))
            elements.append(area_table)
            
            # Tabla de datos
            table_data = [headers]
            total_area = Decimal('0.00')
            
            for detalle in detalles_area:
                db = detalle.salario_dia_base or Decimal('0')
                tarifa_he = ((db / Decimal('8')) * Decimal('2')) if db > 0 else Decimal('0')
                total_usd = (detalle.ingreso_total / planilla.tipo_cambio).quantize(Decimal('0.01')) if planilla.tipo_cambio > 0 else Decimal('0.00')
                total_area += detalle.ingreso_total
                
                row = [
                    str(numero_trabajador),
                    f"{detalle.trabajador.nombre_completo}"[:22],
                    (detalle.trabajador.numero_cedula or '')[:14],
                    (detalle.cargo or '')[:12],
                    str(detalle.dias_laborados),
                    f"{float(db):,.2f}",
                    f"{float(detalle.valor_septimo_dia or 0):,.2f}",
                    f"{float(detalle.salario_devengado or 0):,.2f}",
                    f"{float(detalle.vacaciones or 0):,.2f}",
                    f"{float(detalle.aguinaldo or 0):,.2f}",
                    f"{float(detalle.antiguedad or 0):,.2f}",
                    f"{float(detalle.salario_prestacionado or 0):,.2f}",
                    f"{float(detalle.horas_extras):.1f}",
                    f"{float(tarifa_he):,.2f}",
                    f"{float(detalle.salario_horas_extras):,.2f}",
                    str(detalle.dias_feriados or 0),
                    f"{float(detalle.horas_feriado or 0):.1f}",
                    f"{float(detalle.ingreso_dia_feriado or 0):,.2f}",
                    f"{float(detalle.bonos):,.2f}",
                    f"{float(detalle.combustible):,.2f}",
                    f"{float(detalle.otros_ingresos or 0):,.2f}",
                    f"{float(detalle.deducciones):,.2f}",
                    f"{float(detalle.ingreso_total):,.2f}",
                    f"{float(total_usd):,.2f}",
                ]
                table_data.append(row)
                numero_trabajador += 1
            
            # Subtotal
            subtotal_row = [''] * 22 + [f"{float(total_area):,.2f}", '']
            subtotal_row[1] = 'SUBTOTAL'
            table_data.append(subtotal_row)
            
            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                # Headers
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 5),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                # Datos
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 5.5),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (4, 1), (4, -1), 'CENTER'),
                ('ALIGN', (5, 1), (-1, -1), 'RIGHT'),
                ('ALIGN', (12, 1), (12, -1), 'CENTER'),
                ('ALIGN', (15, 1), (16, -1), 'CENTER'),
                # Bordes
                ('GRID', (0, 0), (-1, -2), 0.4, colors.HexColor('#D1D5DB')),
                ('LINEBELOW', (0, 0), (-1, 0), 0.8, colors.HexColor('#1F4788')),
                # Subtotal
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 6),
                ('LINEABOVE', (0, -1), (-1, -1), 0.8, colors.HexColor('#4472C4')),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EBF5FB')),
                # Zebra
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8FAFC')]),
                # Padding
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 1.5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1.5),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ]))
            
            elements.append(t)
            elements.append(Spacer(1, 0.06 * inch))
        
        # ============================================================
        # 4. TOTALES (solo Total C$ y Total U$ como el Excel)
        # ============================================================
        
        total_data = [
            ['TOTAL GENERAL:', f"C$ {float(planilla.total_cordobas):,.2f}",
             'Total U$:', f"$ {float(planilla.total_dolares):,.2f}"],
        ]
        total_table = Table(total_data, colWidths=[1.2*inch, 1.8*inch, 1.2*inch, 1.8*inch])
        total_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1F4788')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(total_table)
        elements.append(Spacer(1, 0.3 * inch))
        
        # ============================================================
        # 5. FIRMAS
        # ============================================================
        
        firmas_data = [
            ['____________________', '____________________', '____________________'],
            [
                planilla.generada_por.get_full_name() if planilla.generada_por else '',
                planilla.aprobada_gerente_por.get_full_name() if planilla.aprobada_gerente_por else 'Pendiente',
                planilla.aprobada_contador_por.get_full_name() if planilla.aprobada_contador_por else 'Pendiente',
            ],
            ['Generado por', 'Aprobado Gerente', 'Aprobado Contador'],
        ]
        firmas_table = Table(firmas_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
        firmas_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(firmas_table)
        
        # ============================================================
        # 6. CONSTRUIR Y ENVIAR
        # ============================================================
        
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        filename = f"Planilla_{planilla.codigo}_{planilla.proyecto.nombre.replace(' ', '_')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
